"""발주 컨트롤 REST API"""
import io
import json
import locale
import os
import re
import sqlite3
import subprocess
import sys
import threading
import time
from datetime import datetime
from pathlib import Path
from flask import Blueprint, request, jsonify, current_app, send_file

from src.utils.logger import get_logger
from src.settings.constants import DEFAULT_STORE_ID
from src.infrastructure.database.connection import DBRouter
from src.web.routes.api_auth import admin_required

# 입력 검증 패턴
_STORE_ID_PATTERN = re.compile(r'^[0-9]{4,6}$')
_CATEGORY_CODE_PATTERN = re.compile(r'^[0-9]{3}$')

logger = get_logger(__name__)

from src.prediction.improved_predictor import ImprovedPredictor, PredictionResult
from src.prediction.categories.default import (
    CATEGORY_NAMES,
)
from src.report.daily_order_report import DailyOrderReport
from src.infrastructure.database.repos import (
    AutoOrderItemRepository,
    SmartOrderItemRepository,
    AppSettingsRepository,
)

order_bp = Blueprint("order", __name__)

# API 응답 캐시 (TTL 기반)
import time as _time
_categories_cache = {}  # {store_id: {"data": ..., "expires": ...}}
_CATEGORIES_CACHE_TTL = 60  # 초


@order_bp.route("/stores", methods=["GET"])
def get_stores():
    """활성 점포 목록"""
    stores_path = Path(__file__).parent.parent.parent.parent / "config" / "stores.json"
    try:
        with open(stores_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        stores = [
            {"store_id": s["store_id"], "store_name": s["store_name"], "description": s.get("description", "")}
            for s in data.get("stores", []) if s.get("is_active", False)
        ]
    except Exception as e:
        logger.warning(f"점포 목록 로드 실패: {e}")
        stores = [{"store_id": DEFAULT_STORE_ID, "store_name": "기본점포", "description": ""}]
    return jsonify({"stores": stores})


@order_bp.route("/predict", methods=["POST"])
@admin_required
def predict():
    """예측 실행 (admin only)"""
    data = request.get_json(force=True) or {}
    max_items = data.get("max_items", 0)
    store_id = request.args.get('store_id', DEFAULT_STORE_ID)

    try:
        predictor = ImprovedPredictor(store_id=store_id)
        candidates = predictor.get_order_candidates(min_order_qty=0)

        # 자동발주 상품 제외 (DB 설정 반영, 매장별)
        settings_repo = AppSettingsRepository(store_id=store_id)
        if settings_repo.get("EXCLUDE_AUTO_ORDER", True):
            repo = AutoOrderItemRepository(store_id=store_id)
            auto_codes = set(repo.get_all_item_codes(store_id=store_id))
            if auto_codes:
                candidates = [c for c in candidates if c.item_cd not in auto_codes]

        # 스마트발주 상품 제외 (DB 설정 반영)
        if settings_repo.get("EXCLUDE_SMART_ORDER", True):
            smart_repo = SmartOrderItemRepository(store_id=store_id)
            smart_codes = set(smart_repo.get_all_item_codes(store_id=store_id))
            if smart_codes:
                candidates = [c for c in candidates if c.item_cd not in smart_codes]

        if max_items and max_items > 0:
            candidates = candidates[:max_items]

        # 결과 캐시 (매장별)
        current_app.config.setdefault("LAST_PREDICTIONS", {})[store_id] = candidates

        # DailyOrderReport 로직 재활용
        report = DailyOrderReport()
        summary = report._calc_summary(candidates)
        category_data = report._group_by_category(candidates)
        items = report._build_item_table(candidates)
        skipped = report._build_skipped_list(candidates)
        safety_dist = report._build_safety_distribution(candidates)

        return jsonify({
            "summary": summary,
            "category_data": category_data,
            "items": items,
            "skipped": skipped,
            "safety_dist": safety_dist,
        })
    except Exception as e:
        logger.error(f"발주 데이터 조회 실패: {e}")
        return jsonify({"error": "발주 데이터 조회에 실패했습니다"}), 500


@order_bp.route("/adjust", methods=["POST"])
def adjust():
    """발주량 수동 조정"""
    data = request.get_json(force=True)
    adjustments = data.get("adjustments", [])
    store_id = request.args.get('store_id', DEFAULT_STORE_ID)
    pred_cache = current_app.config.get("LAST_PREDICTIONS", {})
    predictions = pred_cache.get(store_id)

    if not predictions:
        return jsonify({"error": "예측 결과가 없습니다. 먼저 예측을 실행하세요."}), 400

    adj_map = {a["item_cd"]: a["order_qty"] for a in adjustments}
    count = 0
    for p in predictions:
        if p.item_cd in adj_map:
            p.order_qty = adj_map[p.item_cd]
            count += 1

    total_qty = sum(p.order_qty for p in predictions if p.order_qty > 0)
    return jsonify({
        "status": "ok",
        "adjusted_count": count,
        "new_total_qty": total_qty,
    })


@order_bp.route("/categories", methods=["GET"])
def categories():
    """카테고리 목록 (60초 캐시, store별 분리)"""
    now = _time.time()
    store_id = request.args.get('store_id') or DEFAULT_STORE_ID
    cache_key = store_id

    # store별 캐시 확인
    cached = _categories_cache.get(cache_key)
    if cached and cached["data"] is not None and now < cached["expires"]:
        return jsonify(cached["data"])

    try:
        # daily_sales는 store DB에 있음 → DBRouter 사용
        conn = DBRouter.get_store_connection(store_id)
        cursor = conn.cursor()
        store_filter_sql = "AND store_id = ?" if store_id else ""
        store_params = (store_id,) if store_id else ()
        cursor.execute(f"""
            SELECT DISTINCT mid_cd FROM daily_sales
            WHERE sales_date >= date('now', '-30 days') {store_filter_sql}
            ORDER BY mid_cd
        """, store_params)
        rows = cursor.fetchall()
        conn.close()
    except Exception as e:
        logger.warning(f"카테고리 목록 조회 실패: {e}")
        rows = []

    cats = []
    for (code,) in rows:
        cats.append({
            "code": code,
            "name": CATEGORY_NAMES.get(code, code),
        })

    data = {"categories": cats}
    _categories_cache[cache_key] = {"data": data, "expires": now + _CATEGORIES_CACHE_TTL}

    return jsonify(data)


@order_bp.route("/partial-summary", methods=["POST"])
def partial_summary():
    """선택된 카테고리의 발주 요약"""
    data = request.get_json(force=True) or {}
    cat_codes = data.get("categories", [])

    store_id = data.get("store_id") or request.args.get('store_id', DEFAULT_STORE_ID)
    pred_cache = current_app.config.get("LAST_PREDICTIONS", {})
    predictions = pred_cache.get(store_id)
    if not predictions:
        return jsonify({"error": "예측 결과가 없습니다. 먼저 예측을 실행하세요."}), 400

    if not cat_codes:
        return jsonify({"error": "카테고리를 선택하세요."}), 400

    cat_set = set(cat_codes)
    filtered = [p for p in predictions if p.mid_cd in cat_set and p.order_qty > 0]
    all_order = [p for p in predictions if p.order_qty > 0]

    return jsonify({
        "selected_categories": len(cat_set),
        "total_categories": len(set(p.mid_cd for p in all_order)),
        "item_count": len(filtered),
        "total_items": len(all_order),
        "total_qty": sum(p.order_qty for p in filtered),
        "all_qty": sum(p.order_qty for p in all_order),
    })


# === 엑셀 내보내기 ===

@order_bp.route("/export-excel", methods=["POST"])
def export_excel():
    """미리보기 결과를 Excel로 다운로드

    LAST_PREDICTIONS에 캐시된 PredictionResult 리스트를 기반으로
    예측 적용값, 재고, 입고예정 등 상세 정보가 포함된 Excel 파일을 생성합니다.
    """
    store_id = request.args.get('store_id', DEFAULT_STORE_ID)
    pred_cache = current_app.config.get("LAST_PREDICTIONS", {})
    predictions = pred_cache.get(store_id)
    if not predictions:
        return jsonify({"error": "예측 결과가 없습니다. 먼저 미리보기를 실행하세요."}), 400

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl 라이브러리가 설치되지 않았습니다."}), 500

    wb = Workbook()

    # ── Sheet 1: 발주 상세 ──
    ws = wb.active
    ws.title = "발주 상세"

    headers = [
        "No.",
        "상품코드",
        "상품명",
        "중분류",
        "최종 발주량",
        # ── 재고/입고 ──
        "현재 재고",
        "입고예정(미입고)",
        "재고+입고예정",
        # ── 예측 적용값 ──
        "기본 예측량",
        "요일계수",
        "조정 예측량",
        "안전재고",
        "일평균 판매",
        # ── 판단 근거 ──
        "신뢰도",
        "데이터일수",
        "판매일비율",
        "입수(발주단위)",
        # ── 카테고리 패턴 정보 ──
        "카테고리 패턴",
        # ── 발주 계산 과정 ──
        "필요량 산식",
    ]

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    section_fill_stock = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    section_fill_pred = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
    section_fill_meta = PatternFill(start_color="FF7030A0", end_color="FF7030A0", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 컬럼별 헤더 색상 매핑 (섹션별 구분)
    header_fills = {
        1: header_fill, 2: header_fill, 3: header_fill, 4: header_fill, 5: header_fill,
        6: section_fill_stock, 7: section_fill_stock, 8: section_fill_stock,
        9: section_fill_pred, 10: section_fill_pred, 11: section_fill_pred,
        12: section_fill_pred, 13: section_fill_pred,
        14: section_fill_meta, 15: section_fill_meta, 16: section_fill_meta,
        17: section_fill_meta,
        18: header_fill,
        19: header_fill,
    }

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fills.get(col_idx, header_fill)
        cell.alignment = header_align
        cell.border = thin_border

    # 컬럼 폭
    col_widths = [5, 18, 30, 14, 12, 10, 14, 14, 12, 10, 12, 10, 12, 8, 10, 10, 12, 20, 35]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 데이터 행
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 조건부 서식 색상
    fill_green = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    fill_yellow = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    fill_red = PatternFill(start_color="FFFCE4EC", end_color="FFFCE4EC", fill_type="solid")

    for row_idx, p in enumerate(
        sorted(predictions, key=lambda x: x.order_qty, reverse=True), 2
    ):
        cat_name = CATEGORY_NAMES.get(p.mid_cd, p.mid_cd)
        mid_display = f"{p.mid_cd} {cat_name}" if cat_name != p.mid_cd else p.mid_cd

        stock_plus_pending = p.current_stock + p.pending_qty

        # 카테고리 패턴 정보
        pattern_info = _get_pattern_info(p)

        # 발주 계산 산식
        formula_desc = _get_order_formula(p)

        row_data = [
            row_idx - 1,                                    # A: No.
            p.item_cd,                                      # B: 상품코드
            p.item_nm or p.item_cd,                         # C: 상품명
            mid_display,                                    # D: 중분류
            p.order_qty,                                    # E: 최종 발주량
            p.current_stock,                                # F: 현재 재고
            p.pending_qty,                                  # G: 입고예정(미입고)
            stock_plus_pending,                             # H: 재고+입고예정
            round(p.predicted_qty, 1),                      # I: 기본 예측량
            round(p.weekday_coef, 2),                       # J: 요일계수
            round(p.adjusted_qty, 1),                       # K: 조정 예측량
            round(p.safety_stock, 1),                       # L: 안전재고
            round(p.predicted_qty, 2) if p.predicted_qty else 0,  # M: 일평균 판매
            p.confidence,                                   # N: 신뢰도
            p.data_days,                                    # O: 데이터일수
            round(p.sell_day_ratio, 2),                     # P: 판매일비율
            getattr(p, 'order_unit_qty', 1) or 1,          # Q: 입수
            pattern_info,                                   # R: 카테고리 패턴
            formula_desc,                                   # S: 필요량 산식
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if col_idx in (1, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17):
                cell.alignment = right_align
            elif col_idx in (14,):
                cell.alignment = center_align
            elif col_idx in (18, 19):
                cell.alignment = left_align
            else:
                cell.alignment = center_align

        # ── 발주량 색상 코딩 (E열) ──
        qty_cell = ws.cell(row=row_idx, column=5)
        if p.order_qty >= 10:
            qty_cell.font = Font(bold=True, color="FFFF0000")
        elif p.order_qty >= 5:
            qty_cell.font = Font(bold=True, color="FFFF8C00")
        elif p.order_qty > 0:
            qty_cell.font = Font(bold=True)

        # ── 재고 상태 색상 (F열: 현재 재고) ──
        stock_cell = ws.cell(row=row_idx, column=6)
        if p.current_stock <= 0:
            stock_cell.fill = fill_red
            stock_cell.font = Font(bold=True, color="FFFF0000")
        elif p.current_stock <= 2:
            stock_cell.fill = fill_yellow

        # ── 신뢰도 색상 (N열) ──
        conf_cell = ws.cell(row=row_idx, column=14)
        if p.confidence == "high":
            conf_cell.fill = fill_green
        elif p.confidence == "low":
            conf_cell.fill = fill_red

    # 필터 + 고정
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(predictions) + 1}"
    ws.freeze_panes = "A2"

    # ── Sheet 2: 요약 ──
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="발주 미리보기 요약").font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")

    row = 3
    ordered = [p for p in predictions if p.order_qty > 0]
    skipped = [p for p in predictions if p.order_qty <= 0]

    meta_items = [
        ("생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("총 상품 수", len(predictions)),
        ("발주 대상", len(ordered)),
        ("스킵 상품", len(skipped)),
        ("총 발주량", sum(p.order_qty for p in ordered)),
        ("카테고리 수", len(set(p.mid_cd for p in predictions))),
    ]
    for label, value in meta_items:
        ws2.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=value)
        row += 1

    # 중분류별 통계
    row += 1
    ws2.cell(row=row, column=1, value="중분류별 통계").font = Font(bold=True, size=12)
    row += 1

    cat_headers = ["중분류", "상품 수", "총 발주량", "평균 재고", "평균 미입고"]
    cat_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = cat_fill
    row += 1

    from collections import defaultdict
    cat_stats = defaultdict(lambda: {"count": 0, "qty": 0, "stock": 0, "pending": 0})
    for p in predictions:
        cat_name = CATEGORY_NAMES.get(p.mid_cd, p.mid_cd)
        key = f"{p.mid_cd} {cat_name}" if cat_name != p.mid_cd else p.mid_cd
        cat_stats[key]["count"] += 1
        cat_stats[key]["qty"] += p.order_qty
        cat_stats[key]["stock"] += p.current_stock
        cat_stats[key]["pending"] += p.pending_qty

    for key in sorted(cat_stats.keys()):
        stats = cat_stats[key]
        cnt = stats["count"] or 1
        ws2.cell(row=row, column=1, value=key)
        ws2.cell(row=row, column=2, value=stats["count"])
        ws2.cell(row=row, column=3, value=stats["qty"])
        ws2.cell(row=row, column=4, value=round(stats["stock"] / cnt, 1))
        ws2.cell(row=row, column=5, value=round(stats["pending"] / cnt, 1))
        row += 1

    # 신뢰도 분포
    row += 1
    ws2.cell(row=row, column=1, value="신뢰도 분포").font = Font(bold=True, size=12)
    row += 1
    conf_counts = defaultdict(int)
    for p in predictions:
        conf_counts[p.confidence] += 1
    for conf in ["high", "medium", "low"]:
        ws2.cell(row=row, column=1, value=conf)
        ws2.cell(row=row, column=2, value=conf_counts.get(conf, 0))
        row += 1

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 12

    # Excel → 메모리 버퍼
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"order_preview_{timestamp}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _get_pattern_info(p) -> str:
    """PredictionResult에서 카테고리별 패턴 정보를 문자열로 추출"""
    parts = []

    # 담배
    if p.tobacco_skip_order:
        parts.append(f"담배: 스킵({p.tobacco_skip_reason})")
    elif p.carton_buffer > 0 or p.carton_frequency > 0:
        parts.append(f"담배: 보루버퍼={p.carton_buffer}, 보루빈도={p.carton_frequency:.1f}")
        if p.tobacco_available_space > 0:
            parts.append(f"여유={p.tobacco_available_space}")

    # 라면
    if p.ramen_skip_order:
        parts.append("라면: 상한초과 스킵")
    elif p.ramen_turnover_level:
        parts.append(f"라면: 회전={p.ramen_turnover_level}, 안전일수={p.ramen_safety_days:.1f}")

    # 맥주
    if p.beer_skip_order:
        parts.append(f"맥주: 스킵({p.beer_skip_reason})")
    elif p.beer_weekday_coef is not None:
        parts.append(f"맥주: 요일계수={p.beer_weekday_coef:.2f}, 안전일수={p.beer_safety_days}")

    # 소주
    if p.soju_skip_order:
        parts.append(f"소주: 스킵({p.soju_skip_reason})")
    elif p.soju_weekday_coef is not None:
        parts.append(f"소주: 요일계수={p.soju_weekday_coef:.2f}, 안전일수={p.soju_safety_days}")

    # 푸드류
    if p.food_expiration_days is not None:
        parts.append(
            f"푸드: 유통기한={p.food_expiration_days}일, "
            f"그룹={p.food_expiry_group}, "
            f"폐기계수={p.food_disuse_coef:.2f}"
        )

    # 간헐적 수요
    if p.intermittent_adjusted:
        parts.append(f"간헐적수요보정(판매일비율={p.sell_day_ratio:.2f})")

    return " | ".join(parts) if parts else "-"


def _get_order_formula(p) -> str:
    """발주량 계산 과정을 읽기 쉬운 산식으로 표현"""
    stock_total = p.current_stock + p.pending_qty
    need = round(p.safety_stock + p.adjusted_qty - stock_total, 1)
    return (
        f"안전재고({p.safety_stock:.1f}) + "
        f"예측({p.adjusted_qty:.1f}) - "
        f"(재고({p.current_stock}) + 미입고({p.pending_qty})) "
        f"= {need} → 발주 {p.order_qty}"
    )


# === 스크립트 실행 관련 ===

# 프로젝트 루트 (bgf_auto/) 기준 스크립트 경로 매핑
_SCRIPT_MAP = {
    "preview":    {"script": "scripts/run_auto_order.py", "args": ["--preview"]},
    "dry-report": {"script": "scripts/dry_order.py",      "args": ["--no-login"]},
    "dry-run":    {"script": "scripts/run_auto_order.py",  "args": []},
    "full-test":  {"script": "scripts/run_full_flow.py",   "args": ["--no-collect"]},
    "real-order": {"script": "scripts/run_auto_order.py",  "args": ["--run"]},
}


def _read_output(pipe, output_list):
    """서브프로세스 stdout/stderr를 라인 단위로 읽어 리스트에 추가"""
    try:
        for line in iter(pipe.readline, ""):
            if line:
                output_list.append(line)
        pipe.close()
    except Exception as e:
        logger.warning(f"발주 데이터 조회 실패: {e}")


@order_bp.route("/run-script", methods=["POST"])
@admin_required
def run_script():
    """스크립트 백그라운드 실행 (admin only)"""
    # 이미 실행 중인 작업이 있는지 확인
    task = current_app.config.get("SCRIPT_TASK")
    if task and task.get("process") and task["process"].poll() is None:
        return jsonify({"error": "이미 실행 중인 스크립트가 있습니다."}), 409

    data = request.get_json(force=True) or {}
    mode = data.get("mode", "")
    max_items = data.get("max_items", 0)
    min_qty = data.get("min_qty", 0)

    if mode not in _SCRIPT_MAP:
        return jsonify({"error": f"알 수 없는 모드: {mode}"}), 400

    cfg = _SCRIPT_MAP[mode]
    # bgf_auto/ 디렉토리 기준 경로
    project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(
        os.path.abspath(__file__)
    ))))
    script_path = os.path.join(project_root, cfg["script"])

    cmd = [sys.executable, script_path] + list(cfg["args"])

    # store_id 인자 추가 (입력 검증)
    store_id = data.get("store_id", DEFAULT_STORE_ID)
    if not _STORE_ID_PATTERN.match(str(store_id)):
        return jsonify({"error": "유효하지 않은 점포 코드입니다"}), 400
    cmd += ["--store-id", str(store_id)]

    # categories 인자 추가 (부분 발주, 입력 검증)
    cat_codes = data.get("categories")
    has_categories = cat_codes and isinstance(cat_codes, list) and len(cat_codes) > 0
    if has_categories:
        for code in cat_codes:
            if not _CATEGORY_CODE_PATTERN.match(str(code)):
                return jsonify({"error": f"유효하지 않은 카테고리 코드: {code}"}), 400
        cmd += ["--categories", ",".join(str(c) for c in cat_codes)]

    # max_items, min_qty 인자 추가
    # 부분 발주 시에는 max_items를 전달하지 않음 (카테고리 필터가 대신 제한)
    if not has_categories and mode not in ("preview",) and max_items and int(max_items) > 0:
        cmd += ["--max-items", str(int(max_items))]
    if mode not in ("preview", "dry-report") and min_qty and int(min_qty) > 0:
        cmd += ["--min-qty", str(int(min_qty))]

    output_lines = []
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            cwd=project_root,
            encoding=locale.getpreferredencoding(False),
            errors="replace",
        )
    except Exception as e:
        logger.error(f"스크립트 실행 실패: {e}")
        return jsonify({"error": "스크립트 실행에 실패했습니다"}), 500

    # stdout 읽기 스레드
    reader = threading.Thread(target=_read_output, args=(proc.stdout, output_lines), daemon=True)
    reader.start()

    current_app.config["SCRIPT_TASK"] = {
        "process": proc,
        "output": output_lines,
        "started_at": time.time(),
        "mode": mode,
        "reader": reader,
    }

    return jsonify({"status": "started", "mode": mode, "pid": proc.pid})


@order_bp.route("/script-status", methods=["GET"])
def script_status():
    """실행 상태/로그 조회"""
    task = current_app.config.get("SCRIPT_TASK")
    if not task:
        return jsonify({"running": False, "mode": None, "output": "", "exit_code": None, "elapsed": 0})

    proc = task["process"]
    running = proc.poll() is None
    elapsed = round(time.time() - task["started_at"], 1)
    output_text = "".join(task["output"])

    return jsonify({
        "running": running,
        "mode": task["mode"],
        "output": output_text,
        "exit_code": proc.returncode,
        "elapsed": elapsed,
    })


@order_bp.route("/stop-script", methods=["POST"])
def stop_script():
    """실행 중인 스크립트 중단"""
    task = current_app.config.get("SCRIPT_TASK")
    if not task or not task.get("process"):
        return jsonify({"error": "실행 중인 스크립트가 없습니다."}), 404

    proc = task["process"]
    if proc.poll() is None:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
        task["output"].append("\n[사용자에 의해 중단됨]\n")

    return jsonify({"status": "stopped"})


@order_bp.route("/exclusions", methods=["GET"])
def get_exclusions():
    """발주 제외 목록 조회 (점포별)"""
    store_id = request.args.get('store_id', DEFAULT_STORE_ID)
    settings_repo = AppSettingsRepository(store_id=store_id)
    auto_repo = AutoOrderItemRepository(store_id=store_id)
    smart_repo = SmartOrderItemRepository(store_id=store_id)
    return jsonify({
        "store_id": store_id,
        "auto_order": {
            "enabled": settings_repo.get("EXCLUDE_AUTO_ORDER", True),
            "count": auto_repo.get_count(store_id=store_id),
            "last_updated": auto_repo.get_last_updated(store_id=store_id),
            "items": auto_repo.get_all_detail(store_id=store_id),
        },
        "smart_order": {
            "enabled": settings_repo.get("EXCLUDE_SMART_ORDER", True),
            "count": smart_repo.get_count(store_id=store_id),
            "last_updated": smart_repo.get_last_updated(store_id=store_id),
            "items": smart_repo.get_all_detail(store_id=store_id),
        }
    })


@order_bp.route("/exclusions/toggle", methods=["POST"])
def toggle_exclusion():
    """발주 제외 설정 토글 (매장별 DB 영속)"""
    data = request.get_json(force=True) or {}
    kind = data.get("kind", "auto")
    enabled = data.get("enabled", True)
    store_id = data.get("store_id", request.args.get('store_id', DEFAULT_STORE_ID))

    settings_repo = AppSettingsRepository(store_id=store_id)
    if kind == "smart":
        settings_repo.set("EXCLUDE_SMART_ORDER", enabled)
    else:
        settings_repo.set("EXCLUDE_AUTO_ORDER", enabled)

    return jsonify({"status": "ok", "kind": kind, "enabled": enabled, "store_id": store_id})
