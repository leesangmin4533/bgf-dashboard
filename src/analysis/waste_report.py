"""
폐기 보고서 엑셀 생성기

날짜별 폐기 보고서를 엑셀로 출력:
  시트1: 일별 폐기 상세 (상품코드, 상품명, 카테고리, 입고일, 폐기일, 폐기수량)
  시트2: 카테고리 집계 (카테고리별 폐기건수, 폐기수량 합계)
  시트3: 주간 트렌드 (최근 4주간 일별 폐기수량 추이 차트)
  시트4: 월간 트렌드 (최근 3개월 주별 폐기율 추이, TOP10)

출력 경로: data/expiry_reports/YYYY-MM-DD_폐기보고서.xlsx
"""

import sys
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional

sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

from src.infrastructure.database.repos import (
    InventoryBatchRepository,
    OrderTrackingRepository,
    ReceivingRepository,
)
from src.infrastructure.database.connection import get_connection, DBRouter
from src.utils.logger import get_logger

logger = get_logger(__name__)

# 출력 디렉토리
REPORT_DIR = Path(__file__).parent.parent.parent / "data" / "expiry_reports"

# 스타일 정의
HEADER_FONT = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14)
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


class WasteReportGenerator:
    """폐기 보고서 엑셀 생성기"""

    def __init__(self, store_id: Optional[str] = None) -> None:
        self.store_id = store_id
        self.batch_repo = InventoryBatchRepository(store_id=self.store_id)
        self.tracking_repo = OrderTrackingRepository(store_id=self.store_id)
        self.receiving_repo = ReceivingRepository(store_id=self.store_id)

    def _get_store_conn(self):
        """store DB 커넥션 (order_tracking, inventory_batches 등 store 테이블용)"""
        if self.store_id:
            return DBRouter.get_store_connection_with_common(self.store_id)
        return get_connection()

    def generate(self, target_date: Optional[str] = None) -> Path:
        """폐기 보고서 생성

        Args:
            target_date: 기준 날짜 (기본: 오늘, YYYY-MM-DD)

        Returns:
            생성된 엑셀 파일 경로
        """
        if target_date is None:
            target_date = datetime.now().strftime("%Y-%m-%d")

        REPORT_DIR.mkdir(parents=True, exist_ok=True)
        output_path = REPORT_DIR / f"{target_date}_폐기보고서.xlsx"

        wb = Workbook()

        # 시트 생성
        self._create_daily_detail_sheet(wb, target_date)
        self._create_category_summary_sheet(wb, target_date)
        self._create_weekly_trend_sheet(wb, target_date)
        self._create_monthly_trend_sheet(wb, target_date)

        # 기본 시트 삭제 (Workbook 생성 시 자동으로 만들어진 빈 시트)
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(output_path))
        logger.info(f"폐기 보고서 생성: {output_path}")
        return output_path

    def _apply_header_style(self, ws, row: int, cols: int) -> None:
        """헤더 행 스타일 적용"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    def _apply_data_style(self, ws, row: int, cols: int) -> None:
        """데이터 행 스타일 적용"""
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

    def _auto_column_width(self, ws) -> None:
        """컬럼 너비 자동 조정"""
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    # =========================================================================
    # 시트 1: 일별 폐기 상세
    # =========================================================================
    def _build_receiving_map(self, target_date: str) -> Dict[str, int]:
        """입고일 기준 receiving_history에서 상품별 실제 입고 수량 조회

        Args:
            target_date: 입고일 (YYYY-MM-DD)

        Returns:
            {item_cd: receiving_qty} 딕셔너리
        """
        recv_map: Dict[str, int] = {}
        try:
            recv_list = self.receiving_repo.get_receiving_by_date(
                target_date, store_id=self.store_id
            )
            for r in recv_list:
                item_cd = r.get('item_cd', '')
                qty = r.get('receiving_qty', 0) or 0
                recv_map[item_cd] = recv_map.get(item_cd, 0) + qty
        except Exception as e:
            logger.debug(f"입고 내역 조회 실패 ({target_date}): {e}")
        return recv_map

    def _create_daily_detail_sheet(self, wb: Workbook, target_date: str) -> None:
        """일별 폐기 상세 시트"""
        ws = wb.create_sheet("일별 폐기 상세", 0)

        # 제목
        ws.cell(row=1, column=1, value=f"폐기 상세 ({target_date})")
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells("A1:I1")

        # 헤더 (실제입고수량 컬럼 추가)
        headers = ["상품코드", "상품명", "카테고리", "입고일", "폐기예정일",
                    "입고수량", "실제입고수량", "폐기수량", "폐기율(%)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        # 데이터: 비-푸드 배치
        batch_data = self.batch_repo.get_expired_batches(target_date, days_back=1, store_id=self.store_id)

        # 데이터: 푸드류 (order_tracking에서 expired 상태)
        food_data = self._get_food_waste_for_date(target_date)

        all_data = food_data + batch_data

        # receiving_history에서 실제 입고 수량 맵 구축
        recv_dates = set()
        for item in all_data:
            rd = item.get('receiving_date', item.get('order_date', ''))
            if rd:
                recv_dates.add(rd)
        recv_map: Dict[str, int] = {}
        for rd in recv_dates:
            recv_map.update(self._build_receiving_map(rd))

        row = 4
        for item in all_data:
            item_cd = item.get('item_cd', '')
            ws.cell(row=row, column=1, value=item_cd)
            ws.cell(row=row, column=2, value=item.get('item_nm', ''))
            ws.cell(row=row, column=3, value=item.get('mid_cd', ''))
            ws.cell(row=row, column=4, value=item.get('receiving_date', item.get('order_date', '')))
            ws.cell(row=row, column=5, value=item.get('expiry_date', item.get('expiry_time', '')))
            ws.cell(row=row, column=6, value=item.get('initial_qty', item.get('order_qty', 0)))

            # 실제 입고수량 (receiving_history 기반)
            actual_recv = recv_map.get(item_cd, '')
            ws.cell(row=row, column=7, value=actual_recv if actual_recv else '')

            waste_qty = item.get('remaining_qty', 0)
            ws.cell(row=row, column=8, value=waste_qty)

            # 폐기율: 실제 입고수량 우선, 없으면 배치 initial_qty 사용
            base_qty = actual_recv if actual_recv else item.get('initial_qty', item.get('order_qty', 1))
            waste_rate = round(waste_qty / base_qty * 100, 1) if base_qty and base_qty > 0 else 0
            ws.cell(row=row, column=9, value=waste_rate)

            self._apply_data_style(ws, row, len(headers))

            # 폐기율 높으면 하이라이트
            if waste_rate >= 50:
                ws.cell(row=row, column=9).fill = RED_FILL
            elif waste_rate >= 30:
                ws.cell(row=row, column=9).fill = YELLOW_FILL

            row += 1

        # 합계
        if all_data:
            ws.cell(row=row + 1, column=5, value="합계")
            ws.cell(row=row + 1, column=5).font = Font(bold=True)
            ws.cell(row=row + 1, column=8, value=sum(
                d.get('remaining_qty', 0) for d in all_data
            ))
            ws.cell(row=row + 1, column=8).font = Font(bold=True)

        self._auto_column_width(ws)

    def _get_food_waste_for_date(self, target_date: str) -> List[Dict[str, Any]]:
        """푸드류 폐기 데이터 조회 (order_tracking 기반)"""
        conn = self._get_store_conn()
        try:
            store_filter = "AND store_id = ?" if self.store_id else ""
            store_params = (self.store_id,) if self.store_id else ()
            cursor = conn.cursor()
            cursor.execute(
                f"""
                SELECT item_cd, item_nm, mid_cd, order_date,
                       expiry_time, order_qty, remaining_qty, delivery_type
                FROM order_tracking
                WHERE status IN ('expired', 'disposed')
                AND date(expiry_time) = ?
                AND remaining_qty > 0
                {store_filter}
                ORDER BY mid_cd, item_nm
                """,
                (target_date,) + store_params
            )
            return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"푸드 폐기 데이터 조회 실패 | store_id={self.store_id} | date={target_date}: {e}")
            return []
        finally:
            conn.close()

    # =========================================================================
    # 시트 2: 카테고리 집계
    # =========================================================================
    def _create_category_summary_sheet(self, wb: Workbook, target_date: str) -> None:
        """카테고리별 폐기 집계 시트"""
        ws = wb.create_sheet("카테고리 집계")

        ws.cell(row=1, column=1, value=f"카테고리별 폐기 집계 ({target_date})")
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells("A1:E1")

        headers = ["카테고리(중분류)", "폐기 건수", "폐기 수량", "입고 수량", "폐기율(%)"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        # 배치 기반 통계 (최근 7일)
        batch_summary = self.batch_repo.get_waste_summary(days_back=7, store_id=self.store_id)

        # 푸드류 통계
        food_summary = self._get_food_waste_summary(days_back=7)

        all_summary = food_summary + batch_summary
        row = 4
        total_items = 0
        total_waste = 0
        total_initial = 0

        for item in all_summary:
            mid_cd = item.get('mid_cd', '')
            items_count = item.get('total_items', 0)
            waste_qty = item.get('total_waste_qty', 0)
            initial_qty = item.get('total_initial_qty', 0)
            waste_rate = item.get('waste_rate_pct', 0)

            ws.cell(row=row, column=1, value=mid_cd)
            ws.cell(row=row, column=2, value=items_count)
            ws.cell(row=row, column=3, value=waste_qty)
            ws.cell(row=row, column=4, value=initial_qty)
            ws.cell(row=row, column=5, value=waste_rate)
            self._apply_data_style(ws, row, len(headers))

            total_items += items_count
            total_waste += waste_qty
            total_initial += initial_qty
            row += 1

        # 합계 행
        ws.cell(row=row, column=1, value="합계")
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_items)
        ws.cell(row=row, column=3, value=total_waste)
        ws.cell(row=row, column=4, value=total_initial)
        if total_initial > 0:
            ws.cell(row=row, column=5, value=round(total_waste / total_initial * 100, 1))
        self._apply_data_style(ws, row, len(headers))

        self._auto_column_width(ws)

    def _get_food_waste_summary(self, days_back: int = 7) -> List[Dict[str, Any]]:
        """푸드류 폐기 통계 (order_tracking 기반)"""
        conn = self._get_store_conn()
        try:
            store_filter = "AND store_id = ?" if self.store_id else ""
            store_params = (self.store_id,) if self.store_id else ()
            cursor = conn.cursor()
            cursor.execute(
                f"""
                SELECT
                    mid_cd,
                    COUNT(*) as total_items,
                    SUM(remaining_qty) as total_waste_qty,
                    SUM(order_qty) as total_initial_qty,
                    ROUND(
                        CAST(SUM(remaining_qty) AS REAL) /
                        NULLIF(SUM(order_qty), 0) * 100, 1
                    ) as waste_rate_pct
                FROM order_tracking
                WHERE status IN ('expired', 'disposed')
                AND date(expiry_time) >= date('now', '-' || ? || ' days')
                AND remaining_qty > 0
                {store_filter}
                GROUP BY mid_cd
                ORDER BY total_waste_qty DESC
                """,
                (days_back,) + store_params
            )
            return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"푸드 폐기 통계 조회 실패 | store_id={self.store_id} | days_back={days_back}: {e}")
            return []
        finally:
            conn.close()

    # =========================================================================
    # 시트 3: 주간 트렌드 (최근 4주)
    # =========================================================================
    def _create_weekly_trend_sheet(self, wb: Workbook, target_date: str) -> None:
        """주간 폐기 트렌드 시트 (차트 포함)"""
        ws = wb.create_sheet("주간 트렌드")

        ws.cell(row=1, column=1, value="주간 폐기 추이 (최근 28일)")
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells("A1:C1")

        headers = ["날짜", "폐기 건수", "폐기 수량"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        # 배치 일별 트렌드
        batch_trend = self.batch_repo.get_daily_waste_trend(days_back=28, store_id=self.store_id)

        # 푸드류 일별 트렌드
        food_trend = self._get_food_daily_waste_trend(days_back=28)

        # 날짜별 병합
        merged = {}
        for item in batch_trend + food_trend:
            date = item.get('expiry_date', '')
            if date not in merged:
                merged[date] = {'waste_items': 0, 'waste_qty': 0}
            merged[date]['waste_items'] += item.get('waste_items', 0)
            merged[date]['waste_qty'] += item.get('waste_qty', 0)

        row = 4
        for date in sorted(merged.keys()):
            data = merged[date]
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=data['waste_items'])
            ws.cell(row=row, column=3, value=data['waste_qty'])
            self._apply_data_style(ws, row, len(headers))
            row += 1

        # 차트 생성 (데이터가 있을 때만)
        if len(merged) >= 2:
            chart = BarChart()
            chart.type = "col"
            chart.title = "일별 폐기 수량 추이"
            chart.y_axis.title = "수량"
            chart.x_axis.title = "날짜"
            chart.style = 10

            data_ref = Reference(ws, min_col=3, min_row=3, max_row=row - 1)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 20
            chart.height = 12

            ws.add_chart(chart, "E3")

        self._auto_column_width(ws)

    def _get_food_daily_waste_trend(self, days_back: int = 28) -> List[Dict[str, Any]]:
        """푸드류 일별 폐기 트렌드"""
        conn = self._get_store_conn()
        try:
            store_filter = "AND store_id = ?" if self.store_id else ""
            store_params = (self.store_id,) if self.store_id else ()
            cursor = conn.cursor()
            cursor.execute(
                f"""
                SELECT
                    date(expiry_time) as expiry_date,
                    COUNT(*) as waste_items,
                    SUM(remaining_qty) as waste_qty
                FROM order_tracking
                WHERE status IN ('expired', 'disposed')
                AND date(expiry_time) >= date('now', '-' || ? || ' days')
                AND remaining_qty > 0
                {store_filter}
                GROUP BY date(expiry_time)
                ORDER BY expiry_date
                """,
                (days_back,) + store_params
            )
            return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"푸드 일별 폐기 트렌드 조회 실패 | store_id={self.store_id} | days_back={days_back}: {e}")
            return []
        finally:
            conn.close()

    # =========================================================================
    # 시트 4: 월간 트렌드 + TOP10
    # =========================================================================
    def _create_monthly_trend_sheet(self, wb: Workbook, target_date: str) -> None:
        """월간 폐기 트렌드 + 폐기율 TOP10 시트"""
        ws = wb.create_sheet("월간 트렌드")

        # === 주별 폐기 추이 ===
        ws.cell(row=1, column=1, value="주별 폐기 추이 (최근 12주)")
        ws.cell(row=1, column=1).font = TITLE_FONT
        ws.merge_cells("A1:D1")

        headers = ["주차", "시작일", "종료일", "폐기 수량"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=header)
        self._apply_header_style(ws, 3, len(headers))

        weekly_data = self._get_weekly_waste_trend(weeks=12)
        row = 4
        for idx, week in enumerate(weekly_data, 1):
            ws.cell(row=row, column=1, value=f"W{idx}")
            ws.cell(row=row, column=2, value=week.get('week_start', ''))
            ws.cell(row=row, column=3, value=week.get('week_end', ''))
            ws.cell(row=row, column=4, value=week.get('waste_qty', 0))
            self._apply_data_style(ws, row, len(headers))
            row += 1

        # 주별 차트
        if len(weekly_data) >= 2:
            chart = LineChart()
            chart.title = "주별 폐기 수량 추이"
            chart.y_axis.title = "수량"
            chart.style = 10

            data_ref = Reference(ws, min_col=4, min_row=3, max_row=row - 1)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 18
            chart.height = 10

            ws.add_chart(chart, "F3")

        # === 폐기율 TOP10 ===
        top10_start = row + 2
        ws.cell(row=top10_start, column=1, value="상품별 폐기율 TOP10 (최근 30일)")
        ws.cell(row=top10_start, column=1).font = TITLE_FONT
        ws.merge_cells(f"A{top10_start}:F{top10_start}")

        headers2 = ["순위", "상품코드", "상품명", "입고수량", "폐기수량", "폐기율(%)"]
        for col, header in enumerate(headers2, 1):
            ws.cell(row=top10_start + 2, column=col, value=header)
        self._apply_header_style(ws, top10_start + 2, len(headers2))

        top10 = self._get_top_waste_items(limit=10, days_back=30)
        row2 = top10_start + 3
        for idx, item in enumerate(top10, 1):
            ws.cell(row=row2, column=1, value=idx)
            ws.cell(row=row2, column=2, value=item.get('item_cd', ''))
            ws.cell(row=row2, column=3, value=item.get('item_nm', ''))
            ws.cell(row=row2, column=4, value=item.get('total_initial', 0))
            ws.cell(row=row2, column=5, value=item.get('total_waste', 0))
            ws.cell(row=row2, column=6, value=item.get('waste_rate', 0))
            self._apply_data_style(ws, row2, len(headers2))

            if item.get('waste_rate', 0) >= 50:
                ws.cell(row=row2, column=6).fill = RED_FILL
            row2 += 1

        self._auto_column_width(ws)

    def _get_weekly_waste_trend(self, weeks: int = 12) -> List[Dict[str, Any]]:
        """주별 폐기 추이"""
        conn = self._get_store_conn()
        try:
            batch_store_filter = "AND store_id = ?" if self.store_id else ""
            ot_store_filter = "AND store_id = ?" if self.store_id else ""
            store_params = (self.store_id,) if self.store_id else ()
            cursor = conn.cursor()
            result = []
            today = datetime.now().date()

            for w in range(weeks - 1, -1, -1):
                week_end = today - timedelta(days=w * 7)
                week_start = week_end - timedelta(days=6)

                # 배치 폐기
                cursor.execute(
                    f"""
                    SELECT COALESCE(SUM(remaining_qty), 0) as qty
                    FROM inventory_batches
                    WHERE status = 'expired'
                    AND expiry_date >= ? AND expiry_date <= ?
                    {batch_store_filter}
                    """,
                    (week_start.isoformat(), week_end.isoformat()) + store_params
                )
                batch_qty = cursor.fetchone()['qty']

                # 푸드 폐기
                cursor.execute(
                    f"""
                    SELECT COALESCE(SUM(remaining_qty), 0) as qty
                    FROM order_tracking
                    WHERE status IN ('expired', 'disposed')
                    AND date(expiry_time) >= ? AND date(expiry_time) <= ?
                    AND remaining_qty > 0
                    {ot_store_filter}
                    """,
                    (week_start.isoformat(), week_end.isoformat()) + store_params
                )
                food_qty = cursor.fetchone()['qty']

                result.append({
                    'week_start': week_start.isoformat(),
                    'week_end': week_end.isoformat(),
                    'waste_qty': batch_qty + food_qty,
                })

            return result
        except Exception as e:
            logger.warning(f"주별 폐기 추이 조회 실패 | store_id={self.store_id} | weeks={weeks}: {e}")
            return []
        finally:
            conn.close()

    def _get_top_waste_items(self, limit: int = 10, days_back: int = 30) -> List[Dict[str, Any]]:
        """폐기율 TOP N 상품

        receiving_history와 LEFT JOIN하여 실제 입고수량 대비 폐기율을 산출한다.
        receiving_history에 레코드가 없으면 배치/발주 수량을 기준으로 폴백.
        """
        conn = self._get_store_conn()
        try:
            batch_sf = "AND ib.store_id = ?" if self.store_id else ""
            ot_sf = "AND ot.store_id = ?" if self.store_id else ""
            rh_sf = "AND rh.store_id = ?" if self.store_id else ""
            sp = (self.store_id,) if self.store_id else ()
            cursor = conn.cursor()

            cursor.execute(
                f"""
                SELECT
                    ib.item_cd,
                    ib.item_nm,
                    SUM(ib.initial_qty) as total_initial,
                    COALESCE(SUM(rh.receiving_qty), SUM(ib.initial_qty)) as total_receiving,
                    SUM(ib.remaining_qty) as total_waste,
                    ROUND(
                        CAST(SUM(ib.remaining_qty) AS REAL) /
                        NULLIF(COALESCE(SUM(rh.receiving_qty), SUM(ib.initial_qty)), 0) * 100, 1
                    ) as waste_rate
                FROM inventory_batches ib
                LEFT JOIN receiving_history rh
                    ON ib.item_cd = rh.item_cd
                    AND ib.receiving_date = rh.receiving_date
                    {rh_sf}
                WHERE ib.status = 'expired'
                AND ib.expiry_date >= date('now', '-' || ? || ' days')
                {batch_sf}
                GROUP BY ib.item_cd
                HAVING total_waste > 0

                UNION ALL

                SELECT
                    ot.item_cd,
                    ot.item_nm,
                    SUM(ot.order_qty) as total_initial,
                    SUM(ot.order_qty) as total_receiving,
                    SUM(ot.remaining_qty) as total_waste,
                    ROUND(
                        CAST(SUM(ot.remaining_qty) AS REAL) /
                        NULLIF(SUM(ot.order_qty), 0) * 100, 1
                    ) as waste_rate
                FROM order_tracking ot
                WHERE ot.status IN ('expired', 'disposed')
                AND date(ot.expiry_time) >= date('now', '-' || ? || ' days')
                AND ot.remaining_qty > 0
                {ot_sf}
                GROUP BY ot.item_cd
                HAVING total_waste > 0

                ORDER BY waste_rate DESC
                LIMIT ?
                """,
                (days_back,) + sp + sp + (days_back,) + sp + (limit,)
            )
            return [dict(row) for row in cursor.fetchall()]
        except Exception as e:
            logger.warning(f"폐기율 TOP 상품 조회 실패 | store_id={self.store_id} | limit={limit}: {e}")
            return []
        finally:
            conn.close()


def generate_waste_report(target_date: Optional[str] = None, store_id: Optional[str] = None) -> Optional[Path]:
    """폐기 보고서 생성 (외부 호출용)

    Args:
        target_date: 기준 날짜 (기본: 오늘)
        store_id: 매장 ID (기본: None)

    Returns:
        생성된 파일 경로 (실패 시 None)
    """
    try:
        generator = WasteReportGenerator(store_id=store_id)
        return generator.generate(target_date)
    except Exception as e:
        logger.error(f"폐기 보고서 생성 실패: {e}")
        return None


# 테스트
if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding='utf-8')

    print("=== 폐기 보고서 생성 ===")
    path = generate_waste_report()
    if path:
        print(f"생성 완료: {path}")
    else:
        print("생성 실패")
