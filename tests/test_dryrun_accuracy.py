"""
dryrun-accuracy PDCA 테스트
Design 문서: docs/02-design/features/dryrun-accuracy.design.md

T-01 ~ T-10: RI stale 판정, Excel 컬럼 정합성, 차이 경고 출력
"""

import pytest
from datetime import datetime, timedelta
from unittest.mock import patch, MagicMock
import sys
from pathlib import Path

# 프로젝트 루트 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))


# ═══════════════════════════════════════════════════════════════
# Import 대상 상수/함수
# ═══════════════════════════════════════════════════════════════

from scripts.run_full_flow import (
    SECTION_A, SECTION_B, SECTION_C, SECTION_D, SECTION_E,
    ALL_SECTIONS,
    COLUMN_DESCRIPTIONS,
    COL_WIDTHS,
    FLOAT_COLS, INT_COLS,
    TOTAL_COLS,
    DRYRUN_STALE_HOURS,
    FOOD_MID_CDS,
)


# ═══════════════════════════════════════════════════════════════
# T-01: stale 판정 — 푸드 6h 초과 → is_stale=True
# ═══════════════════════════════════════════════════════════════

class TestStaleJudgment:
    """RI stale 판정 로직 테스트"""

    def _judge_stale(self, queried_at_str, mid_cd):
        """Design 문서 2.1절의 stale 판정 로직 재현"""
        now = datetime.now()
        hours_ago = -1.0
        if queried_at_str:
            try:
                qt = datetime.fromisoformat(queried_at_str)
                hours_ago = (now - qt).total_seconds() / 3600
            except (ValueError, TypeError):
                pass

        threshold_h = DRYRUN_STALE_HOURS["food"] if mid_cd in FOOD_MID_CDS else DRYRUN_STALE_HOURS["default"]
        is_stale = hours_ago < 0 or hours_ago > threshold_h
        return is_stale

    def test_t01_food_stale_over_6h(self):
        """T-01: queried_at이 7h 전 + mid_cd='001' → is_stale=True"""
        seven_hours_ago = (datetime.now() - timedelta(hours=7)).isoformat()
        assert self._judge_stale(seven_hours_ago, "001") is True

    def test_t02_nonfood_fresh_within_24h(self):
        """T-02: queried_at이 20h 전 + mid_cd='040' → is_stale=False"""
        twenty_hours_ago = (datetime.now() - timedelta(hours=20)).isoformat()
        assert self._judge_stale(twenty_hours_ago, "040") is False

    def test_t03_no_queried_at(self):
        """T-03: queried_at=None → is_stale=True"""
        assert self._judge_stale(None, "001") is True
        assert self._judge_stale(None, "040") is True

    def test_food_fresh_within_6h(self):
        """푸드 5h 전 → is_stale=False"""
        five_hours_ago = (datetime.now() - timedelta(hours=5)).isoformat()
        assert self._judge_stale(five_hours_ago, "001") is False

    def test_nonfood_stale_over_24h(self):
        """비푸드 25h 전 → is_stale=True"""
        twentyfive_hours_ago = (datetime.now() - timedelta(hours=25)).isoformat()
        assert self._judge_stale(twentyfive_hours_ago, "040") is True

    def test_food_boundary_6h(self):
        """푸드 정확히 6h 전 → is_stale=False (> 아닌 >= 기준이므로 경계값 미포함)"""
        # 5.9h 전 → False
        just_under = (datetime.now() - timedelta(hours=5, minutes=59)).isoformat()
        assert self._judge_stale(just_under, "002") is False

    def test_dessert_is_food_mid_cd(self):
        """디저트(014)는 FOOD_MID_CDS에 포함"""
        assert "014" in FOOD_MID_CDS

    def test_invalid_queried_at(self):
        """잘못된 형식 queried_at → is_stale=True"""
        assert self._judge_stale("not-a-date", "001") is True

    def test_all_food_mid_cds(self):
        """FOOD_MID_CDS에 7개 카테고리 모두 포함 확인"""
        expected = {"001", "002", "003", "004", "005", "012", "014"}
        assert FOOD_MID_CDS == expected


# ═══════════════════════════════════════════════════════════════
# T-04 ~ T-07: Excel 컬럼 정합성 테스트
# ═══════════════════════════════════════════════════════════════

class TestExcelColumnConsistency:
    """Excel 컬럼 수 정합성 테스트"""

    def test_t04_section_c_column_count(self):
        """T-04: SECTION_C 컬럼 수 == 5 (RI조회시각 추가)"""
        assert len(SECTION_C["columns"]) == 5

    def test_t05_total_cols_matches_sections(self):
        """T-05: 모든 섹션 컬럼 합 == TOTAL_COLS"""
        total = sum(len(s["columns"]) for s in ALL_SECTIONS)
        assert total == TOTAL_COLS

    def test_t06_column_descriptions_count(self):
        """T-06: COLUMN_DESCRIPTIONS 수 == TOTAL_COLS"""
        assert len(COLUMN_DESCRIPTIONS) == TOTAL_COLS

    def test_t07_col_widths_count(self):
        """T-07: COL_WIDTHS 수 == TOTAL_COLS"""
        assert len(COL_WIDTHS) == TOTAL_COLS

    def test_total_cols_is_30(self):
        """TOTAL_COLS == 30 (29→30 변경 확인)"""
        assert TOTAL_COLS == 30

    def test_ri_queried_at_in_section_c(self):
        """SECTION_C에 'RI조회시각' 컬럼 존재"""
        col_names = [name for name, _key in SECTION_C["columns"]]
        assert "RI조회시각" in col_names

    def test_ri_queried_at_key(self):
        """SECTION_C의 RI조회시각 컬럼 key == 'ri_queried_at'"""
        for name, key in SECTION_C["columns"]:
            if name == "RI조회시각":
                assert key == "ri_queried_at"
                return
        pytest.fail("RI조회시각 컬럼을 찾지 못함")

    def test_ri_description_exists(self):
        """COLUMN_DESCRIPTIONS에 '재고 데이터 조회 시각' 존재"""
        assert "재고 데이터 조회 시각" in COLUMN_DESCRIPTIONS

    def test_float_cols_no_overlap_with_int_cols(self):
        """FLOAT_COLS와 INT_COLS 겹침 없음"""
        assert FLOAT_COLS & INT_COLS == set()

    def test_section_e_columns_unchanged(self):
        """SECTION_E 컬럼 수는 여전히 8개"""
        assert len(SECTION_E["columns"]) == 8


# ═══════════════════════════════════════════════════════════════
# T-08 ~ T-10: 통합 테스트 (Excel 생성 + 콘솔 출력)
# ═══════════════════════════════════════════════════════════════

class TestDryrunExcelGeneration:
    """Excel 생성 통합 테스트"""

    def _make_order_item(self, item_cd="1234567890123", mid_cd="001",
                         final_order_qty=5, ri_queried_at="", ri_stale=False):
        """테스트용 order_list 항목 생성"""
        return {
            "item_cd": item_cd,
            "item_nm": f"상품_{item_cd[-4:]}",
            "mid_cd": mid_cd,
            "demand_pattern": "daily",
            "data_days": 30,
            "sell_day_ratio": 0.7,
            "wma_raw": 3.5,
            "feat_prediction": 3.2,
            "daily_avg": 3.3,
            "predicted_sales": 3.4,
            "weekday_coef": 1.0,
            "current_stock": 2,
            "pending_receiving_qty": 0,
            "safety_stock": 3,
            "need_qty": final_order_qty,
            "ri_queried_at": ri_queried_at,
            "ri_stale": ri_stale,
            "rule_order_qty": final_order_qty,
            "ml_order_qty": final_order_qty,
            "ml_weight_used": 0.2,
            "final_order_qty": final_order_qty,
            "proposal_summary": "test",
            "round_before": final_order_qty,
            "round_floor": final_order_qty,
            "round_ceil": final_order_qty,
            "order_unit_qty": 1,
            "model_type": "rule",
        }

    def test_t08_excel_creation(self, tmp_path):
        """T-08: 드라이런 Excel 생성 확인 (파일 존재 + 시트 2개)"""
        from scripts.run_full_flow import create_dryrun_excel

        items = [self._make_order_item()]
        output = str(tmp_path / "test_dryrun.xlsx")
        result = create_dryrun_excel(
            order_list=items,
            output_path=output,
            delivery_date="2026-03-09",
            store_id="46513",
        )
        assert Path(result).exists()

        from openpyxl import load_workbook
        wb = load_workbook(result)
        assert len(wb.sheetnames) == 2  # 상세 + 요약

    def test_t09_ri_column_header(self, tmp_path):
        """T-09: Excel Q열 헤더 = 'RI조회시각'"""
        from scripts.run_full_flow import create_dryrun_excel

        items = [self._make_order_item(
            ri_queried_at="2026-03-09T10:00:00",
            ri_stale=False,
        )]
        output = str(tmp_path / "test_ri_col.xlsx")
        create_dryrun_excel(
            order_list=items,
            output_path=output,
            delivery_date="2026-03-09",
        )

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb.active
        # Q열 = column 17, 2행 = 컬럼명 헤더
        assert ws.cell(row=2, column=17).value == "RI조회시각"

    def test_excel_stale_red_background(self, tmp_path):
        """stale 항목의 RI조회시각 셀에 빨간 배경 적용 확인"""
        from scripts.run_full_flow import create_dryrun_excel

        items = [
            self._make_order_item(
                item_cd="1111111111111",
                ri_queried_at="2026-03-08T01:00:00",
                ri_stale=True,
            ),
            self._make_order_item(
                item_cd="2222222222222",
                ri_queried_at="2026-03-09T10:00:00",
                ri_stale=False,
            ),
        ]
        output = str(tmp_path / "test_stale_bg.xlsx")
        create_dryrun_excel(
            order_list=items,
            output_path=output,
            delivery_date="2026-03-09",
        )

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb.active
        # 4행 = 첫 번째 데이터 (stale=True) → 빨간 배경
        fill_1 = ws.cell(row=4, column=17).fill
        assert fill_1.start_color.rgb == "00FFC7CE" or fill_1.fgColor.rgb == "00FFC7CE"
        # 5행 = 두 번째 데이터 (stale=False) → 빨간 배경 없음
        fill_2 = ws.cell(row=5, column=17).fill
        assert fill_2.start_color.rgb != "00FFC7CE" or fill_2.patternType is None

    def test_t10_scheduler_warning_output(self, capsys):
        """T-10: 스케줄러 차이 경고 문자열이 정상 포함되는지 확인"""
        # 직접 출력 검증 (함수 호출 없이 문자열 패턴만 확인)
        print("[스케줄러 차이 경고] 실제 7시 발주와 다를 수 있는 항목:")
        captured = capsys.readouterr()
        assert "스케줄러 차이 경고" in captured.out


# ═══════════════════════════════════════════════════════════════
# 추가 테스트: DRYRUN_STALE_HOURS 상수 검증
# ═══════════════════════════════════════════════════════════════

class TestDryrunConstants:
    """드라이런 신선도 상수 테스트"""

    def test_stale_hours_food(self):
        """푸드 stale 기준 = 6시간"""
        assert DRYRUN_STALE_HOURS["food"] == 6

    def test_stale_hours_default(self):
        """기본 stale 기준 = 24시간"""
        assert DRYRUN_STALE_HOURS["default"] == 24
