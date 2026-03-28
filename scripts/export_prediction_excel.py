"""46513 매장 3/28 예측 결과 엑셀 내보내기"""
import sqlite3
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

store_id = "46513"
eval_date = "2026-03-28"
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

conn = sqlite3.connect(os.path.join(base_dir, f"data/stores/{store_id}.db"))
conn.row_factory = sqlite3.Row
common = sqlite3.connect(os.path.join(base_dir, "data/common.db"))
common.row_factory = sqlite3.Row

# Lookups
mid_map = {r["mid_cd"]: r["mid_nm"] for r in common.execute("SELECT mid_cd, mid_nm FROM mid_categories")}
prod_map = {r["item_cd"]: r["item_nm"] for r in common.execute("SELECT item_cd, item_nm FROM products")}
detail_map = {}
for r in common.execute("SELECT item_cd, expiration_days, demand_pattern, order_unit_qty, promo_type FROM product_details"):
    detail_map[r["item_cd"]] = dict(r)

# Main data
rows = conn.execute("""
    SELECT item_cd, mid_cd, decision, daily_avg, current_stock, pending_qty,
           predicted_qty, actual_order_qty, order_status, sell_price, margin_rate,
           promo_type, trend_score, stockout_freq, exposure_days, popularity_score
    FROM eval_outcomes
    WHERE store_id = ? AND eval_date = ?
    ORDER BY mid_cd, item_cd
""", (store_id, eval_date)).fetchall()

pred_map = {}
for r in conn.execute("""
    SELECT item_cd, safety_stock, weekday_coef, ml_order_qty, ml_weight_used, confidence
    FROM prediction_logs WHERE store_id = ? AND prediction_date = ?
""", (store_id, eval_date)):
    pred_map[r["item_cd"]] = dict(r)

# Create workbook
wb = openpyxl.Workbook()
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
order_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
force_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
skip_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# ===== Sheet 1: 전체 목록 =====
ws = wb.active
ws.title = "전체 예측 목록"

headers = [
    "No", "상품코드", "상품명", "중분류코드", "중분류명",
    "판정", "일평균판매", "현재고", "입고예정",
    "예측수량", "실제발주수량", "발주상태",
    "판매가", "마진율(%)", "행사",
    "트렌드", "품절빈도", "노출일수", "인기점수",
    "수요패턴", "유통기한(일)", "배수단위",
    "안전재고", "요일계수", "ML수량", "ML가중치", "신뢰도",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[1].height = 35

for i, row in enumerate(rows, 1):
    d = dict(row)
    item_cd = d["item_cd"]
    pred = pred_map.get(item_cd, {})
    detail = detail_map.get(item_cd, {})

    values = [
        i,
        item_cd,
        prod_map.get(item_cd, ""),
        d["mid_cd"],
        mid_map.get(d["mid_cd"], ""),
        d["decision"],
        round(d["daily_avg"], 2) if d["daily_avg"] else 0,
        d["current_stock"] or 0,
        d["pending_qty"] or 0,
        d["predicted_qty"] if d["predicted_qty"] is not None else "",
        d["actual_order_qty"] if d["actual_order_qty"] is not None else "",
        d["order_status"] or "",
        d["sell_price"] or 0,
        round(d["margin_rate"], 1) if d["margin_rate"] else 0,
        d["promo_type"] or "",
        round(d["trend_score"], 3) if d["trend_score"] else "",
        round(d["stockout_freq"], 3) if d["stockout_freq"] else "",
        round(d["exposure_days"], 1) if d["exposure_days"] else 0,
        round(d["popularity_score"], 4) if d["popularity_score"] else 0,
        detail.get("demand_pattern", ""),
        detail.get("expiration_days", ""),
        detail.get("order_unit_qty", ""),
        round(pred.get("safety_stock", 0), 2) if pred.get("safety_stock") else "",
        round(pred.get("weekday_coef", 0), 3) if pred.get("weekday_coef") else "",
        pred.get("ml_order_qty", ""),
        round(pred.get("ml_weight_used", 0), 3) if pred.get("ml_weight_used") else "",
        pred.get("confidence", ""),
    ]

    for col, v in enumerate(values, 1):
        cell = ws.cell(row=i + 1, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col != 3 else "left")

        decision = d["decision"]
        if decision in ("NORMAL_ORDER", "URGENT_ORDER") and col in (6, 10, 11):
            cell.fill = order_fill
        elif decision == "FORCE_ORDER" and col in (6, 10, 11):
            cell.fill = force_fill
        elif decision == "SKIP" and col == 6:
            cell.fill = skip_fill

widths = [6, 16, 30, 8, 14, 14, 10, 8, 8, 8, 10, 10, 8, 8, 8, 8, 8, 8, 8, 10, 10, 8, 8, 8, 8, 8, 8]
for idx, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

# ===== Sheet 2: 발주 상품만 =====
ws2 = wb.create_sheet("발주 상품")
ordered_rows = [r for r in rows if dict(r).get("actual_order_qty") and dict(r)["actual_order_qty"] > 0]

headers2 = [
    "No", "상품코드", "상품명", "중분류명", "판정",
    "발주수량", "배수단위", "현재고", "일평균",
    "판매가", "마진율(%)", "행사", "수요패턴", "발주상태",
]

for col, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws2.row_dimensions[1].height = 30

for i, row in enumerate(ordered_rows, 1):
    d = dict(row)
    item_cd = d["item_cd"]
    detail = detail_map.get(item_cd, {})
    values = [
        i, item_cd, prod_map.get(item_cd, ""),
        mid_map.get(d["mid_cd"], ""), d["decision"],
        d["actual_order_qty"], detail.get("order_unit_qty", ""),
        d["current_stock"] or 0,
        round(d["daily_avg"], 2) if d["daily_avg"] else 0,
        d["sell_price"] or 0,
        round(d["margin_rate"], 1) if d["margin_rate"] else 0,
        d["promo_type"] or "", detail.get("demand_pattern", ""),
        d["order_status"] or "",
    ]
    for col, v in enumerate(values, 1):
        cell = ws2.cell(row=i + 1, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col != 3 else "left")

widths2 = [6, 16, 30, 14, 14, 10, 8, 8, 8, 8, 8, 8, 10, 10]
for idx, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(idx)].width = w
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{len(ordered_rows) + 1}"

# ===== Sheet 3: 요약 =====
ws3 = wb.create_sheet("요약")
ws3.cell(row=1, column=1, value=f"{store_id} 매장 발주 예측 요약 ({eval_date})").font = Font(bold=True, size=14)

summary_font = Font(bold=True, size=11)
for col, h in enumerate(["구분", "건수", "발주수량"], 1):
    ws3.cell(row=3, column=col, value=h).font = summary_font

decisions = {}
for row in rows:
    d = dict(row)
    dec = d["decision"]
    if dec not in decisions:
        decisions[dec] = {"count": 0, "qty": 0}
    decisions[dec]["count"] += 1
    decisions[dec]["qty"] += (d["actual_order_qty"] or 0)

r = 4
for dec in ["PASS", "NORMAL_ORDER", "FORCE_ORDER", "URGENT_ORDER", "SKIP"]:
    if dec in decisions:
        ws3.cell(row=r, column=1, value=dec)
        ws3.cell(row=r, column=2, value=decisions[dec]["count"])
        ws3.cell(row=r, column=3, value=decisions[dec]["qty"])
        r += 1

ws3.cell(row=r, column=1, value="합계").font = summary_font
ws3.cell(row=r, column=2, value=sum(v["count"] for v in decisions.values())).font = summary_font
ws3.cell(row=r, column=3, value=sum(v["qty"] for v in decisions.values())).font = summary_font

r += 2
ws3.cell(row=r, column=1, value="중분류별 발주 현황").font = Font(bold=True, size=13)
r += 1
for col, h in enumerate(["중분류", "중분류명", "발주건수", "총수량"], 1):
    ws3.cell(row=r, column=col, value=h).font = summary_font
r += 1

mid_summary = {}
for row in rows:
    d = dict(row)
    if d["actual_order_qty"] and d["actual_order_qty"] > 0:
        mid = d["mid_cd"]
        if mid not in mid_summary:
            mid_summary[mid] = {"count": 0, "qty": 0}
        mid_summary[mid]["count"] += 1
        mid_summary[mid]["qty"] += d["actual_order_qty"]

for mid in sorted(mid_summary.keys()):
    ws3.cell(row=r, column=1, value=mid)
    ws3.cell(row=r, column=2, value=mid_map.get(mid, ""))
    ws3.cell(row=r, column=3, value=mid_summary[mid]["count"])
    ws3.cell(row=r, column=4, value=mid_summary[mid]["qty"])
    r += 1

ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 18
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 12

# Save
output_dir = os.path.join(base_dir, "data", "reports")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, f"{store_id}_prediction_{eval_date.replace('-', '')}.xlsx")
wb.save(output_path)

print(f"Excel saved: {output_path}")
print(f"  Sheet 1 (전체 예측 목록): {len(rows)}건")
print(f"  Sheet 2 (발주 상품): {len(ordered_rows)}건")
print(f"  Sheet 3 (요약): 판정별 + 중분류별")

conn.close()
common.close()
