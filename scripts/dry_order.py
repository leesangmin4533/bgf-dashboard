"""
모의 발주 리포트 생성 스크립트

⚠️ 안전장치: 이 스크립트는 절대로 실제 발주를 실행하지 않습니다.
- OrderExecutor import 없음
- execute(), run_daily_order(), execute_orders(), confirm_order() 호출 없음
- get_recommendations() + prefetch_pending_quantities() 만 사용

실제 발주와 100% 동일한 목록을 생성하되, 결과를 Excel + 텍스트로만 출력합니다.
daily_job.py → _run_auto_order_with_driver() 와 동일한 파라미터:
  min_order_qty=1, max_items=None, prefetch_pending=True
"""

import sys
import io
import argparse

# Windows CP949 콘솔 → UTF-8 래핑 (한글·특수문자 깨짐 방지)
if sys.platform == "win32":
    for _sn in ("stdout", "stderr"):
        _s = getattr(sys, _sn, None)
        if _s and hasattr(_s, "buffer") and (getattr(_s, "encoding", "") or "").lower().replace("-", "") != "utf8":
            setattr(sys, _sn, io.TextIOWrapper(_s.buffer, encoding="utf-8", errors="replace", line_buffering=True))
import time
import traceback
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from typing import Any, Dict, List, Optional

# 프로젝트 루트를 path에 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.order.auto_order import AutoOrderSystem
from src.prediction.pre_order_evaluator import EvalDecision, DECISION_LABELS
from src.infrastructure.database.repos import (
    SalesRepository,
    RealtimeInventoryRepository,
    ProductDetailRepository,
)
from src.settings.constants import DEFAULT_STORE_ID


# =============================================================================
# 상수
# =============================================================================
OUTPUT_DIR = Path(__file__).parent.parent.parent / "test report"

# 판정 색상 (openpyxl hex, ARGB)
DECISION_COLORS = {
    EvalDecision.FORCE_ORDER:  "FFFF0000",   # 빨강
    EvalDecision.URGENT_ORDER: "FFFF8C00",   # 주황
    EvalDecision.NORMAL_ORDER: "FFFFFF00",    # 노랑
    EvalDecision.PASS:         "FF87CEEB",    # 하늘색
    EvalDecision.SKIP:         "FFC0C0C0",    # 회색
}

DECISION_SORT_ORDER = {
    EvalDecision.FORCE_ORDER: 0,
    EvalDecision.URGENT_ORDER: 1,
    EvalDecision.NORMAL_ORDER: 2,
    EvalDecision.PASS: 3,
    EvalDecision.SKIP: 4,
}


# =============================================================================
# 중분류 이름 조회
# =============================================================================
def _get_mid_category_names(store_id: str = None) -> Dict[str, str]:
    """DB에서 mid_cd → mid_nm 매핑 조회"""
    try:
        repo = SalesRepository(store_id=store_id)
        categories = repo.get_all_mid_categories()
        return {c["mid_cd"]: c["mid_nm"] for c in categories}
    except Exception:
        return {}


# =============================================================================
# 메인 실행 플로우
# =============================================================================
def run_dry_order(
    no_login: bool = False,
    no_pending: bool = False,
    max_items: Optional[int] = None,
    store_id: str = DEFAULT_STORE_ID,
) -> None:
    """
    모의 발주 리포트 생성

    Args:
        no_login: True면 로그인 없이 DB 데이터만으로 예측
        no_pending: True면 미입고 조회 스킵
        max_items: 상품 수 제한 (None이면 전체)
        store_id: 점포 코드
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    print("\n" + "=" * 70)
    print("모의 발주 리포트 생성")
    print(f"시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"모드: {'오프라인 (DB만)' if no_login else '온라인 (로그인)'}")
    print(f"미입고 조회: {'스킵' if no_pending else '실행'}")
    print(f"상품 제한: {max_items if max_items else '없음 (전체)'}")
    print(f"점포: {store_id}")
    print("=" * 70)

    analyzer = None
    system = None
    driver = None

    try:
        # ── 1단계: 로그인 ──
        if not no_login:
            print("\n[1/7] BGF 시스템 로그인...")
            from src.sales_analyzer import SalesAnalyzer

            analyzer = SalesAnalyzer()
            analyzer.setup_driver()
            analyzer.connect()

            if not analyzer.do_login():
                print("[ERROR] 로그인 실패")
                return
            print("[OK] 로그인 성공")
            time.sleep(2)
            driver = analyzer.driver
        else:
            print("\n[1/7] 로그인 스킵 (--no-login)")

        # ── 2단계: 시스템 초기화 ──
        print("\n[2/7] 자동 발주 시스템 초기화 (예측 전용)...")
        system = AutoOrderSystem(driver=driver, use_improved_predictor=True, store_id=store_id)

        # DB에서 미취급 상품 목록 로드
        system.load_unavailable_from_db()

        # DB에서 발주중지(CUT) 상품 목록 로드
        system.load_cut_items_from_db()

        # DB 인벤토리 캐시 로드
        system.load_inventory_cache_from_db()

        # ── 3단계: 발주 추천 생성 ──
        print("\n[3/7] 발주 추천 목록 생성 (전체 상품)...")
        order_list = system.get_recommendations(
            min_order_qty=1,
            max_items=None,  # 전체 상품 대상
        )

        if not order_list:
            print("[INFO] 발주 대상 상품 없음")
            return

        print(f"  → 추천 상품: {len(order_list)}개")

        # ── 4단계: 미입고 조회 ──
        pending_data = {}
        stock_data = {}
        if not no_pending and not no_login and system.pending_collector:
            print(f"\n[4/7] 미입고 수량 사전 조회 ({len(order_list)}개 상품)...")
            candidate_items = [
                item.get("item_cd") for item in order_list if item.get("item_cd")
            ]
            pending_data = system.prefetch_pending_quantities(
                item_codes=candidate_items,
                max_items=len(candidate_items),
            )
            stock_data = getattr(system, "_last_stock_data", {})
            print(f"  → 미입고 데이터: {len(pending_data)}개 상품")
            print(f"  → 실시간 재고: {len(stock_data)}개 상품")
        else:
            reason = "--no-pending" if no_pending else ("--no-login" if no_login else "드라이버 없음")
            print(f"\n[4/7] 미입고 조회 스킵 ({reason})")

        # ── 5단계: 미입고/재고 반영 ──
        if pending_data or stock_data:
            print("\n[5/7] 미입고/재고 반영 조정...")
            before_count = len(order_list)
            before_qty = sum(item.get("final_order_qty", 0) for item in order_list)

            order_list = system._apply_pending_and_stock_to_order_list(
                order_list=order_list,
                pending_data=pending_data,
                stock_data=stock_data,
                min_order_qty=1,
            )

            after_count = len(order_list)
            after_qty = sum(item.get("final_order_qty", 0) for item in order_list)
            print(f"  → 상품: {before_count} → {after_count}개")
            print(f"  → 수량: {before_qty} → {after_qty}개 ({before_qty - after_qty}개 감소)")
        else:
            print("\n[5/7] 조정 스킵 (미입고 데이터 없음)")

        # max_items 적용 (리포트 제한용)
        if max_items and len(order_list) > max_items:
            order_list = order_list[:max_items]
            print(f"  → 상위 {max_items}개로 제한")

        # ── 발주중지(CUT) 상품 수집 ──
        cut_items_detail = []
        if system._cut_items:
            inv_repo = RealtimeInventoryRepository(store_id=store_id)
            cut_items_detail = inv_repo.get_cut_items_detail()
            print(f"  → 발주중지(CUT) 상품: {len(cut_items_detail)}개")

        # ── 6단계: 사전 평가 결과 수집 ──
        print("\n[6/7] 사전 평가 결과 매핑...")
        eval_results = {}
        try:
            eval_results = system.pre_order_evaluator.evaluate_all(write_log=False)
        except Exception as e:
            print(f"  [WARN] 사전 평가 재실행 실패: {e}")

        # 중분류 이름 매핑
        mid_names = _get_mid_category_names(store_id=store_id)

        # 매가/이익율 매핑
        product_details_map = {}
        try:
            pd_repo = ProductDetailRepository()
            for pd in pd_repo.get_all():
                product_details_map[pd["item_cd"]] = pd
        except Exception as e:
            print(f"  [WARN] 상품 상세 로드 실패: {e}")

        margin_count = sum(1 for pd in product_details_map.values() if pd.get("margin_rate") is not None)
        print(f"  → 상품 상세: {len(product_details_map)}개 (매가/이익율: {margin_count}개)")

        # ── 7단계: 리포트 생성 ──
        print("\n[7/7] 리포트 생성...")
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

        xlsx_path = OUTPUT_DIR / f"dry_order_{timestamp}.xlsx"
        txt_path = OUTPUT_DIR / f"dry_order_{timestamp}.txt"

        _write_excel_report(order_list, eval_results, mid_names, xlsx_path, cut_items_detail, product_details_map)
        _write_text_report(order_list, eval_results, mid_names, txt_path, cut_items_detail, product_details_map)

        print(f"\n{'=' * 70}")
        print("모의 발주 리포트 생성 완료")
        print(f"  Excel: {xlsx_path}")
        print(f"  텍스트: {txt_path}")
        print(f"  총 상품: {len(order_list)}개")
        print(f"  총 발주량: {sum(item.get('final_order_qty', 0) for item in order_list)}개")
        print(f"{'=' * 70}")

    except Exception as e:
        print(f"\n[ERROR] 오류 발생: {e}")
        traceback.print_exc()

    finally:
        if system:
            try:
                system.close()
            except Exception:
                pass
        if analyzer:
            try:
                analyzer.close()
            except Exception:
                pass


# =============================================================================
# Excel 리포트 생성
# =============================================================================
def _write_excel_report(
    order_list: List[Dict[str, Any]],
    eval_results: Dict[str, Any],
    mid_names: Dict[str, str],
    output_path: Path,
    cut_items: Optional[List[Dict[str, Any]]] = None,
    product_details_map: Optional[Dict[str, Dict[str, Any]]] = None,
) -> None:
    """Excel 리포트 생성 (Sheet 1: 발주 상세, Sheet 2: 요약)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: 발주 상세 ──
    ws = wb.active
    ws.title = "발주 상세"

    headers = [
        "No.", "상품코드", "상품명", "중분류",
        "발주 판정", "판정 사유", "최종 발주량",
        "현재 재고", "미입고", "예측 판매량",
        "안전재고", "일평균", "요일계수",
        "신뢰도", "데이터일수", "입수",
        "매가", "이익율(%)",
        "인기도", "노출일수",
    ]

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFFFF", size=10)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 컬럼 폭 설정
    col_widths = [5, 16, 30, 14, 12, 30, 12, 10, 8, 12, 10, 8, 8, 8, 10, 6, 10, 10, 8, 8]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 데이터 행
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    pd_map = product_details_map or {}

    for row_idx, item in enumerate(order_list, 2):
        item_cd = item.get("item_cd", "")
        mid_cd = item.get("mid_cd", "")
        eval_r = eval_results.get(item_cd)
        final_qty = item.get("final_order_qty", 0)
        pd_info = pd_map.get(item_cd, {})

        # 판정 정보
        if eval_r:
            decision = eval_r.decision
            decision_label = DECISION_LABELS.get(decision, str(decision.value))
            reason = eval_r.reason
            popularity = eval_r.popularity_score
            exposure = eval_r.exposure_days
        else:
            decision = None
            decision_label = "-"
            reason = "-"
            popularity = 0.0
            exposure = 0.0

        # 매가/이익율
        sell_price = pd_info.get("sell_price")
        margin_rate = pd_info.get("margin_rate")

        # 중분류 표시
        mid_nm = mid_names.get(mid_cd, "")
        mid_display = f"{mid_cd} {mid_nm}" if mid_nm else mid_cd

        row_data = [
            row_idx - 1,                              # A: No.
            item_cd,                                   # B: 상품코드
            item.get("item_nm", ""),                   # C: 상품명
            mid_display,                               # D: 중분류
            decision_label,                            # E: 발주 판정
            reason,                                    # F: 판정 사유
            final_qty,                                 # G: 최종 발주량
            item.get("current_stock", 0),              # H: 현재 재고
            item.get("pending_receiving_qty", 0),      # I: 미입고
            item.get("predicted_sales", 0),            # J: 예측 판매량
            round(item.get("safety_stock", 0), 1),     # K: 안전재고
            round(item.get("daily_avg", 0), 2),        # L: 일평균
            round(item.get("weekday_coef", 0), 2),     # M: 요일계수
            item.get("confidence", ""),                 # N: 신뢰도
            item.get("data_days", 0),                   # O: 데이터일수
            item.get("order_unit_qty", 1),              # P: 입수
            sell_price if sell_price else "",           # Q: 매가
            round(margin_rate, 2) if margin_rate else "",  # R: 이익율(%)
            round(popularity, 3),                       # S: 인기도
            round(exposure, 1),                         # T: 노출일수
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # 숫자 열 우측 정렬 (매가=17, 이익율=18 추가)
            if col_idx in (1, 7, 8, 9, 10, 11, 12, 13, 15, 16, 17, 18, 19, 20):
                cell.alignment = right_align
            elif col_idx in (5, 14):
                cell.alignment = center_align

        # ── 판정 색상 코딩 (E열) ──
        if decision and decision in DECISION_COLORS:
            fill_color = DECISION_COLORS[decision]
            ws.cell(row=row_idx, column=5).fill = PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type="solid"
            )
            # FORCE/URGENT 텍스트는 흰색
            if decision in (EvalDecision.FORCE_ORDER, EvalDecision.URGENT_ORDER):
                ws.cell(row=row_idx, column=5).font = Font(bold=True, color="FFFFFFFF")

        # ── 대량 발주 경고 (G열) ──
        qty_cell = ws.cell(row=row_idx, column=7)
        qty_cell.font = Font(bold=True)
        if final_qty >= 10:
            qty_cell.font = Font(bold=True, color="FFFF0000")  # 빨간 굵은 글씨
        elif final_qty >= 5:
            qty_cell.font = Font(bold=True, color="FFFF8C00")  # 주황 굵은 글씨

    # 필터 추가
    ws.auto_filter.ref = f"A1:T{len(order_list) + 1}"

    # 첫 행 고정
    ws.freeze_panes = "A2"

    # ── Sheet 2: 요약 ──
    ws2 = wb.create_sheet("요약")

    # 제목
    ws2.cell(row=1, column=1, value="모의 발주 리포트 요약").font = Font(bold=True, size=14)
    ws2.merge_cells("A1:D1")

    # 메타 정보
    row = 3
    meta_items = [
        ("생성 시각", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("총 상품 수", len(order_list)),
        ("총 발주량", sum(item.get("final_order_qty", 0) for item in order_list)),
        ("발주중지 상품", len(cut_items) if cut_items else 0),
    ]
    for label, value in meta_items:
        ws2.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=row, column=2, value=value)
        row += 1

    # ── 판정별 분포 ──
    row += 1
    ws2.cell(row=row, column=1, value="판정별 분포").font = Font(bold=True, size=12)
    row += 1

    summary_headers = ["판정", "상품 수", "총 발주량", "비율"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    row += 1

    # 판정별 집계
    decision_stats = defaultdict(lambda: {"count": 0, "qty": 0})
    for item in order_list:
        item_cd = item.get("item_cd", "")
        eval_r = eval_results.get(item_cd)
        dec = eval_r.decision if eval_r else EvalDecision.PASS
        decision_stats[dec]["count"] += 1
        decision_stats[dec]["qty"] += item.get("final_order_qty", 0)

    total_count = len(order_list) or 1
    for dec in [EvalDecision.FORCE_ORDER, EvalDecision.URGENT_ORDER,
                EvalDecision.NORMAL_ORDER, EvalDecision.PASS, EvalDecision.SKIP]:
        stats = decision_stats.get(dec, {"count": 0, "qty": 0})
        label = DECISION_LABELS.get(dec, str(dec.value))
        ratio = f"{stats['count'] / total_count * 100:.1f}%"
        ws2.cell(row=row, column=1, value=label)
        ws2.cell(row=row, column=2, value=stats["count"])
        ws2.cell(row=row, column=3, value=stats["qty"])
        ws2.cell(row=row, column=4, value=ratio)
        # 색상
        if dec in DECISION_COLORS:
            ws2.cell(row=row, column=1).fill = PatternFill(
                start_color=DECISION_COLORS[dec],
                end_color=DECISION_COLORS[dec],
                fill_type="solid",
            )
        row += 1

    # ── 중분류별 통계 ──
    row += 1
    ws2.cell(row=row, column=1, value="중분류별 통계").font = Font(bold=True, size=12)
    row += 1

    cat_headers = ["중분류", "상품 수", "총 발주량", "평균 발주량"]
    for col, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=row, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    row += 1

    cat_stats = defaultdict(lambda: {"count": 0, "qty": 0})
    for item in order_list:
        mid_cd = item.get("mid_cd", "")
        mid_nm = mid_names.get(mid_cd, "")
        key = f"{mid_cd} {mid_nm}" if mid_nm else mid_cd
        cat_stats[key]["count"] += 1
        cat_stats[key]["qty"] += item.get("final_order_qty", 0)

    for key in sorted(cat_stats.keys()):
        stats = cat_stats[key]
        avg = stats["qty"] / stats["count"] if stats["count"] else 0
        ws2.cell(row=row, column=1, value=key)
        ws2.cell(row=row, column=2, value=stats["count"])
        ws2.cell(row=row, column=3, value=stats["qty"])
        ws2.cell(row=row, column=4, value=round(avg, 1))
        row += 1

    # ── 발주중지(CUT) 상품 ──
    if cut_items:
        row += 1
        ws2.cell(row=row, column=1, value="발주중지(CUT) 상품").font = Font(bold=True, size=12)
        ws2.cell(row=row, column=1).fill = PatternFill(
            start_color="FF00CED1", end_color="FF00CED1", fill_type="solid"
        )
        row += 1

        cut_headers = ["상품코드", "상품명"]
        for col, h in enumerate(cut_headers, 1):
            cell = ws2.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
        row += 1

        for ci in sorted(cut_items, key=lambda x: x.get("item_cd", "")):
            ws2.cell(row=row, column=1, value=ci.get("item_cd", ""))
            ws2.cell(row=row, column=2, value=ci.get("item_nm", ""))
            row += 1

    # 컬럼 폭
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 12

    wb.save(str(output_path))
    print(f"  Excel 저장: {output_path}")


# =============================================================================
# 텍스트 리포트 생성
# =============================================================================
def _write_text_report(
    order_list: List[Dict[str, Any]],
    eval_results: Dict[str, Any],
    mid_names: Dict[str, str],
    output_path: Path,
    cut_items: Optional[List[Dict[str, Any]]] = None,
    product_details_map: Optional[Dict[str, Dict[str, Any]]] = None,
) -> None:
    """텍스트 리포트 생성"""
    lines = []
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_qty = sum(item.get("final_order_qty", 0) for item in order_list)

    lines.append("=" * 80)
    lines.append("모의 발주 리포트")
    lines.append(f"생성: {now_str}")
    lines.append(f"총 상품: {len(order_list)}개 | 총 발주량: {total_qty}개")
    lines.append("[주의] 이 리포트는 검증용이며, 실제 발주는 실행되지 않았습니다.")
    lines.append("=" * 80)

    # ── 판정별 요약 ──
    lines.append("\n[판정별 분포]")
    decision_stats = defaultdict(lambda: {"count": 0, "qty": 0})
    for item in order_list:
        item_cd = item.get("item_cd", "")
        eval_r = eval_results.get(item_cd)
        dec = eval_r.decision if eval_r else EvalDecision.PASS
        decision_stats[dec]["count"] += 1
        decision_stats[dec]["qty"] += item.get("final_order_qty", 0)

    for dec in [EvalDecision.FORCE_ORDER, EvalDecision.URGENT_ORDER,
                EvalDecision.NORMAL_ORDER, EvalDecision.PASS, EvalDecision.SKIP]:
        stats = decision_stats.get(dec, {"count": 0, "qty": 0})
        label = DECISION_LABELS.get(dec, str(dec.value))
        lines.append(f"  {label:<12}: {stats['count']:>4}개 상품, 발주량 {stats['qty']:>5}개")

    # ── 중분류별 요약 ──
    lines.append("\n[중분류별 통계]")
    cat_stats = defaultdict(lambda: {"count": 0, "qty": 0})
    for item in order_list:
        mid_cd = item.get("mid_cd", "")
        mid_nm = mid_names.get(mid_cd, "")
        key = f"{mid_cd} {mid_nm}" if mid_nm else mid_cd
        cat_stats[key]["count"] += 1
        cat_stats[key]["qty"] += item.get("final_order_qty", 0)

    for key in sorted(cat_stats.keys()):
        stats = cat_stats[key]
        lines.append(f"  {key:<16}: {stats['count']:>4}개 상품, 발주량 {stats['qty']:>5}개")

    # ── 대량 발주 경고 ──
    large_orders = [
        item for item in order_list if item.get("final_order_qty", 0) >= 5
    ]
    if large_orders:
        lines.append(f"\n[대량 발주 경고] ({len(large_orders)}개 상품)")
        large_orders.sort(key=lambda x: -x.get("final_order_qty", 0))
        for item in large_orders[:30]:
            qty = item.get("final_order_qty", 0)
            marker = "[!!]" if qty >= 10 else "[! ]"
            item_cd = item.get("item_cd", "")
            item_nm = (item.get("item_nm", "") or "")[:25]
            eval_r = eval_results.get(item_cd)
            dec_label = DECISION_LABELS.get(eval_r.decision, "-") if eval_r else "-"
            lines.append(
                f"  {marker} {item_cd} {item_nm:<25} "
                f"발주={qty:>3}개  재고={item.get('current_stock', 0):>3}  "
                f"미입고={item.get('pending_receiving_qty', 0):>3}  "
                f"판정={dec_label}"
            )
        if len(large_orders) > 30:
            lines.append(f"  ... 외 {len(large_orders) - 30}개")

    # ── 발주중지(CUT) 상품 ──
    if cut_items:
        lines.append(f"\n[발주중지(CUT) 상품] ({len(cut_items)}개)")
        for ci in sorted(cut_items, key=lambda x: x.get("item_cd", "")):
            item_cd = ci.get("item_cd", "")
            item_nm = (ci.get("item_nm", "") or "")[:30]
            lines.append(f"  {item_cd}  {item_nm}")

    # ── 상세 목록 ──
    pd_map = product_details_map or {}

    lines.append(f"\n{'=' * 130}")
    lines.append("[상세 목록]")
    lines.append(
        f"{'No.':<5} {'상품코드':<16} {'상품명':<25} {'판정':<10} "
        f"{'발주':>5} {'재고':>5} {'미입고':>5} {'예측':>5} {'안전재고':>6} {'일평균':>6} "
        f"{'매가':>8} {'이익율':>7}"
    )
    lines.append("-" * 130)

    for i, item in enumerate(order_list, 1):
        item_cd = item.get("item_cd", "")
        item_nm = (item.get("item_nm", "") or "")[:23]
        eval_r = eval_results.get(item_cd)
        dec_label = DECISION_LABELS.get(eval_r.decision, "-") if eval_r else "-"
        pd_info = pd_map.get(item_cd, {})
        sell_price = pd_info.get("sell_price")
        margin_rate = pd_info.get("margin_rate")
        price_str = f"{sell_price:>8,}" if sell_price else f"{'':>8}"
        margin_str = f"{margin_rate:>6.1f}%" if margin_rate else f"{'':>7}"

        lines.append(
            f"{i:<5} {item_cd:<16} {item_nm:<25} {dec_label:<10} "
            f"{item.get('final_order_qty', 0):>5} "
            f"{item.get('current_stock', 0):>5} "
            f"{item.get('pending_receiving_qty', 0):>5} "
            f"{item.get('predicted_sales', 0):>5} "
            f"{item.get('safety_stock', 0):>6.1f} "
            f"{item.get('daily_avg', 0):>6.2f} "
            f"{price_str} {margin_str}"
        )

    lines.append("=" * 130)

    text = "\n".join(lines)
    output_path.write_text(text, encoding="utf-8")
    print(f"  텍스트 저장: {output_path}")


# =============================================================================
# CLI 진입점
# =============================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="모의 발주 리포트 생성 (실제 발주 없음)"
    )
    parser.add_argument(
        "--no-login",
        action="store_true",
        help="오프라인 모드 (로그인 없이 DB 데이터만으로 예측)",
    )
    parser.add_argument(
        "--no-pending",
        action="store_true",
        help="미입고 조회 스킵 (빠른 실행)",
    )
    parser.add_argument(
        "--max-items",
        type=int,
        default=None,
        help="리포트에 포함할 최대 상품 수",
    )
    parser.add_argument(
        "--store-id",
        type=str,
        default=DEFAULT_STORE_ID,
        help="점포 코드 (예: 46513)",
    )
    args = parser.parse_args()

    run_dry_order(
        no_login=args.no_login,
        no_pending=args.no_pending,
        max_items=args.max_items,
        store_id=args.store_id,
    )
