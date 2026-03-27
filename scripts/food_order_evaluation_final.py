"""
푸드 발주 평가 최종본 엑셀 리포트
기간: 2026-03-01 ~ 2026-03-07

평가 철학: "폐기를 최소화하면서 팔 수 있는 만큼 파는 것"
- 핵심 평가: 이천호반베르디움점 auto 발주, buy_qty > 0 건만
- 참고: 나머지 매장/출처는 '참고 수준' 명시
- 데이터 품질 이슈 문서화

시트 구성:
  1. 핵심 평가 (이천호반베르디움 auto)
  2. 과발주 상품 (폐기 발생)
  3. 수요초과 가능성 상품
  4. 전매장 참고 현황
  5. 데이터 품질 이슈 기록
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 ──
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
COMMON_DB = os.path.join(DATA_DIR, "common.db")
STORES_DIR = os.path.join(DATA_DIR, "stores")
OUTPUT_PATH = os.path.join(
    DATA_DIR, "exports", "food_order_evaluation_final_0301_0307.xlsx"
)

# ── 설정 ──
ORDER_START = "2026-03-01"
ORDER_END = "2026-03-07"
SALES_END = "2026-03-09"
FOOD_MID_PREFIXES = ("001", "002", "003", "004", "005", "012")
SELLOUT_THRESHOLD_DAYS = 5  # 7일 중 N일 이상 완판 → 수요초과 가능성

# 핵심 평가 매장
CORE_STORE_ID = "46513"  # 이천호반베르디움점
CORE_ORDER_SOURCE = "auto"

# ── 스타일 ──
BLUE_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ORANGE_HEADER = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
PURPLE_HEADER = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
TEAL_HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="맑은 고딕", color="9C0006")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="맑은 고딕", color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="맑은 고딕", color="9C6500")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BLUE_FONT = Font(name="맑은 고딕", color="2F5496")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

NORMAL_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
ITALIC_FONT = Font(name="맑은 고딕", size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SOURCE_LABELS = {"auto": "자동", "site": "사이트", "manual": "수동", "receiving": "입고감지"}


# ────────────────────────────────────────
# 데이터 수집
# ────────────────────────────────────────

def get_store_names():
    conn = sqlite3.connect(COMMON_DB)
    rows = conn.execute("SELECT store_id, store_name FROM stores").fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}


def get_store_dbs():
    store_dbs = []
    for f in os.listdir(STORES_DIR):
        if (f.endswith(".db") and not f.endswith(".db.db")
                and "backup" not in f and "bak" not in f
                and f not in ("store.db", "TEST.db", "99999.db")):
            store_id = f.replace(".db", "")
            store_dbs.append((store_id, os.path.join(STORES_DIR, f)))
    return store_dbs


def determine_delivery(item_nm):
    if not item_nm or not item_nm.strip():
        return None
    last = item_nm.strip()[-1]
    return "1차" if last == "1" else ("2차" if last == "2" else None)


def get_sales_dates(order_date_str, delivery_type):
    od = datetime.strptime(order_date_str, "%Y-%m-%d")
    if delivery_type == "1차":
        return [(od + timedelta(days=1)).strftime("%Y-%m-%d")]
    else:
        return [
            (od + timedelta(days=1)).strftime("%Y-%m-%d"),
            (od + timedelta(days=2)).strftime("%Y-%m-%d"),
        ]


def collect_data():
    """
    Returns (core_records, all_records, buy0_records, source_stats)
    - core_records: 이천호반베르디움점, auto, buy_qty>0 (핵심 평가)
    - all_records: 전매장, 전출처, buy_qty>0 (참고 현황)
    - buy0_records: buy_qty=0 건 (데이터 품질)
    - source_stats: 매장별 order_source 건수
    """
    store_names = get_store_names()
    store_dbs = get_store_dbs()
    all_records = []
    all_buy0_records = []
    all_source_stats = {}

    mid_filter_ot = " OR ".join([f"ot.mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES])
    mid_filter_ds = " OR ".join([f"ds.mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES])

    for store_id, db_path in store_dbs:
        store_name = store_names.get(store_id, f"{store_id}점")
        conn = sqlite3.connect(db_path)
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]
        if "order_tracking" not in tables or "daily_sales" not in tables:
            conn.close()
            continue

        # ── order_source별 건수 집계 (전체 카테고리) ──
        src_rows = conn.execute("""
            SELECT order_source, COUNT(*)
            FROM order_tracking
            WHERE order_date BETWEEN ? AND ?
            GROUP BY order_source
        """, (ORDER_START, ORDER_END)).fetchall()
        store_src = {"store_name": store_name}
        total_all = 0
        for src, cnt in src_rows:
            store_src[src or "unknown"] = cnt
            total_all += cnt
        store_src["total"] = total_all
        all_source_stats[store_id] = store_src

        # ── 발주 데이터 (order_source 포함) ──
        orders = conn.execute(f"""
            SELECT ot.order_date, ot.item_cd, ot.item_nm, ot.mid_cd,
                   ot.delivery_type, ot.order_qty, ot.order_source
            FROM order_tracking ot
            WHERE ot.order_date BETWEEN ? AND ?
              AND ({mid_filter_ot})
            ORDER BY ot.order_date, ot.item_cd
        """, (ORDER_START, ORDER_END)).fetchall()

        # ── 판매/폐기/입고 데이터 로드 ──
        sales_data = {}
        for row in conn.execute(f"""
            SELECT ds.sales_date, ds.item_cd, ds.sale_qty, ds.disuse_qty, ds.buy_qty
            FROM daily_sales ds
            WHERE ds.sales_date BETWEEN ? AND ?
              AND ({mid_filter_ds})
        """, (ORDER_START, SALES_END)).fetchall():
            key = (row[0], row[1])
            if key not in sales_data:
                sales_data[key] = {"sale_qty": row[2] or 0, "disuse_qty": row[3] or 0, "buy_qty": row[4] or 0}
            else:
                sales_data[key]["sale_qty"] = max(sales_data[key]["sale_qty"], row[2] or 0)
                sales_data[key]["disuse_qty"] = max(sales_data[key]["disuse_qty"], row[3] or 0)
                sales_data[key]["buy_qty"] = max(sales_data[key]["buy_qty"], row[4] or 0)

        # ── hourly_sales_detail 로드 ──
        hourly_data = defaultdict(lambda: defaultdict(int))
        if "hourly_sales_detail" in tables:
            for row in conn.execute("""
                SELECT sales_date, item_cd, hour, sale_qty
                FROM hourly_sales_detail
                WHERE sales_date BETWEEN ? AND ?
            """, (ORDER_START, SALES_END)).fetchall():
                hourly_data[(row[0], row[1])][row[2]] += row[3] or 0

        # ── 레코드 생성 ──
        for order_date, item_cd, item_nm, mid_cd, db_delivery, order_qty, order_source in orders:
            delivery = determine_delivery(item_nm)
            if delivery is None:
                delivery = db_delivery if db_delivery in ("1차", "2차") else None
            if delivery is None:
                continue
            order_qty = order_qty or 0
            if order_qty <= 0:
                continue

            sales_dates = get_sales_dates(order_date, delivery)
            sale_qty = 0
            disuse_qty = 0
            buy_qty = 0
            for sd in sales_dates:
                info = sales_data.get((sd, item_cd), {})
                sale_qty += info.get("sale_qty", 0)
                disuse_qty += info.get("disuse_qty", 0)
                buy_qty += info.get("buy_qty", 0)

            # 오전(7~13시) 판매 비중
            morning_qty = 0
            total_hourly = 0
            for sd in sales_dates:
                hd = hourly_data.get((sd, item_cd), {})
                for h, q in hd.items():
                    total_hourly += q
                    if 7 <= h <= 13:
                        morning_qty += q
            morning_ratio = (morning_qty / total_hourly * 100) if total_hourly > 0 else None

            rec = {
                "store_id": store_id,
                "store_name": store_name,
                "order_date": order_date,
                "item_cd": item_cd,
                "item_nm": item_nm,
                "mid_cd": mid_cd,
                "delivery": delivery,
                "order_source": order_source or "unknown",
                "order_qty": order_qty,
                "buy_qty": buy_qty,
                "sale_qty": sale_qty,
                "disuse_qty": disuse_qty,
                "morning_ratio": morning_ratio,
            }

            # buy_qty=0 → 별도 수집
            if buy_qty <= 0:
                rec["sell_rate"] = 0
                rec["sellout"] = 0
                all_buy0_records.append(rec)
                continue

            rec["sellout"] = 1 if (sale_qty >= buy_qty and disuse_qty == 0) else 0
            rec["sell_rate"] = round(sale_qty / buy_qty * 100, 1) if buy_qty > 0 else 0
            all_records.append(rec)

        conn.close()

    # ── 7일 완판 횟수 계산 ──
    sellout_counts = defaultdict(int)
    for rec in all_records:
        if rec["sellout"]:
            sellout_counts[(rec["store_id"], rec["item_cd"])] += 1
    for rec in all_records:
        rec["sellout_7d"] = sellout_counts.get((rec["store_id"], rec["item_cd"]), 0)

    # ── 판정 ──
    for rec in all_records:
        if rec["disuse_qty"] > 0:
            rec["verdict"] = "과발주"
        elif rec["sellout"] and rec["sellout_7d"] >= SELLOUT_THRESHOLD_DAYS:
            rec["verdict"] = "수요초과가능성"
        elif rec["sale_qty"] < rec["buy_qty"] * 0.7:
            rec["verdict"] = "저조"
        else:
            rec["verdict"] = "적정"

    # ── 핵심 평가 분리: 호반베르디움 auto만 ──
    core_records = [
        r for r in all_records
        if r["store_id"] == CORE_STORE_ID and r["order_source"] == CORE_ORDER_SOURCE
    ]

    return core_records, all_records, all_buy0_records, all_source_stats


# ────────────────────────────────────────
# 스타일 헬퍼
# ────────────────────────────────────────

def style_header(ws, row, headers, widths, fill):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 30


def apply_verdict_style(cell, verdict):
    if verdict == "과발주":
        cell.fill = RED_FILL
        cell.font = RED_FONT
    elif verdict == "적정":
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
    elif verdict == "수요초과가능성":
        cell.fill = BLUE_FILL
        cell.font = BLUE_FONT
    elif verdict == "저조":
        cell.fill = YELLOW_FILL
        cell.font = YELLOW_FONT


def write_row(ws, row_idx, values, center_from=1):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if col >= center_from:
            cell.alignment = Alignment(horizontal="center")
    return row_idx + 1


def write_note(ws, row, col, text, merge_to=None):
    """안내 텍스트 작성"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = ITALIC_FONT
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return row + 1


# ────────────────────────────────────────
# 시트 1: 핵심 평가 (이천호반베르디움 auto)
# ────────────────────────────────────────

def write_sheet1_core(wb, core_records):
    ws = wb.active
    ws.title = "핵심 평가"

    # 안내 문구
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1,
                         value="이천호반베르디움점 | 자동발주(auto) | buy_qty > 0 | 2026-03-01 ~ 03-07")
    title_cell.font = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["발주일", "상품명", "차수", "발주량", "입고량",
               "판매량", "폐기량", "완판율(%)", "판정"]
    widths = [12, 38, 7, 9, 9, 9, 9, 11, 14]
    style_header(ws, 2, headers, widths, BLUE_HEADER)

    core_sorted = sorted(core_records, key=lambda r: (r["order_date"], r["item_nm"]))

    row = 3
    for rec in core_sorted:
        vals = [rec["order_date"], rec["item_nm"], rec["delivery"],
                rec["order_qty"], rec["buy_qty"], rec["sale_qty"],
                rec["disuse_qty"], rec["sell_rate"], rec["verdict"]]
        write_row(ws, row, vals, center_from=3)
        apply_verdict_style(ws.cell(row=row, column=9), rec["verdict"])
        row += 1

    # 요약
    row += 1
    verdict_counts = defaultdict(int)
    for r in core_records:
        verdict_counts[r["verdict"]] += 1
    total_buy = sum(r["buy_qty"] for r in core_records)
    total_disuse = sum(r["disuse_qty"] for r in core_records)
    waste_rate = round(total_disuse / total_buy * 100, 1) if total_buy > 0 else 0

    ws.cell(row=row, column=1,
            value=f"총 {len(core_records)}건 | 폐기율 {waste_rate}% "
                  f"(과발주 {verdict_counts['과발주']}건, 적정 {verdict_counts['적정']}건, "
                  f"수요초과 {verdict_counts['수요초과가능성']}건, 저조 {verdict_counts['저조']}건)").font = \
        Font(name="맑은 고딕", bold=True, size=11)

    ws.auto_filter.ref = f"A2:I{max(row - 1, 3)}"
    ws.freeze_panes = "A3"


# ────────────────────────────────────────
# 시트 2: 과발주 상품 (폐기 발생) — 호반베르디움 auto만
# ────────────────────────────────────────

def write_sheet2_overorder(wb, core_records):
    ws = wb.create_sheet("과발주 상품")
    headers = ["상품명", "폐기발생횟수", "총폐기량", "평균입고량",
               "평균판매량", "폐기율(%)", "원인추정"]
    widths = [38, 13, 11, 11, 11, 11, 16]
    style_header(ws, 1, headers, widths, ORANGE_HEADER)

    waste_stats = defaultdict(lambda: {
        "item_nm": "", "count": 0, "total_disuse": 0,
        "total_buy": 0, "total_sale": 0, "days": 0,
    })
    for rec in core_records:
        if rec["disuse_qty"] > 0:
            key = rec["item_cd"]
            ws_item = waste_stats[key]
            ws_item["item_nm"] = rec["item_nm"]
            ws_item["count"] += 1
            ws_item["total_disuse"] += rec["disuse_qty"]
            ws_item["total_buy"] += rec["buy_qty"]
            ws_item["total_sale"] += rec["sale_qty"]
            ws_item["days"] += 1

    sorted_waste = sorted(waste_stats.items(),
                          key=lambda x: x[1]["total_disuse"], reverse=True)

    row = 2
    for icd, st in sorted_waste:
        avg_buy = round(st["total_buy"] / st["days"], 1)
        avg_sale = round(st["total_sale"] / st["days"], 1)
        denom = st["total_sale"] + st["total_disuse"]
        waste_rate = round(st["total_disuse"] / denom * 100, 1) if denom > 0 else 0

        if waste_rate > 30:
            cause = "과발주"
        elif waste_rate >= 10:
            cause = "발주 조정 검토"
        else:
            cause = "허용 범위"

        vals = [st["item_nm"], st["count"], st["total_disuse"],
                avg_buy, avg_sale, waste_rate, cause]
        write_row(ws, row, vals, center_from=2)

        cause_cell = ws.cell(row=row, column=7)
        if cause == "과발주":
            cause_cell.fill = RED_FILL
            cause_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")
        elif cause == "발주 조정 검토":
            cause_cell.fill = YELLOW_FILL
            cause_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
        else:
            cause_cell.fill = GREEN_FILL
            cause_cell.font = Font(name="맑은 고딕", bold=True, color="006100")
        row += 1

    row += 1
    high_waste_cnt = sum(1 for _, s in sorted_waste
                         if s["total_disuse"] / max(s["total_sale"] + s["total_disuse"], 1) * 100 > 30)
    ws.cell(row=row, column=1,
            value=f"과발주 상품 총 {len(sorted_waste)}건 (폐기율 30%+: {high_waste_cnt}건)").font = \
        Font(name="맑은 고딕", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1,
            value="폐기율 = 폐기량 / (판매량 + 폐기량) x 100").font = ITALIC_FONT

    ws.auto_filter.ref = f"A1:G{max(row - 2, 2)}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 3: 수요초과 가능성 상품 — 호반베르디움 auto만
# ────────────────────────────────────────

def write_sheet3_excess(wb, core_records):
    ws = wb.create_sheet("수요초과 가능성")
    headers = ["상품명", "완판횟수/7일", "평균입고량", "오전판매비중(7~13시, %)"]
    widths = [38, 14, 13, 22]
    style_header(ws, 1, headers, widths, PURPLE_HEADER)

    # 상품별 집계 (핵심 평가 데이터에서)
    item_stats = defaultdict(lambda: {
        "item_nm": "", "sellout_days": 0, "total_days": 0,
        "total_buy": 0, "morning_ratios": [],
    })
    for rec in core_records:
        key = rec["item_cd"]
        ist = item_stats[key]
        ist["item_nm"] = rec["item_nm"]
        ist["total_days"] += 1
        ist["total_buy"] += rec["buy_qty"]
        if rec["sellout"]:
            ist["sellout_days"] += 1
        if rec["morning_ratio"] is not None:
            ist["morning_ratios"].append(rec["morning_ratio"])

    # 7일 중 5일+ 완판 AND 폐기=0 상품
    excess_items = []
    for icd, ist in item_stats.items():
        if ist["sellout_days"] >= SELLOUT_THRESHOLD_DAYS:
            # 폐기 발생 여부 확인
            has_disuse = any(r["item_cd"] == icd and r["disuse_qty"] > 0 for r in core_records)
            if not has_disuse:
                excess_items.append((icd, ist))

    excess_items.sort(key=lambda x: x[1]["sellout_days"], reverse=True)

    row = 2
    for icd, ist in excess_items:
        avg_buy = round(ist["total_buy"] / ist["total_days"], 1) if ist["total_days"] > 0 else 0
        avg_morning = (round(sum(ist["morning_ratios"]) / len(ist["morning_ratios"]), 1)
                       if ist["morning_ratios"] else None)
        morning_str = f"{avg_morning}%" if avg_morning is not None else "데이터없음"

        vals = [ist["item_nm"],
                f"{ist['sellout_days']}/{ist['total_days']}",
                avg_buy, morning_str]
        write_row(ws, row, vals, center_from=2)

        if avg_morning is not None and avg_morning >= 60:
            mc = ws.cell(row=row, column=4)
            mc.fill = YELLOW_FILL
            mc.font = YELLOW_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value=f"수요초과 가능성 상품: {len(excess_items)}개").font = \
        Font(name="맑은 고딕", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1,
            value=f"기준: 7일 중 {SELLOUT_THRESHOLD_DAYS}일 이상 완판 + 폐기 0").font = ITALIC_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="오전비중 60%+ → 1차 배송 증량 검토 필요").font = ITALIC_FONT

    ws.auto_filter.ref = f"A1:D{max(row - 3, 2)}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 4: 전매장 참고 현황 (신뢰도 명시)
# ────────────────────────────────────────

def write_sheet4_all_stores(wb, all_records, buy0_records, source_stats):
    ws = wb.create_sheet("전매장 참고 현황")

    # 안내 문구
    ws.merge_cells("A1:J1")
    note_cell = ws.cell(row=1, column=1,
                        value="참고 자료: 매장별 auto 비율에 따라 신뢰도 차이가 있습니다")
    note_cell.font = Font(name="맑은 고딕", bold=True, size=11, color="9C6500")
    note_cell.fill = YELLOW_FILL
    note_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["매장명", "auto비율(%)", "평가 신뢰도", "건수",
               "총입고량", "총판매량", "총폐기량", "폐기율(%)",
               "과발주", "적정", "비고"]
    widths = [20, 12, 12, 8, 11, 11, 11, 11, 9, 9, 40]
    style_header(ws, 2, headers, widths, GREEN_HEADER)

    # 매장별 집계
    store_agg = defaultdict(lambda: {
        "name": "", "count": 0, "buy": 0, "sale": 0, "disuse": 0,
        "over": 0, "ok": 0, "low": 0, "excess": 0,
    })
    for rec in all_records:
        sid = rec["store_id"]
        sa = store_agg[sid]
        sa["name"] = rec["store_name"]
        sa["count"] += 1
        sa["buy"] += rec["buy_qty"]
        sa["sale"] += rec["sale_qty"]
        sa["disuse"] += rec["disuse_qty"]
        v = rec["verdict"]
        if v == "과발주": sa["over"] += 1
        elif v == "적정": sa["ok"] += 1
        elif v == "저조": sa["low"] += 1
        elif v == "수요초과가능성": sa["excess"] += 1

    # buy_qty=0 건수 매장별
    buy0_by_store = defaultdict(int)
    buy0_sale_by_store = defaultdict(int)
    for rec in buy0_records:
        buy0_by_store[rec["store_id"]] += 1
        if rec.get("sale_qty", 0) > 0:
            buy0_sale_by_store[rec["store_id"]] += 1

    store_ids = sorted(store_agg.keys())

    row = 3
    for sid in store_ids:
        ss = source_stats.get(sid, {})
        auto_cnt = ss.get("auto", 0)
        total_cnt = ss.get("total", 1)
        auto_pct = round(auto_cnt / total_cnt * 100, 1) if total_cnt > 0 else 0

        if auto_pct >= 80:
            reliability = "높음"
        elif auto_pct >= 50:
            reliability = "중간"
        else:
            reliability = "낮음"

        sa = store_agg[sid]
        wr = round(sa["disuse"] / sa["buy"] * 100, 1) if sa["buy"] > 0 else 0

        # 비고 생성
        notes = []
        if sid == CORE_STORE_ID:
            notes.append("핵심 평가 대상")
        if reliability == "낮음":
            notes.append(f"auto 비율 낮음({auto_pct}%) - site/manual 발주 중심")
        elif reliability == "중간":
            notes.append(f"auto+site 혼합 발주({auto_pct}%)")

        b0 = buy0_by_store.get(sid, 0)
        b0s = buy0_sale_by_store.get(sid, 0)
        if b0 > 0:
            notes.append(f"buy_qty=0 {b0}건 (판매발생 {b0s}건)")

        if sa["excess"] > 0:
            notes.append(f"수요초과 {sa['excess']}건")
        if sa["low"] > 0:
            notes.append(f"저조 {sa['low']}건")

        note_str = " | ".join(notes) if notes else ""

        vals = [sa["name"], auto_pct, reliability, sa["count"],
                sa["buy"], sa["sale"], sa["disuse"], wr,
                sa["over"], sa["ok"], note_str]
        write_row(ws, row, vals, center_from=2)
        # 비고는 왼쪽 정렬
        ws.cell(row=row, column=11).alignment = Alignment(horizontal="left", wrap_text=True)

        # 신뢰도 색상
        rel_cell = ws.cell(row=row, column=3)
        if reliability == "높음":
            rel_cell.fill = GREEN_FILL
            rel_cell.font = Font(name="맑은 고딕", bold=True, color="006100")
        elif reliability == "중간":
            rel_cell.fill = YELLOW_FILL
            rel_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
        else:
            rel_cell.fill = RED_FILL
            rel_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")

        # 핵심 평가 대상 행 강조
        if sid == CORE_STORE_ID:
            for c in range(1, 12):
                cell = ws.cell(row=row, column=c)
                if c != 3:  # 신뢰도 셀은 이미 색칠
                    cell.fill = LIGHT_BLUE_FILL
                cell.font = Font(name="맑은 고딕", size=10, bold=True)
                cell.border = THIN_BORDER
        row += 1

    # 전체 합계 행
    row += 1
    total_count = sum(sa["count"] for sa in store_agg.values())
    total_buy = sum(sa["buy"] for sa in store_agg.values())
    total_sale = sum(sa["sale"] for sa in store_agg.values())
    total_disuse = sum(sa["disuse"] for sa in store_agg.values())
    total_wr = round(total_disuse / total_buy * 100, 1) if total_buy > 0 else 0
    total_over = sum(sa["over"] for sa in store_agg.values())
    total_ok = sum(sa["ok"] for sa in store_agg.values())

    total_vals = ["합계", "", "", total_count,
                  total_buy, total_sale, total_disuse, total_wr,
                  total_over, total_ok, ""]
    write_row(ws, row, total_vals, center_from=2)
    for c in range(1, 12):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_GRAY_FILL
        cell.border = THIN_BORDER

    # 범례
    row += 2
    ws.cell(row=row, column=1,
            value="auto비율 기준: 전체 order_tracking(전카테고리) 중 auto 건수 비율").font = ITALIC_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="신뢰도: 높음(80%+) = bgf_auto 단독 평가 가능 | 중간(50~79%) = 참고 수준 | 낮음(<50%) = 제한적").font = ITALIC_FONT

    ws.auto_filter.ref = f"A2:K{max(row - 3, 3)}"
    ws.freeze_panes = "A3"


# ────────────────────────────────────────
# 시트 5: 데이터 품질 이슈 기록
# ────────────────────────────────────────

def write_sheet5_data_quality(wb, buy0_records, all_records, source_stats):
    ws = wb.create_sheet("데이터 품질 이슈")

    # ── 섹션 A: buy_qty 미기록 목록 ──
    ws.merge_cells("A1:H1")
    ws.cell(row=1, column=1,
            value="A. buy_qty=0 이지만 판매 발생 (입고 미기록)").font = \
        Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
    ws.row_dimensions[1].height = 26

    headers_a = ["매장명", "발주일", "상품명", "발주출처", "차수",
                 "발주량", "입고량(0)", "판매량"]
    widths_a = [20, 12, 38, 10, 7, 9, 10, 9]
    style_header(ws, 2, headers_a, widths_a, TEAL_HEADER)

    # sale_qty > 0 인 케이스만
    sale_b0 = [r for r in buy0_records if r.get("sale_qty", 0) > 0]
    sale_b0.sort(key=lambda r: (r["store_name"], r["order_date"], r["item_nm"]))

    row = 3
    for rec in sale_b0:
        src_label = SOURCE_LABELS.get(rec["order_source"], rec["order_source"])
        vals = [rec["store_name"], rec["order_date"], rec["item_nm"],
                src_label, rec["delivery"], rec["order_qty"], 0, rec["sale_qty"]]
        write_row(ws, row, vals, center_from=4)
        # 판매량 셀 강조
        ws.cell(row=row, column=8).fill = YELLOW_FILL
        ws.cell(row=row, column=8).font = YELLOW_FONT
        row += 1

    # 요약
    sale_sum = sum(r.get("sale_qty", 0) for r in sale_b0)
    total_main_sale = sum(r["sale_qty"] for r in all_records)
    pct = round(sale_sum / (total_main_sale + sale_sum) * 100, 1) if (total_main_sale + sale_sum) > 0 else 0

    row += 1
    ws.cell(row=row, column=1,
            value=f"입고 미기록 + 판매 발생: {len(sale_b0)}건, "
                  f"판매량 합계: {sale_sum}개, "
                  f"전체 판매 대비 {pct}%").font = BOLD_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="원인: BGF 수집 시점에 입고 미확정 또는 데이터 타이밍 이슈").font = ITALIC_FONT

    # ── 섹션 B: actual_receiving_qty NULL 현황 ──
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1,
            value="B. actual_receiving_qty 컬럼 NULL 현황").font = \
        Font(name="맑은 고딕", bold=True, size=12, color="2F5496")

    row += 1
    headers_b = ["매장명", "전체 order_tracking 건수", "actual_receiving_qty NULL 건수",
                 "NULL 비율(%)", "보정 가능 여부"]
    widths_b = [20, 22, 26, 14, 16]
    for col, (h, w) in enumerate(zip(headers_b, widths_b), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = TEAL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0, w)

    row += 1
    store_names = get_store_names()
    for store_id, db_path in get_store_dbs():
        sname = store_names.get(store_id, f"{store_id}점")
        conn = sqlite3.connect(db_path)
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        if "order_tracking" not in tables:
            conn.close()
            continue
        # 전체 건수
        total = conn.execute("""
            SELECT COUNT(*) FROM order_tracking
            WHERE order_date BETWEEN ? AND ?
        """, (ORDER_START, ORDER_END)).fetchone()[0]
        # NULL 건수
        null_cnt = conn.execute("""
            SELECT COUNT(*) FROM order_tracking
            WHERE order_date BETWEEN ? AND ?
              AND actual_receiving_qty IS NULL
        """, (ORDER_START, ORDER_END)).fetchone()[0]
        conn.close()

        null_pct = round(null_cnt / total * 100, 1) if total > 0 else 0
        can_fix = "불가" if null_pct >= 99 else "부분 가능"

        vals = [sname, total, null_cnt, null_pct, can_fix]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        fix_cell = ws.cell(row=row, column=5)
        if can_fix == "불가":
            fix_cell.fill = RED_FILL
            fix_cell.font = RED_FONT
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value="actual_receiving_qty가 100% NULL이므로 buy_qty=0 건의 보정 불가").font = BOLD_FONT

    # ── 섹션 C: 개선 과제 ──
    row += 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.cell(row=row, column=1,
            value="C. 다음 개선 과제").font = \
        Font(name="맑은 고딕", bold=True, size=12, color="2F5496")

    improvements = [
        ("1", "buy_qty 수집 개선",
         "daily_sales 수집 시 입고 확정 시점 이후에 buy_qty 재수집 또는 order_tracking status=arrived 기록 활용"),
        ("2", "actual_receiving_qty 채우기",
         "센터매입조회 입고확정 데이터를 order_tracking.actual_receiving_qty에 기록하여 buy_qty 보정 가능하게"),
        ("3", "site/manual 발주 추적 강화",
         "BGF 사이트 직접 발주(site) 또는 수동 발주(manual) 건의 결과를 추적하여 전체 평가 신뢰도 향상"),
        ("4", "이중발주 방지",
         "auto 발주 후 site/manual로 동일 상품 추가 발주 시 알림 또는 차단 로직 추가"),
        ("5", "폐기 전표 대조",
         "검수전표 폐기 데이터와 daily_sales disuse_qty 대조를 통한 폐기량 정확도 검증"),
    ]

    row += 1
    headers_c = ["#", "과제", "상세 내용"]
    widths_c = [5, 24, 65]
    for col, (h, w) in enumerate(zip(headers_c, widths_c), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = TEAL_HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = max(
            ws.column_dimensions[get_column_letter(col)].width or 0, w)

    row += 1
    for num, title, detail in improvements:
        ws.cell(row=row, column=1, value=num).font = NORMAL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=title).font = BOLD_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3, value=detail).font = NORMAL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 30
        row += 1

    ws.freeze_panes = "A3"


# ────────────────────────────────────────
# 메인
# ────────────────────────────────────────

def main():
    print("=" * 60)
    print("푸드 발주 평가 최종본 리포트 생성")
    print(f"기간: {ORDER_START} ~ {ORDER_END}")
    print(f"핵심 평가: 이천호반베르디움점(auto, buy_qty>0)")
    print("=" * 60)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    core_records, all_records, buy0_records, source_stats = collect_data()

    print(f"\n[데이터 수집 결과]")
    print(f"  전체 레코드 (buy_qty>0): {len(all_records)}건")
    print(f"  핵심 평가 (호반베르디움 auto): {len(core_records)}건")
    print(f"  buy_qty=0 제외 건: {len(buy0_records)}건 "
          f"(판매발생: {sum(1 for r in buy0_records if r.get('sale_qty',0)>0)}건)")

    if not core_records:
        print("핵심 평가 데이터 없음")
        return

    # 핵심 평가 통계
    verdict_counts = defaultdict(int)
    for r in core_records:
        verdict_counts[r["verdict"]] += 1
    total_buy = sum(r["buy_qty"] for r in core_records)
    total_disuse = sum(r["disuse_qty"] for r in core_records)
    waste_rate = round(total_disuse / total_buy * 100, 1) if total_buy > 0 else 0

    print(f"\n[핵심 평가 통계]")
    print(f"  건수: {len(core_records)}")
    print(f"  폐기율: {waste_rate}%")
    for v in ["과발주", "적정", "저조", "수요초과가능성"]:
        print(f"  {v}: {verdict_counts[v]}건")

    # 매장별 auto 비율
    print(f"\n[매장별 auto 비율]")
    for sid, ss in sorted(source_stats.items()):
        if ss.get("total", 0) == 0:
            continue
        apct = round(ss.get("auto", 0) / ss["total"] * 100, 1)
        rel = "높음" if apct >= 80 else ("중간" if apct >= 50 else "낮음")
        print(f"  {ss['store_name']}: {apct}% ({rel})")

    # 엑셀 생성
    wb = Workbook()
    write_sheet1_core(wb, core_records)
    write_sheet2_overorder(wb, core_records)
    write_sheet3_excess(wb, core_records)
    write_sheet4_all_stores(wb, all_records, buy0_records, source_stats)
    write_sheet5_data_quality(wb, buy0_records, all_records, source_stats)

    wb.save(OUTPUT_PATH)
    print(f"\n리포트 저장 완료: {OUTPUT_PATH}")
    print(f"시트 구성: {wb.sheetnames}")


if __name__ == "__main__":
    main()
