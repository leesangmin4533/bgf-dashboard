"""
전체 자동화 플로우 실행
1. BGF 로그인
2. 판매 데이터 수집
3. DB 현황 확인
4. 발주 실행 (DailyOrderFlow 위임)
5. 드라이런 Excel 리포트 (30컬럼 상세)

오프라인 모드 (--export-excel):
  브라우저 없이 DB 기반 예측 → 엑셀 내보내기
"""

import sys
import io
import math
from pathlib import Path
from collections import defaultdict, Counter

# Windows CP949 콘솔 -> UTF-8 래핑 (한글/특수문자 깨짐 방지)
if sys.platform == "win32":
    for _sn in ("stdout", "stderr"):
        _s = getattr(sys, _sn, None)
        if _s and hasattr(_s, "buffer") and (getattr(_s, "encoding", "") or "").lower().replace("-", "") != "utf8":
            setattr(sys, _sn, io.TextIOWrapper(_s.buffer, encoding="utf-8", errors="replace", line_buffering=True))

# 프로젝트 루트를 path에 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# 로그 정리 (30일 초과 삭제, 50MB 초과 잘라내기)
from src.utils.logger import cleanup_old_logs
cleanup_old_logs(max_age_days=30, max_file_mb=50)

import time
from datetime import datetime
from typing import Any, Dict, List, Optional

from src.sales_analyzer import SalesAnalyzer
from src.infrastructure.database.repos import SalesRepository
from src.collectors.sales_collector import SalesCollector
from src.settings.constants import DEFAULT_STORE_ID

# ─── 드라이런 데이터 신선도 기준 ───
DRYRUN_STALE_HOURS = {
    "food": 6,      # 001~005, 012, 014: 6시간
    "default": 24,   # 기타: 24시간
}
FOOD_MID_CDS = {"001", "002", "003", "004", "005", "012", "014"}


# ═══════════════════════════════════════════════════════════════════════
# Excel 리포트: 섹션별 컬럼 정의 (5개 섹션, 총 30컬럼 A~AD)
# ═══════════════════════════════════════════════════════════════════════

SECTION_A = {
    "name": "A. 기본정보",
    "columns": [
        ("No",          "no"),
        ("상품코드",    "item_cd"),
        ("상품명",      "item_nm"),
        ("중분류",      "mid_cd"),
        ("수요패턴",    "demand_pattern"),
        ("데이터일수",  "data_days"),
        ("판매일비율",  "sell_day_ratio"),
    ],
}

SECTION_B = {
    "name": "B. 예측단계",
    "columns": [
        ("WMA(원본)",    "wma_raw"),
        ("Feature예측",  "feat_prediction"),
        ("블렌딩결과",   "predicted_qty"),
        ("요일계수",     "weekday_coef"),
        ("조정예측",     "adjusted_qty"),
    ],
}

SECTION_C = {
    "name": "C. 재고/필요량",
    "columns": [
        ("현재재고",    "current_stock"),
        ("미입고",      "pending_receiving_qty"),
        ("안전재고",    "safety_stock"),
        ("필요량",      "need_qty"),
        ("RI조회시각",  "ri_queried_at"),
    ],
}

SECTION_D = {
    "name": "D. 조정과정",
    "columns": [
        ("Rule발주",    "rule_order_qty"),
        ("ML예측",      "ml_order_qty"),
        ("ML가중치",    "ml_weight_used"),
        ("ML후발주",    "final_order_qty"),
        ("조정이력",    "proposal_summary"),
    ],
}

SECTION_E = {
    "name": "E. 배수정렬+BGF입력",
    "columns": [
        ("정렬전수량",      "round_before"),
        ("내림후보",        "round_floor"),
        ("올림후보",        "round_ceil"),
        ("정렬결과",        "final_order_qty"),
        ("발주단위(입수)",  "order_unit_qty"),
        ("PYUN_QTY(배수)",  "pyun_qty"),
        ("TOT_QTY(발주량)", "tot_qty"),
        ("모델타입",        "model_type"),
    ],
}

ALL_SECTIONS = [SECTION_A, SECTION_B, SECTION_C, SECTION_D, SECTION_E]

# ─── 섹션별 색상: 1~2행(헤더) / 3행(설명) ───
SECTION_STYLES = [
    {"header": "BDD7EE", "desc": "DEEAF1"},  # A. 기본정보 (파랑 계열)
    {"header": "C6EFCE", "desc": "E2F0D9"},  # B. 예측단계 (초록 계열)
    {"header": "FFE699", "desc": "FFF2CC"},  # C. 재고/필요량 (주황 계열)
    {"header": "F4B8D1", "desc": "FCE4D6"},  # D. 조정과정 (분홍 계열)
    {"header": "D9D9D9", "desc": "EDEDED"},  # E. 배수정렬+BGF입력 (회색 계열)
]

# ─── 컬럼별 설명 (3행, 30개) ───
COLUMN_DESCRIPTIONS = [
    "순번", "바코드 번호", "상품 이름", "카테고리 코드",
    "판매 빈도 유형", "학습에 사용된 날수", "실제 판매된 날 비율",
    "가중평균 판매량(보정 전)", "ML 특성 기반 예측값",
    "WMA+ML 혼합 예측값", "오늘 요일 판매 보정치", "요일 반영 최종 예측값",
    "지금 매장에 있는 수량", "발주했지만 아직 안 온 수량",
    "최소 유지해야 할 수량", "실제로 더 필요한 수량",
    "재고 데이터 조회 시각",
    "규칙 기반 발주 수량", "ML 모델 발주 제안량",
    "ML 의견 반영 비율", "ML 반영 후 발주량", "발주 결정 과정 로그",
    "배수 맞추기 전 수량", "내림 적용 시 수량", "올림 적용 시 수량",
    "최종 배수 정렬 수량", "한 박스에 들어가는 개수",
    "박스 단위 수량", "최종 발주 수량", "사용된 예측 모델 종류",
]

# ─── 열 너비 (A~AD, 30개) ───
COL_WIDTHS = [
    5, 16, 22, 8, 10, 10, 10,           # A~G  (기본정보)
    11, 11, 11, 11, 11,                  # H~L  (예측단계)
    10, 10, 10, 10, 16,                  # M~Q  (재고/필요량 + RI조회시각)
    10, 8, 10, 10, 45,                   # R~V  (조정과정)
    13, 13, 13, 13, 13, 13, 13, 13,      # W~AD (배수정렬+BGF입력)
]

# ─── 숫자 포맷 컬럼 세트 ───
# 소수점 2자리: H~T (cols 8~20) — Q열(17)은 RI조회시각(텍스트)이므로 제외하지 않음
#   Q열은 텍스트이므로 FLOAT에서 자동 무시됨 (값이 문자열)
# V열(23): round_before
FLOAT_COLS = set(range(8, 21)) | {23}
# 정수: X~AC (cols 24~29)
INT_COLS = set(range(24, 30))

TOTAL_COLS = 30  # A~AD


# ═══════════════════════════════════════════════════════════════════════
# Excel 헬퍼 함수
# ═══════════════════════════════════════════════════════════════════════

def _safe_get(item: dict, key: str, default=0):
    """None-safe getter"""
    v = item.get(key, default)
    return default if v is None else v


def _compute_bgf_fields(item: dict) -> dict:
    """PYUN_QTY / TOT_QTY 계산 + round_before"""
    final = _safe_get(item, "final_order_qty", 0)
    unit = _safe_get(item, "order_unit_qty", 1) or 1

    if final > 0 and unit > 1:
        pyun = math.ceil(final / unit)
        tot = pyun * unit
    elif final > 0:
        pyun = final
        tot = final
    else:
        pyun = 0
        tot = 0

    r_floor = _safe_get(item, "round_floor", 0)
    r_ceil = _safe_get(item, "round_ceil", 0)
    if r_floor > 0 or r_ceil > 0:
        if final == r_floor:
            round_before = r_floor
        elif final == r_ceil:
            round_before = r_ceil
        else:
            round_before = final
    else:
        round_before = final

    return {"pyun_qty": pyun, "tot_qty": tot, "round_before": round_before}


def _get_cell_value(item: dict, key: str, computed: dict):
    """컬럼 key에 따른 셀 값 반환"""
    if key in computed:
        return computed[key]
    if key == "predicted_qty":
        return _safe_get(item, "daily_avg", 0)
    if key == "adjusted_qty":
        return _safe_get(item, "predicted_sales", 0)
    return _safe_get(item, key, "")


# ═══════════════════════════════════════════════════════════════════════
# 요약 시트 생성
# ═══════════════════════════════════════════════════════════════════════

def _create_summary_sheet(wb, order_list: List[Dict[str, Any]]):
    """요약 시트: 중분류별 / 수요패턴별 / 모델타입별 현황 3개 표"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("요약")

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    h_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    h_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    d_font = Font(name="맑은 고딕", size=10)
    b_font = Font(name="맑은 고딕", bold=True, size=10)
    ca = Alignment(horizontal="center", vertical="center")
    even_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    def _hdr(row, cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font, c.fill, c.alignment, c.border = h_font, h_fill, ca, thin

    def _row(row, vals, bold=False):
        is_even = row % 2 == 0
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = b_font if bold else d_font
            c.alignment = ca
            c.border = thin
            if is_even and not bold:
                c.fill = even_fill
            if isinstance(v, float):
                c.number_format = "0.00"

    # ────── 표 1: 중분류별 발주 현황 ──────
    _hdr(1, ["중분류", "상품수", "총발주량", "평균발주량", "발주0건수", "주요모델"])

    mid_grp = defaultdict(list)
    for it in order_list:
        mid_grp[_safe_get(it, "mid_cd", "?")].append(it)

    r = 2
    g_items, g_qty, g_zero = 0, 0, 0
    for mid in sorted(mid_grp):
        items = mid_grp[mid]
        cnt = len(items)
        qty = sum(_compute_bgf_fields(it)["tot_qty"] for it in items)
        avg = round(qty / cnt, 2) if cnt else 0
        z = sum(1 for it in items if _compute_bgf_fields(it)["tot_qty"] == 0)
        mc = Counter(_safe_get(it, "model_type", "?") for it in items)
        top_model = mc.most_common(1)[0][0] if mc else ""
        _row(r, [mid, cnt, qty, avg, z, top_model])
        g_items += cnt
        g_qty += qty
        g_zero += z
        r += 1

    _row(r, ["합계", g_items, g_qty,
             round(g_qty / g_items, 2) if g_items else 0, g_zero, ""], bold=True)
    r += 1

    # ────── 표 2: 수요패턴별 현황 (2행 간격) ──────
    r += 2
    _hdr(r, ["수요패턴", "상품수", "총발주량", "평균ML가중치"])
    r += 1

    pat_grp = defaultdict(list)
    for it in order_list:
        pat_grp[_safe_get(it, "demand_pattern", "미분류") or "미분류"].append(it)

    for pat in sorted(pat_grp):
        items = pat_grp[pat]
        cnt = len(items)
        qty = sum(_compute_bgf_fields(it)["tot_qty"] for it in items)
        wts = [_safe_get(it, "ml_weight_used", 0) or 0 for it in items]
        avg_w = round(sum(wts) / cnt, 4) if cnt else 0
        _row(r, [pat, cnt, qty, avg_w])
        r += 1

    # ────── 표 3: 모델타입별 현황 (2행 간격) ──────
    r += 2
    _hdr(r, ["모델타입", "상품수", "비율(%)"])
    r += 1

    total_n = len(order_list)
    mc = Counter(_safe_get(it, "model_type", "?") for it in order_list)
    for model, cnt in mc.most_common():
        pct = round(cnt / total_n * 100, 1) if total_n else 0
        _row(r, [model, cnt, pct])
        r += 1

    # 열 너비 자동 조정 (한글 2배 너비)
    for ci in range(1, 7):
        mx = 8
        for ri in range(1, r):
            v = ws.cell(row=ri, column=ci).value
            if v is not None:
                sl = sum(2 if ord(ch) > 127 else 1 for ch in str(v))
                mx = max(mx, sl + 2)
        ws.column_dimensions[get_column_letter(ci)].width = min(mx, 30)


# ═══════════════════════════════════════════════════════════════════════
# 메인 Excel 생성 (3행 헤더 + 조건부 서식 + 요약 시트)
# ═══════════════════════════════════════════════════════════════════════

def create_dryrun_excel(
    order_list: List[Dict[str, Any]],
    output_path: Optional[str] = None,
    delivery_date: Optional[str] = None,
    store_id: Optional[str] = None,
) -> str:
    """order_list → 5섹션 3행헤더 컬러구분 엑셀 + 요약 시트.

    헤더 구조:
        1행: 섹션명 (병합)
        2행: 컬럼명 (AutoFilter)
        3행: 컬럼 설명 (이탤릭 회색)
        4행~: 데이터
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not delivery_date:
        delivery_date = datetime.now().strftime("%Y-%m-%d")

    if not output_path:
        export_dir = project_root / "data" / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        suffix = f"_{store_id}" if store_id else ""
        output_path = str(export_dir / f"dryrun_detail_{delivery_date}{suffix}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = f"발주상세_{delivery_date}"

    # ─── 공통 스타일 ───
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    ca = Alignment(horizontal="center", vertical="center")

    # ─── 섹션별 컬럼 범위 계산 ───
    col_keys = []           # [(key, section_index), ...]
    section_col_ranges = [] # [(start_col, end_col, section_index), ...]
    ci = 1
    for si, sec in enumerate(ALL_SECTIONS):
        n = len(sec["columns"])
        section_col_ranges.append((ci, ci + n - 1, si))
        for _col_name, key in sec["columns"]:
            col_keys.append((key, si))
        ci += n

    def _section_for_col(col_num):
        for s, e, si in section_col_ranges:
            if s <= col_num <= e:
                return si
        return 0

    # ─── 1행: 섹션명 (병합) ───
    for start, end, si in section_col_ranges:
        styles = SECTION_STYLES[si]
        fill = PatternFill(start_color=styles["header"],
                           end_color=styles["header"], fill_type="solid")
        cell = ws.cell(row=1, column=start, value=ALL_SECTIONS[si]["name"])
        cell.font = Font(name="맑은 고딕", bold=True, size=10)
        cell.fill = fill
        cell.alignment = ca
        cell.border = thin
        if end > start:
            ws.merge_cells(start_row=1, start_column=start,
                           end_row=1, end_column=end)
            for c in range(start + 1, end + 1):
                mc = ws.cell(row=1, column=c)
                mc.fill = fill
                mc.border = thin
    ws.row_dimensions[1].height = 18

    # ─── 2행: 컬럼명 ───
    ci = 1
    for si, sec in enumerate(ALL_SECTIONS):
        styles = SECTION_STYLES[si]
        fill = PatternFill(start_color=styles["header"],
                           end_color=styles["header"], fill_type="solid")
        for col_name, _key in sec["columns"]:
            cell = ws.cell(row=2, column=ci, value=col_name)
            cell.font = Font(name="맑은 고딕", bold=True, size=10)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = thin
            ci += 1
    ws.row_dimensions[2].height = 20

    # ─── 3행: 설명행 ───
    for idx, desc in enumerate(COLUMN_DESCRIPTIONS):
        col_num = idx + 1
        si = _section_for_col(col_num)
        styles = SECTION_STYLES[si]
        fill = PatternFill(start_color=styles["desc"],
                           end_color=styles["desc"], fill_type="solid")
        cell = ws.cell(row=3, column=col_num, value=desc)
        cell.font = Font(name="맑은 고딕", size=8, italic=True, color="888888")
        cell.fill = fill
        cell.alignment = ca
        cell.border = thin
    ws.row_dimensions[3].height = 14

    # ─── 열 너비 ───
    for i, w in enumerate(COL_WIDTHS):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # ─── Freeze Panes ───
    ws.freeze_panes = "A4"

    # ─── AutoFilter ───
    ws.auto_filter.ref = "A2:AD2"

    # ─── 데이터 행: 4행부터 ───
    DATA_START = 4
    for row_i, item in enumerate(order_list, start=1):
        excel_row = DATA_START + row_i - 1
        item["no"] = row_i
        computed = _compute_bgf_fields(item)

        ci = 1
        for key, _si in col_keys:
            value = _get_cell_value(item, key, computed)
            cell = ws.cell(row=excel_row, column=ci, value=value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = thin
            cell.alignment = ca

            # 숫자 포맷
            if ci in FLOAT_COLS:
                cell.number_format = "0.00"
            elif ci in INT_COLS:
                cell.number_format = "0"

            # V열(22): 조정이력 — 왼쪽정렬, 줄바꿈 없음 (Q열 추가로 +1)
            if ci == 22:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=False)

            ci += 1

        ws.row_dimensions[excel_row].height = 15

    data_end = DATA_START + len(order_list) - 1

    # ─── 조건부 서식 — 셀 순회 ───
    COL_S = 20   # ML가중치 (Q열 추가로 +1)
    COL_AB = 29  # TOT_QTY (Q열 추가로 +1)
    COL_AC = 30  # 모델타입 (Q열 추가로 +1)

    for r in range(DATA_START, data_end + 1):
        tot_val = ws.cell(row=r, column=COL_AB).value
        tot_num = tot_val if isinstance(tot_val, (int, float)) else 0

        # 규칙 A: TOT_QTY = 0 → 행 전체 폰트 회색(AAAAAA)
        if tot_num == 0:
            for c in range(1, TOTAL_COLS + 1):
                ws.cell(row=r, column=c).font = Font(
                    name="맑은 고딕", size=10, color="AAAAAA")

        # 규칙 B: ML가중치 >= 0.3 → S열 배경 주황(FFC000)
        ml_val = ws.cell(row=r, column=COL_S).value
        if isinstance(ml_val, (int, float)) and ml_val >= 0.3:
            ws.cell(row=r, column=COL_S).fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid")

        # 규칙 C: TOT_QTY >= 5 → AB열 배경 빨강 + 흰색 Bold
        if tot_num >= 5:
            ws.cell(row=r, column=COL_AB).fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid")
            ws.cell(row=r, column=COL_AB).font = Font(
                name="맑은 고딕", size=10, color="FFFFFF", bold=True)

        # 규칙 D: 모델타입 = 'rule' → AC열 배경 연초록(E2EFDA)
        model_val = ws.cell(row=r, column=COL_AC).value
        if model_val == "rule":
            ws.cell(row=r, column=COL_AC).fill = PatternFill(
                start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        # 규칙 E (NEW): RI stale → RI조회시각 셀 배경 연빨강(FFC7CE)
        COL_RI = 17  # RI조회시각 (Q열)
        item_idx = r - DATA_START
        if item_idx < len(order_list) and order_list[item_idx].get("ri_stale"):
            ws.cell(row=r, column=COL_RI).fill = PatternFill(
                start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # ─── 합계행 ───
    total_row = data_end + 1
    ws.cell(row=total_row, column=1, value="합계").font = Font(
        name="맑은 고딕", bold=True, size=10)
    ws.cell(row=total_row, column=1).border = thin

    sum_keys = {"final_order_qty", "pyun_qty", "tot_qty", "need_qty"}
    for ci_1, (key, _) in enumerate(col_keys, start=1):
        if key in sum_keys:
            total = sum(
                _safe_get(it, key, 0) if key not in ("pyun_qty", "tot_qty")
                else _compute_bgf_fields(it).get(key, 0)
                for it in order_list
            )
            cell = ws.cell(row=total_row, column=ci_1, value=total)
            cell.font = Font(name="맑은 고딕", bold=True, size=10)
            cell.number_format = "0"
            cell.border = thin

    # ─── 정보 행 ───
    info = total_row + 2
    for ri, txt in enumerate([
        f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"배송일: {delivery_date}",
        f"상품수: {len(order_list)}개",
    ]):
        c = ws.cell(row=info + ri, column=1, value=txt)
        c.font = Font(name="맑은 고딕", size=9, color="666666")

    # ─── 요약 시트 ───
    _create_summary_sheet(wb, order_list)

    # ─── 저장 ───
    wb.save(output_path)
    print(f"  Excel 저장: {output_path}")
    return output_path


# ═══════════════════════════════════════════════════════════════════════
# 오프라인 드라이런 (브라우저 없이 DB 기반 예측 → 엑셀 내보내기)
# ═══════════════════════════════════════════════════════════════════════

def run_dryrun_and_export(
    store_id: str = DEFAULT_STORE_ID,
    max_items: int = 999,
) -> str:
    """드라이런 예측 실행 → order_list 획득 → 엑셀 저장.

    실제 7시 스케줄러의 execute() 경로와 동일한 순서로 실행:
      1. 미취급 상품 DB 로드
      2. 발주중지(CUT) 상품 DB 로드
      3. 자동/스마트발주 제외 목록 로드 (DB 캐시)
      4. get_recommendations(min_order_qty=1) — 예측 파이프라인
      5. CUT 재필터 (_refilter_cut_items)
      6. 미입고+실시간재고 조정 (_apply_pending_and_stock)
      7. 수동발주 차감 (_deduct_manual_food_orders)
    발주 실행(executor.execute_orders)만 생략.

    Returns:
        저장된 엑셀 경로
    """
    from src.order.auto_order import AutoOrderSystem

    print(f"[1/5] 드라이런 시작 (store={store_id}, max_items={max_items})")
    print(f"  ※ 실제 발주(execute) 경로와 동일한 순서로 실행")

    system = AutoOrderSystem(
        driver=None,
        use_improved_predictor=True,
        store_id=store_id,
    )

    # ── execute() Step 1: 미취급 상품 DB 로드 ──
    system.load_unavailable_from_db()
    unavail_cnt = len(system._unavailable_items)
    print(f"  미취급 상품 로드: {unavail_cnt}개")

    # ── execute() Step 2: 발주중지(CUT) 상품 DB 로드 ──
    system.load_cut_items_from_db()
    cut_cnt = len(system._cut_items)
    print(f"  발주중지(CUT) 로드: {cut_cnt}개")

    # ── execute() Step 4: 자동/스마트발주 제외 목록 로드 (DB 캐시) ──
    system.load_auto_order_items(skip_site_fetch=True)
    auto_cnt = len(system._auto_order_items)
    smart_cnt = len(system._smart_order_items)
    print(f"  자동발주 제외: {auto_cnt}개, 스마트발주 제외: {smart_cnt}개")

    # ── execute() Step 5: 예측 (get_recommendations) ──
    print(f"[2/5] 예측 파이프라인 실행 (min_order_qty=1)...")
    order_list = system.get_recommendations(
        min_order_qty=1,
        max_items=max_items if max_items > 0 else None,
    )
    pred_cnt = len(order_list)
    print(f"  예측 완료: {pred_cnt}개 상품")

    if not order_list:
        print("  발주 대상 상품 없음")
        delivery_date = datetime.now().strftime("%Y-%m-%d")
        output_path = create_dryrun_excel(
            order_list=[],
            delivery_date=delivery_date,
            store_id=store_id,
        )
        print(f"[5/5] 빈 엑셀 저장: {output_path}")
        return output_path

    # ── execute() Step 6: CUT 재필터 ──
    order_list, _cut_lost = system._refilter_cut_items(order_list)
    if _cut_lost:
        system._cut_lost_items = _cut_lost
        print(f"  CUT 재필터: {len(_cut_lost)}개 제거")

    # ── execute() Step 7: 미입고+실시간재고 DB 조정 ──
    print(f"[3/5] 미입고+재고 DB 조정...")
    ri_freshness_map = {}   # {item_cd: {"queried_at": str, "is_stale": bool, "hours_ago": float}}
    try:
        from src.infrastructure.database.repos import RealtimeInventoryRepository
        from src.infrastructure.database.repos.order_tracking_repo import OrderTrackingRepository
        inv_repo = RealtimeInventoryRepository(store_id=store_id)
        all_inv = inv_repo.get_all(store_id=store_id)

        # ── [NEW-1] RI freshness 분석 (FR-01) ──
        now = datetime.now()

        pending_data = {}
        stock_data = {}
        for item in all_inv:
            item_cd = item.get("item_cd", "")
            if not item_cd:
                continue
            pq = item.get("pending_qty", 0) or 0
            sq = item.get("stock_qty", 0) or 0
            if pq > 0:
                pending_data[item_cd] = pq
            if sq > 0:
                stock_data[item_cd] = sq

            # freshness 분석
            queried_at = item.get("queried_at")
            is_stale = True
            hours_ago = -1.0
            if queried_at:
                try:
                    qt = datetime.fromisoformat(queried_at)
                    hours_ago = (now - qt).total_seconds() / 3600
                    is_stale = False  # 기본값, order_list 매칭 시 mid_cd 기반 재판정
                except (ValueError, TypeError):
                    pass
            ri_freshness_map[item_cd] = {
                "queried_at": queried_at or "",
                "is_stale": is_stale,
                "hours_ago": round(hours_ago, 1),
            }

        # order_tracking 미입고 병합 (중복 발주 방지)
        ot_pending_merged = 0
        try:
            ot_repo = OrderTrackingRepository(store_id=store_id)
            ot_pending = ot_repo.get_pending_qty_sum_batch(store_id=store_id)
            if ot_pending:
                for item_cd, ot_qty in ot_pending.items():
                    ri_qty = pending_data.get(item_cd, 0)
                    if ot_qty > ri_qty:
                        pending_data[item_cd] = ot_qty
                        ot_pending_merged += 1
                print(f"  order_tracking 미입고 병합: {len(ot_pending)}개 중 {ot_pending_merged}개 반영 "
                      f"(RI보다 OT가 큰 상품)")
        except Exception as e:
            print(f"  order_tracking 조회 실패 (RI 데이터만 사용): {e}")

        if pending_data or stock_data:
            before_total = sum(it.get("final_order_qty", 0) for it in order_list)
            order_list = system._apply_pending_and_stock_to_order_list(
                order_list=order_list,
                pending_data=pending_data,
                stock_data=stock_data,
                min_order_qty=1,
            )
            after_total = sum(it.get("final_order_qty", 0) for it in order_list)
            diff = before_total - after_total
            print(f"  미입고({len(pending_data)})+재고({len(stock_data)}) 반영: "
                  f"{before_total}개 → {after_total}개 ({abs(diff)}개 {'감소' if diff > 0 else '증가' if diff < 0 else '변동없음'})")
        else:
            print(f"  미입고/재고 DB 데이터 없음 (조정 생략)")
    except Exception as e:
        print(f"  미입고+재고 조정 실패 (원본 유지): {e}")

    # ── [NEW-2] order_list에 freshness 주입 + mid_cd 기반 stale 재판정 (FR-04) ──
    for item in order_list:
        item_cd = item.get("item_cd", "")
        mid_cd = item.get("mid_cd", "")
        fresh = ri_freshness_map.get(item_cd, {})

        hours_ago = fresh.get("hours_ago", -1)
        threshold_h = DRYRUN_STALE_HOURS["food"] if mid_cd in FOOD_MID_CDS else DRYRUN_STALE_HOURS["default"]
        is_stale = hours_ago < 0 or hours_ago > threshold_h

        item["ri_queried_at"] = fresh.get("queried_at", "")
        item["ri_stale"] = is_stale

    # stale 경고 출력
    stale_in_order = [it for it in order_list if it.get("ri_stale")]
    if stale_in_order:
        food_stale = sum(1 for it in stale_in_order if it.get("mid_cd", "") in FOOD_MID_CDS)
        other_stale = len(stale_in_order) - food_stale
        print(f"  [stale경고] RI 데이터 오래됨: 푸드 {food_stale}개(>{DRYRUN_STALE_HOURS['food']}h), "
              f"기타 {other_stale}개(>{DRYRUN_STALE_HOURS['default']}h)")

    # ── [NEW-6] CUT stale 경고 (FR-03) ──
    if ri_freshness_map and hasattr(system, '_cut_items') and system._cut_items:
        old_cut = 0
        for cd in system._cut_items:
            fresh = ri_freshness_map.get(cd, {})
            if fresh.get("hours_ago", -1) > 72:  # 3일 이상 미확인
                old_cut += 1
        if old_cut > 0:
            print(f"  [경고] CUT 상품 중 {old_cut}개가 72h+ 미확인 "
                  f"(실제 CUT 해제 가능성)")

    # ── execute() Step 8: 캐시 초기화 ──
    system.predictor.clear_pending_qty_cache()
    if system.use_improved_predictor and system.improved_predictor:
        system.improved_predictor.clear_pending_cache()
        system.improved_predictor.clear_stock_cache()

    # ── execute() Step 9: 수동발주 차감 ──
    print(f"[4/5] 수동발주 차감...")
    before_cnt = len(order_list)
    order_list = system._deduct_manual_food_orders(order_list, min_order_qty=1)
    deducted = before_cnt - len(order_list)
    if deducted > 0:
        print(f"  수동발주 차감: {deducted}개 제거 → {len(order_list)}개 잔여")
    else:
        print(f"  수동발주 차감 대상 없음")

    # ── 발주량 0 제외 ──
    order_list = [item for item in order_list if _safe_get(item, "final_order_qty", 0) > 0]
    print(f"  → 최종 발주 대상: {len(order_list)}개 상품")

    total_qty = sum(_safe_get(item, "final_order_qty", 0) for item in order_list)
    total_pyun = sum(_compute_bgf_fields(item)["pyun_qty"] for item in order_list)

    # 배송일 추정
    delivery_date = None
    if order_list:
        delivery_date = order_list[0].get("target_date")
    if not delivery_date:
        delivery_date = datetime.now().strftime("%Y-%m-%d")

    output_path = create_dryrun_excel(
        order_list=order_list,
        delivery_date=delivery_date,
        store_id=store_id,
    )

    print(f"[5/5] 엑셀 저장 완료: {output_path}")
    print(f"  → 상품: {len(order_list)}개, 총발주량: {total_qty}개, 총배수: {total_pyun}")

    # ── [NEW-3] 스케줄러 차이 경고 (FR-05) ──
    print()
    print("─" * 60)
    print("[스케줄러 차이 경고] 실제 7시 발주와 다를 수 있는 항목:")
    stale_cnt = sum(1 for it in order_list if it.get("ri_stale"))
    if stale_cnt > 0:
        print(f"  - RI stale 상품: {stale_cnt}개 (재고/미입고가 실제와 다를 수 있음)")
    try:
        cut_cnt = len(system._cut_items)
        if cut_cnt > 0:
            print(f"  - CUT 상품: {cut_cnt}개 (DB 기준, 실제 CUT 해제 미반영 가능)")
    except Exception:
        pass
    print(f"  - 자동/스마트발주: DB 캐시 사용 (사이트 최신 목록과 다를 수 있음)")
    print(f"  - 미입고 소스: DB(RI+OT) 캐시 (실시간 BGF 조회 아님)")
    print("─" * 60)

    return output_path


# ═══════════════════════════════════════════════════════════════════════
# 메인 플로우 (로그인 → 수집 → 발주 → Excel)
# ═══════════════════════════════════════════════════════════════════════

def run_full_flow(
    dry_run: bool = True,
    collect_sales: bool = True,
    max_order_items: int = 5,
    min_order_qty: int = 1,
    store_id: str = DEFAULT_STORE_ID,
    no_report: bool = False,
) -> Dict[str, Any]:
    """
    전체 플로우 실행

    로그인/수집은 스크립트 고유 로직이므로 직접 처리하고,
    핵심 발주 파이프라인은 DailyOrderFlow에 위임합니다.

    Args:
        dry_run: True면 실제 발주 안함 (테스트)
        collect_sales: 판매 데이터 수집 여부
        max_order_items: 발주할 최대 상품 수
        min_order_qty: 최소 발주량
        store_id: 점포 코드
        no_report: True면 드라이런 Excel 리포트 생성 스킵
    """
    print("\n" + "="*70)
    print("BGF 자동화 시스템 - 전체 플로우")
    print(f"시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("="*70)

    analyzer = None
    start_time = time.time()

    try:
        # ============================================================
        # 1단계: BGF 로그인
        # ============================================================
        print("\n" + "="*70)
        print("[1단계] BGF 시스템 로그인")
        print("="*70)

        analyzer = SalesAnalyzer(store_id=store_id)
        analyzer.setup_driver()
        analyzer.connect()

        if not analyzer.do_login():
            print("[ERROR] 로그인 실패")
            return {"success": False, "step": "login", "message": "login failed"}

        print("[OK] 로그인 성공")

        # 팝업 닫기
        analyzer.close_popup()
        time.sleep(2)

        # ============================================================
        # 2단계: 판매 데이터 수집 (스크립트 고유 — 누락일 백필)
        # ============================================================
        if collect_sales:
            print("\n" + "="*70)
            print("[2단계] 판매 데이터 수집")
            print("="*70)

            # 판매현황 메뉴로 이동
            print("\n[이동] 판매현황 메뉴...")
            if not analyzer.navigate_to_sales_menu():
                print("[WARN] 메뉴 이동 실패, 수집 건너뜀")
            else:
                repo = SalesRepository(store_id=store_id)

                # 과거 30일 중 누락된 날짜 확인
                missing_dates = repo.get_missing_dates(days=30)
                print(f"\n[확인] 과거 30일 중 누락된 날짜: {len(missing_dates)}개")

                if missing_dates:
                    print(f"  누락 날짜: {missing_dates[:5]}{'...' if len(missing_dates) > 5 else ''}")

                    # 날짜 순(오래된 것부터)으로 수집
                    for date_str in missing_dates:
                        # YYYY-MM-DD -> YYYYMMDD 변환
                        target_date = date_str.replace("-", "")
                        print(f"\n[수집] {target_date} (누락 데이터)")

                        try:
                            data = analyzer.collect_all_mid_category_data(target_date)
                            if data:
                                print(f"  -> 수집 완료: {len(data)}건")
                                stats = repo.save_daily_sales(data, date_str)
                                print(f"  -> DB 저장: 신규 {stats.get('new', 0)}건, 업데이트 {stats.get('updated', 0)}건")
                            else:
                                print(f"  -> 데이터 없음")
                        except Exception as e:
                            print(f"  -> 수집 실패: {e}")

                        time.sleep(0.5)  # 요청 간 딜레이
                else:
                    print("  [OK] 과거 30일 데이터 모두 수집됨")

                # 오늘 데이터 수집 (YYYYMMDD 형식)
                target_date = datetime.now().strftime("%Y%m%d")
                print(f"\n[수집] {target_date} (오늘)")

                try:
                    data = analyzer.collect_all_mid_category_data(target_date)
                    if data:
                        print(f"  -> 수집 완료: {len(data)}건")

                        # DB 저장
                        sales_date = datetime.now().strftime("%Y-%m-%d")
                        stats = repo.save_daily_sales(data, sales_date)
                        print(f"  -> DB 저장: 신규 {stats.get('new', 0)}건, 업데이트 {stats.get('updated', 0)}건")
                except Exception as e:
                    print(f"  -> 수집 실패: {e}")

            # 매출분석 탭 닫기 (발주 화면 진입 전 정리)
            print("\n[정리] 매출분석 탭 닫기...")
            sales_collector = SalesCollector(analyzer)
            sales_collector.close_sales_menu()

            time.sleep(1)
        else:
            print("\n[2단계] 판매 데이터 수집 건너뜀")

        # ============================================================
        # 3단계: DB 현황 확인
        # ============================================================
        print("\n" + "="*70)
        print("[3단계] DB 현황 확인")
        print("="*70)

        sales_repo = SalesRepository(store_id=store_id)
        stats = sales_repo.get_stats_summary()

        print(f"  총 상품 수: {stats.get('total_products', 0)}개")
        print(f"  총 카테고리: {stats.get('total_categories', 0)}개")
        print(f"  수집된 일수: {stats.get('total_days', 0)}일")
        print(f"  기간: {stats.get('first_date', 'N/A')} ~ {stats.get('last_date', 'N/A')}")

        # ============================================================
        # 4단계: 발주 실행 (DailyOrderFlow 위임)
        # ============================================================
        print("\n" + "="*70)
        print(f"[4단계] 발주 실행 ({'테스트 모드' if dry_run else '실제 발주'})")
        print("="*70)

        if dry_run:
            print("\n[INFO] dry_run=True -> 실제 발주 없이 시뮬레이션만 수행")

        from src.application.use_cases.daily_order_flow import DailyOrderFlow
        from src.settings.store_context import StoreContext

        ctx = StoreContext.from_store_id(store_id)
        flow = DailyOrderFlow(
            store_ctx=ctx,
            driver=analyzer.driver,
            use_improved_predictor=True,
        )
        flow_result = flow.run_auto_order(
            dry_run=dry_run,
            min_order_qty=min_order_qty,
            max_items=max_order_items if max_order_items > 0 else None,
            prefetch_pending=True,
            max_pending_items=100,
            collect_fail_reasons=True,
        )

        # ============================================================
        # 결과 요약
        # ============================================================
        elapsed = time.time() - start_time

        exec_result = flow_result.get("execution", {})
        fail_reasons = flow_result.get("fail_reasons")

        print("\n" + "="*70)
        print("전체 플로우 완료")
        print("="*70)
        print(f"소요 시간: {elapsed:.1f}초")
        print(f"발주 성공: {flow_result.get('success_count', 0)}건")
        print(f"발주 실패: {flow_result.get('fail_count', 0)}건")
        if fail_reasons:
            print(f"실패 사유 확인: {fail_reasons.get('checked', 0)}건")
        print(f"모드: {'테스트 (dry_run)' if dry_run else '실제 발주'}")

        if exec_result.get('grouped_by_date'):
            print("\n[발주 일정]")
            for date, count in exec_result.get('grouped_by_date', {}).items():
                print(f"  {date}: {count}개 상품")

        # ============================================================
        # 5단계: 드라이런 Excel 리포트 자동 생성
        # ============================================================
        if dry_run and not no_report:
            # order_list는 execution(=auto_order result) 안에 있음
            _exec = flow_result.get("execution", {})
            order_list = _exec.get("order_list", [])

            if order_list:
                print("\n" + "="*70)
                print("[5단계] 드라이런 Excel 리포트 생성 (29컬럼 상세)")
                print("="*70)

                try:
                    export_dir = project_root / "data" / "exports"
                    export_dir.mkdir(parents=True, exist_ok=True)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    xlsx_path = str(export_dir / f"dry_order_{timestamp}.xlsx")

                    create_dryrun_excel(
                        order_list=order_list,
                        output_path=xlsx_path,
                        store_id=store_id,
                    )
                    print(f"  Excel 리포트: {xlsx_path}")
                except Exception as e:
                    print(f"  [WARN] Excel 리포트 생성 실패: {e}")

        return {
            "success": flow_result.get("success", False),
            "elapsed": elapsed,
            "success_count": flow_result.get("success_count", 0),
            "fail_count": flow_result.get("fail_count", 0),
            "dry_run": dry_run
        }

    except Exception as e:
        print(f"\n[ERROR] 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return {"success": False, "message": str(e)}

    finally:
        if analyzer:
            try:
                input("\n브라우저를 닫으려면 Enter를 누르세요...")
            except EOFError:
                pass  # 비대화형 모드에서는 바로 종료
            analyzer.close()


# ═══════════════════════════════════════════════════════════════════════
# CLI
# ═══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="BGF 전체 자동화 플로우")
    parser.add_argument("--run", action="store_true", help="실제 발주 실행")
    parser.add_argument("--no-collect", action="store_true", help="판매 데이터 수집 건너뜀")
    parser.add_argument("--max-items", "-n", type=int, default=5, help="최대 발주 상품 수")
    parser.add_argument("--min-qty", "-q", type=int, default=1, help="최소 발주량")
    parser.add_argument("--store-id", type=str, default=DEFAULT_STORE_ID, help="점포 코드 (예: 46513)")
    parser.add_argument("--export-excel", action="store_true",
                        help="드라이런 후 발주 상세 엑셀 내보내기 (브라우저 없이 예측만)")
    parser.add_argument("--no-report", action="store_true",
                        help="드라이런 시 Excel 리포트 생성 스킵")

    args = parser.parse_args()

    if args.export_excel:
        # 브라우저 없이 예측만 수행 → 엑셀 내보내기
        path = run_dryrun_and_export(
            store_id=args.store_id,
            max_items=args.max_items,
        )
        print(f"\n엑셀 내보내기 완료: {path}")
    else:
        run_full_flow(
            dry_run=not args.run,
            collect_sales=not args.no_collect,
            max_order_items=args.max_items,
            min_order_qty=args.min_qty,
            store_id=args.store_id,
            no_report=args.no_report,
        )
