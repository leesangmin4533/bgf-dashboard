"""46513 매장 3/28 실제 발주 목록 엑셀 내보내기"""
import sqlite3
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

store_id = "46513"
order_date = "2026-03-28"
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

conn = sqlite3.connect(os.path.join(base_dir, f"data/stores/{store_id}.db"))
conn.row_factory = sqlite3.Row
common = sqlite3.connect(os.path.join(base_dir, "data/common.db"))
common.row_factory = sqlite3.Row

prod_map = {r["item_cd"]: r["item_nm"] for r in common.execute("SELECT item_cd, item_nm FROM products")}
mid_map = {r["mid_cd"]: r["mid_nm"] for r in common.execute("SELECT mid_cd, mid_nm FROM mid_categories")}
detail_map = {}
for r in common.execute(
    "SELECT item_cd, expiration_days, demand_pattern, order_unit_qty, promo_type, sell_price, margin_rate FROM product_details"
):
    detail_map[r["item_cd"]] = dict(r)

rows = conn.execute(
    """
    SELECT item_cd, item_nm, mid_cd, delivery_type, order_qty,
           remaining_qty, arrival_time, expiry_time, status,
           order_source, ord_input_id, actual_receiving_qty, created_at
    FROM order_tracking
    WHERE store_id = ? AND order_date = ?
    ORDER BY mid_cd, item_cd
""",
    (store_id, order_date),
).fetchall()

eval_map = {}
for r in conn.execute(
    """
    SELECT item_cd, decision, daily_avg, current_stock, predicted_qty,
           actual_order_qty, sell_price, margin_rate, promo_type, stockout_freq
    FROM eval_outcomes WHERE store_id = ? AND eval_date = ?
""",
    (store_id, order_date),
):
    eval_map[r["item_cd"]] = dict(r)

wb = openpyxl.Workbook()
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
auto_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
site_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ===== Sheet 1: 실제 발주 목록 =====
ws = wb.active
ws.title = "실제 발주 목록"

headers = [
    "No",
    "상품코드",
    "상품명",
    "중분류코드",
    "중분류명",
    "발주수량",
    "배송구분",
    "발주출처",
    "판정",
    "일평균판매",
    "현재고",
    "예측수량",
    "판매가",
    "마진율(%)",
    "행사",
    "도착예정",
    "만료예정",
    "상태",
    "수요패턴",
    "유통기한(일)",
    "배수단위",
    "품절빈도",
    "발주입력ID",
]

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws.row_dimensions[1].height = 35

for i, row in enumerate(rows, 1):
    d = dict(row)
    item_cd = d["item_cd"]
    ev = eval_map.get(item_cd, {})
    det = detail_map.get(item_cd, {})

    values = [
        i,
        item_cd,
        prod_map.get(item_cd, d.get("item_nm", "")),
        d["mid_cd"],
        mid_map.get(d["mid_cd"] or "", ""),
        d["order_qty"],
        d["delivery_type"] or "",
        d["order_source"] or "",
        ev.get("decision", ""),
        round(ev.get("daily_avg", 0), 2) if ev.get("daily_avg") else "",
        ev.get("current_stock", ""),
        ev.get("predicted_qty") if ev.get("predicted_qty") is not None else "",
        ev.get("sell_price") or det.get("sell_price", ""),
        round(ev.get("margin_rate", 0), 1) if ev.get("margin_rate") else "",
        ev.get("promo_type") or det.get("promo_type", ""),
        d["arrival_time"] or "",
        d["expiry_time"] or "",
        d["status"] or "",
        det.get("demand_pattern", ""),
        det.get("expiration_days", ""),
        det.get("order_unit_qty", ""),
        round(ev.get("stockout_freq", 0), 3) if ev.get("stockout_freq") else "",
        d["ord_input_id"] or "",
    ]

    for col, v in enumerate(values, 1):
        cell = ws.cell(row=i + 1, column=col, value=v)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center" if col != 3 else "left")

        source = d.get("order_source", "")
        if col in (6, 8):
            if source == "auto":
                cell.fill = auto_fill
            elif source == "site":
                cell.fill = site_fill

widths = [5, 16, 30, 8, 14, 8, 8, 8, 14, 8, 7, 8, 8, 8, 8, 16, 16, 8, 10, 10, 8, 8, 14]
for idx, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

# ===== Sheet 2: 요약 =====
ws2 = wb.create_sheet("요약")
bold = Font(bold=True, size=11)

ws2.cell(row=1, column=1, value=f"{store_id} 실제 발주 현황 ({order_date})").font = Font(bold=True, size=14)

ws2.cell(row=3, column=1, value="발주 출처별").font = Font(bold=True, size=12)
for col, h in enumerate(["출처", "건수", "총수량"], 1):
    ws2.cell(row=4, column=col, value=h).font = bold

source_map = {}
for row in rows:
    d = dict(row)
    src = d["order_source"] or "unknown"
    if src not in source_map:
        source_map[src] = {"count": 0, "qty": 0}
    source_map[src]["count"] += 1
    source_map[src]["qty"] += d["order_qty"]

r = 5
for src in ["auto", "site", "unknown"]:
    if src in source_map:
        ws2.cell(row=r, column=1, value=src)
        ws2.cell(row=r, column=2, value=source_map[src]["count"])
        ws2.cell(row=r, column=3, value=source_map[src]["qty"])
        r += 1
ws2.cell(row=r, column=1, value="합계").font = bold
ws2.cell(row=r, column=2, value=sum(v["count"] for v in source_map.values())).font = bold
ws2.cell(row=r, column=3, value=sum(v["qty"] for v in source_map.values())).font = bold

r += 2
ws2.cell(row=r, column=1, value="배송구분별").font = Font(bold=True, size=12)
r += 1
for col, h in enumerate(["배송구분", "건수", "총수량"], 1):
    ws2.cell(row=r, column=col, value=h).font = bold
r += 1
dtype_map = {}
for row in rows:
    d = dict(row)
    dt = d["delivery_type"] or "(미지정)"
    if dt not in dtype_map:
        dtype_map[dt] = {"count": 0, "qty": 0}
    dtype_map[dt]["count"] += 1
    dtype_map[dt]["qty"] += d["order_qty"]
for dt in sorted(dtype_map.keys()):
    ws2.cell(row=r, column=1, value=dt)
    ws2.cell(row=r, column=2, value=dtype_map[dt]["count"])
    ws2.cell(row=r, column=3, value=dtype_map[dt]["qty"])
    r += 1

r += 1
ws2.cell(row=r, column=1, value="중분류별 발주").font = Font(bold=True, size=12)
r += 1
for col, h in enumerate(["중분류", "중분류명", "건수", "총수량"], 1):
    ws2.cell(row=r, column=col, value=h).font = bold
r += 1
mid_sum = {}
for row in rows:
    d = dict(row)
    mid = d["mid_cd"] or ""
    if mid not in mid_sum:
        mid_sum[mid] = {"count": 0, "qty": 0}
    mid_sum[mid]["count"] += 1
    mid_sum[mid]["qty"] += d["order_qty"]
for mid in sorted(mid_sum.keys()):
    ws2.cell(row=r, column=1, value=mid)
    ws2.cell(row=r, column=2, value=mid_map.get(mid, ""))
    ws2.cell(row=r, column=3, value=mid_sum[mid]["count"])
    ws2.cell(row=r, column=4, value=mid_sum[mid]["qty"])
    r += 1

ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 18
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10

# Save
output_dir = os.path.join(base_dir, "data", "reports")
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, f"{store_id}_actual_orders_{order_date.replace('-', '')}.xlsx")
wb.save(output_path)

print(f"Excel saved: {output_path}")
print(f"  Sheet 1 (실제 발주 목록): {len(rows)}건, 총 {sum(dict(r)['order_qty'] for r in rows)}개")
print(f"  Sheet 2 (요약): 출처별 + 배송별 + 중분류별")
auto = source_map.get("auto", {"count": 0, "qty": 0})
site = source_map.get("site", {"count": 0, "qty": 0})
print(f"    - auto(자동): {auto['count']}건 ({auto['qty']}개)")
print(f"    - site(수동): {site['count']}건 ({site['qty']}개)")

conn.close()
common.close()
