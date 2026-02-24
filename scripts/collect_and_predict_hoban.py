"""
호반점(46704) 데이터 수집 + 예측 + 엑셀 저장
1단계: BGF 사이트에서 누락/오늘 판매 데이터 수집
2단계: 전체 상품 예측 실행
3단계: 엑셀 파일로 저장
"""

import sys
import io

# Windows CP949 콘솔 → UTF-8 래핑 (한글·특수문자 깨짐 방지)
if sys.platform == "win32":
    for _sn in ("stdout", "stderr"):
        _s = getattr(sys, _sn, None)
        if _s and hasattr(_s, "buffer") and (getattr(_s, "encoding", "") or "").lower().replace("-", "") != "utf8":
            setattr(sys, _sn, io.TextIOWrapper(_s.buffer, encoding="utf-8", errors="replace", line_buffering=True))

import os
from pathlib import Path

# 프로젝트 루트/src를 path에 추가
project_root = Path(__file__).parent.parent
src_root = project_root / "src"
sys.path.insert(0, str(project_root))
sys.path.insert(0, str(src_root))
os.chdir(str(src_root))

import time
import sqlite3
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any

STORE_ID = "46513"
STORE_NAME = "CU 호반점"

# ============================================================
# 1단계: 데이터 수집
# ============================================================
def collect_sales_data():
    """BGF 사이트에서 호반점 판매 데이터 수집"""
    from sales_analyzer import SalesAnalyzer
    from db.repository import SalesRepository

    print("\n" + "=" * 70)
    print(f"[1단계] {STORE_NAME}({STORE_ID}) 판매 데이터 수집")
    print("=" * 70)

    analyzer = None
    try:
        analyzer = SalesAnalyzer(store_id=STORE_ID)
        analyzer.setup_driver()
        analyzer.connect()

        if not analyzer.do_login():
            print("[ERROR] 로그인 실패")
            return False

        print("[OK] 로그인 성공")
        analyzer.close_popup()
        time.sleep(2)

        # 매출분석 메뉴 이동
        if not analyzer.navigate_to_sales_menu():
            print("[WARN] 메뉴 이동 실패")
            return False

        repo = SalesRepository(store_id=STORE_ID)

        # 누락 날짜 수집
        store_db = str(project_root / "data" / "stores" / f"{STORE_ID}.db")
        conn = sqlite3.connect(store_db, timeout=30)
        cur = conn.cursor()

        # 최근 7일 중 누락 확인
        missing = []
        for i in range(7, 0, -1):
            d = (datetime.now() - timedelta(days=i)).strftime("%Y-%m-%d")
            cur.execute("SELECT COUNT(*) FROM daily_sales WHERE sales_date = ?", (d,))
            if cur.fetchone()[0] == 0:
                missing.append(d)
        conn.close()

        print(f"\n[확인] 최근 7일 중 누락: {len(missing)}일")

        for date_str in missing:
            target = date_str.replace("-", "")
            print(f"\n[수집] {target} (누락)")
            try:
                data = analyzer.collect_all_mid_category_data(target)
                if data:
                    print(f"  -> {len(data)}건 수집")
                    stats = repo.save_daily_sales(data, date_str, store_id=STORE_ID)
                    print(f"  -> 저장: 신규 {stats.get('new', 0)}, 업데이트 {stats.get('updated', 0)}")
                else:
                    print(f"  -> 데이터 없음")
            except Exception as e:
                print(f"  -> 실패: {e}")
            time.sleep(0.5)

        # 오늘 데이터 수집
        today_str = datetime.now().strftime("%Y%m%d")
        today_date = datetime.now().strftime("%Y-%m-%d")
        print(f"\n[수집] {today_str} (오늘)")
        try:
            data = analyzer.collect_all_mid_category_data(today_str)
            if data:
                print(f"  -> {len(data)}건 수집")
                stats = repo.save_daily_sales(data, today_date, store_id=STORE_ID)
                print(f"  -> 저장: 신규 {stats.get('new', 0)}, 업데이트 {stats.get('updated', 0)}")
        except Exception as e:
            print(f"  -> 실패: {e}")

        print("\n[OK] 수집 완료")
        return True

    except Exception as e:
        print(f"[ERROR] 수집 중 오류: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if analyzer and hasattr(analyzer, 'driver') and analyzer.driver:
            try:
                analyzer.driver.quit()
            except Exception:
                pass


# ============================================================
# 2단계: 예측 실행
# ============================================================
def run_predictions() -> List[Dict[str, Any]]:
    """호반점 전체 상품 예측"""
    from prediction.improved_predictor import ImprovedPredictor

    print("\n" + "=" * 70)
    print(f"[2단계] {STORE_NAME}({STORE_ID}) 예측 실행")
    print("=" * 70)

    store_db = str(project_root / "data" / "stores" / f"{STORE_ID}.db")
    common_db = str(project_root / "data" / "common.db")

    # ImprovedPredictor는 내부에서 products 테이블을 직접 쿼리하므로
    # store DB(products 없음) 대신 레거시 통합 DB 또는 store_id 기반 사용
    # → db_path=None이면 기본 bgf_sales.db 사용
    predictor = ImprovedPredictor(
        db_path=store_db,
        use_db_inventory=True,
        store_id=STORE_ID
    )

    # products/product_details를 store DB에 ATTACH하여 사용 가능하도록
    # predictor의 _get_connection을 패치: common.db를 ATTACH
    _orig_get_conn = predictor._get_connection
    def _patched_get_connection(timeout=30):
        conn = _orig_get_conn(timeout)
        try:
            conn.execute(f"ATTACH DATABASE '{common_db}' AS common")
            # products, product_details를 main 스키마에 VIEW로 생성
            conn.execute("CREATE TEMP VIEW IF NOT EXISTS products AS SELECT * FROM common.products")
            conn.execute("CREATE TEMP VIEW IF NOT EXISTS product_details AS SELECT * FROM common.product_details")
            conn.execute("CREATE TEMP VIEW IF NOT EXISTS mid_categories AS SELECT * FROM common.mid_categories")
        except Exception:
            pass  # 이미 ATTACH된 경우
        return conn
    predictor._get_connection = _patched_get_connection

    # 활성 상품 조회 (최근 14일 판매 있는 상품)
    conn = sqlite3.connect(store_db, timeout=30)
    conn.execute(f"ATTACH DATABASE '{common_db}' AS common")
    cur = conn.cursor()
    cur.execute("""
        SELECT DISTINCT ds.item_cd, p.item_nm, p.mid_cd
        FROM daily_sales ds
        LEFT JOIN common.products p ON ds.item_cd = p.item_cd
        WHERE ds.sales_date >= date('now', '-14 days')
        AND ds.sale_qty > 0
        ORDER BY p.mid_cd, ds.item_cd
    """)
    active_items = cur.fetchall()
    conn.close()

    print(f"\n활성 상품: {len(active_items)}개")

    results = []
    success = 0
    fail = 0

    target_date = datetime.now() + timedelta(days=1)  # 내일 예측

    for item_cd, item_nm, mid_cd in active_items:
        try:
            pred = predictor.predict(item_cd, target_date=target_date)
            if pred and pred.order_qty >= 0:
                results.append({
                    "item_cd": item_cd,
                    "item_nm": item_nm or "",
                    "mid_cd": mid_cd or "",
                    "daily_avg": round(pred.daily_avg, 2) if hasattr(pred, 'daily_avg') else 0,
                    "predicted_qty": round(pred.predicted_qty, 2) if hasattr(pred, 'predicted_qty') else 0,
                    "safety_stock": round(pred.safety_stock, 2) if hasattr(pred, 'safety_stock') else 0,
                    "current_stock": pred.current_stock if hasattr(pred, 'current_stock') else 0,
                    "pending_qty": pred.pending_qty if hasattr(pred, 'pending_qty') else 0,
                    "order_qty": pred.order_qty,
                    "confidence": pred.confidence if hasattr(pred, 'confidence') else "",
                    "weekday_coef": round(pred.weekday_coef, 3) if hasattr(pred, 'weekday_coef') else 1.0,
                })
                success += 1
        except Exception as e:
            fail += 1

    print(f"예측 완료: 성공 {success}개, 실패 {fail}개")
    print(f"발주 대상 (order_qty > 0): {sum(1 for r in results if r['order_qty'] > 0)}개")

    return results


# ============================================================
# 3단계: 엑셀 저장
# ============================================================
def save_to_excel(results: List[Dict[str, Any]]) -> str:
    """예측 결과를 엑셀 파일로 저장"""
    print("\n" + "=" * 70)
    print("[3단계] 엑셀 파일 저장")
    print("=" * 70)

    # 카테고리명 매핑
    cat_names = {
        "001": "도시락", "002": "주먹밥", "003": "김밥", "004": "샌드위치",
        "005": "햄버거", "006": "조리면", "010": "음료", "012": "빵",
        "013": "유제품", "014": "껌", "015": "캔디", "016": "초콜릿",
        "017": "비스킷", "018": "파이", "019": "스낵", "020": "기타과자",
        "021": "냉동식품", "026": "반찬", "027": "즉석밥", "028": "즉석국",
        "029": "간식류", "030": "시리얼바", "031": "즉석조리1", "032": "면류",
        "033": "즉석조리2", "034": "아이스크림", "035": "즉석조리3",
        "036": "세면용품", "037": "위생용품", "039": "탄산음료",
        "040": "과즙음료", "041": "기능성음료", "042": "커피음료",
        "043": "생수", "044": "기타음료", "045": "유음료", "046": "신선식품",
        "047": "RTD커피", "048": "차음료", "049": "맥주", "050": "소주",
        "052": "양주", "053": "와인", "054": "잡화1", "055": "잡화2",
        "056": "주방용품", "057": "생활용품", "072": "담배", "073": "전자담배",
        "086": "세탁용품", "100": "냉동기타", "900": "소모품",
    }

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl 미설치 — CSV로 대체 저장합니다")
        return save_to_csv(results, cat_names)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주 예측"

    # 헤더
    headers = [
        "상품코드", "상품명", "중분류", "카테고리명",
        "일평균", "예측량", "안전재고", "현재고", "미입고",
        "발주량", "신뢰도", "요일계수"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 발주량 > 0인 것을 먼저, 나머지는 카테고리/상품코드 순
    sorted_results = sorted(results, key=lambda r: (-r['order_qty'], r['mid_cd'], r['item_cd']))

    # 데이터 행
    order_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    data_font = Font(name="맑은 고딕", size=10)

    for row_idx, r in enumerate(sorted_results, 2):
        cat_name = cat_names.get(r['mid_cd'], r['mid_cd'])
        values = [
            r['item_cd'], r['item_nm'], r['mid_cd'], cat_name,
            r['daily_avg'], r['predicted_qty'], r['safety_stock'],
            r['current_stock'], r['pending_qty'],
            r['order_qty'], r['confidence'], r['weekday_coef']
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = data_font
            cell.border = thin_border
            if col >= 5:  # 숫자 컬럼 우측 정렬
                cell.alignment = Alignment(horizontal='right')
            if r['order_qty'] > 0:
                cell.fill = order_fill

    # 컬럼 너비
    widths = [12, 30, 8, 12, 8, 8, 8, 8, 8, 8, 8, 8]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 필터 설정
    ws.auto_filter.ref = f"A1:L{len(sorted_results) + 1}"

    # 요약 시트
    ws2 = wb.create_sheet("요약")
    ws2.cell(row=1, column=1, value="호반점 발주 예측 요약").font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"예측일: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    ws2.cell(row=3, column=1, value=f"대상: 내일 ({(datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d %A')})")
    ws2.cell(row=5, column=1, value="전체 상품 수")
    ws2.cell(row=5, column=2, value=len(results))
    ws2.cell(row=6, column=1, value="발주 대상 (qty>0)")
    ws2.cell(row=6, column=2, value=sum(1 for r in results if r['order_qty'] > 0))
    ws2.cell(row=7, column=1, value="총 발주량")
    ws2.cell(row=7, column=2, value=sum(r['order_qty'] for r in results))

    # 카테고리별 요약
    ws2.cell(row=9, column=1, value="카테고리별 발주 요약").font = Font(bold=True, size=12)
    ws2.cell(row=10, column=1, value="카테고리")
    ws2.cell(row=10, column=2, value="상품수")
    ws2.cell(row=10, column=3, value="발주대상")
    ws2.cell(row=10, column=4, value="총발주량")

    cat_summary = {}
    for r in results:
        mid = r['mid_cd']
        if mid not in cat_summary:
            cat_summary[mid] = {"name": cat_names.get(mid, mid), "total": 0, "order": 0, "qty": 0}
        cat_summary[mid]["total"] += 1
        if r['order_qty'] > 0:
            cat_summary[mid]["order"] += 1
            cat_summary[mid]["qty"] += r['order_qty']

    for row, (mid, info) in enumerate(sorted(cat_summary.items()), 11):
        ws2.cell(row=row, column=1, value=f"{mid} {info['name']}")
        ws2.cell(row=row, column=2, value=info['total'])
        ws2.cell(row=row, column=3, value=info['order'])
        ws2.cell(row=row, column=4, value=info['qty'])

    # 저장
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"호반점_발주예측_{timestamp}.xlsx"
    filepath = str(project_root / "data" / "reports" / filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    wb.save(filepath)

    print(f"\n[OK] 엑셀 저장: {filepath}")
    print(f"  전체: {len(results)}개 상품")
    print(f"  발주대상: {sum(1 for r in results if r['order_qty'] > 0)}개")
    print(f"  총발주량: {sum(r['order_qty'] for r in results)}개")

    return filepath


def save_to_csv(results, cat_names):
    """openpyxl 없을 때 CSV 저장"""
    import csv
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"호반점_발주예측_{timestamp}.csv"
    filepath = str(project_root / "data" / "reports" / filename)
    os.makedirs(os.path.dirname(filepath), exist_ok=True)

    sorted_results = sorted(results, key=lambda r: (-r['order_qty'], r['mid_cd'], r['item_cd']))

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["상품코드", "상품명", "중분류", "카테고리명",
                         "일평균", "예측량", "안전재고", "현재고", "미입고",
                         "발주량", "신뢰도", "요일계수"])
        for r in sorted_results:
            writer.writerow([
                r['item_cd'], r['item_nm'], r['mid_cd'],
                cat_names.get(r['mid_cd'], r['mid_cd']),
                r['daily_avg'], r['predicted_qty'], r['safety_stock'],
                r['current_stock'], r['pending_qty'],
                r['order_qty'], r['confidence'], r['weekday_coef']
            ])

    print(f"\n[OK] CSV 저장: {filepath}")
    return filepath


# ============================================================
# 메인
# ============================================================
if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="호반점 수집+예측+엑셀")
    parser.add_argument("--no-collect", action="store_true", help="수집 건너뛰기")
    args = parser.parse_args()

    print(f"\n{'='*70}")
    print(f"  {STORE_NAME}({STORE_ID}) 수집 + 예측 + 엑셀 저장")
    print(f"  시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*70}")

    # 1단계: 수집
    if not args.no_collect:
        collect_ok = collect_sales_data()
        if not collect_ok:
            print("\n[WARN] 수집 실패 — 기존 데이터로 예측 진행")
    else:
        print("\n[SKIP] 수집 건너뜀 (--no-collect)")

    # 2단계: 예측
    results = run_predictions()

    if not results:
        print("\n[ERROR] 예측 결과 없음")
        sys.exit(1)

    # 3단계: 엑셀
    filepath = save_to_excel(results)

    print(f"\n{'='*70}")
    print(f"  완료: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  파일: {filepath}")
    print(f"{'='*70}")
