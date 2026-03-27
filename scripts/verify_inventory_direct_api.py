"""
재고 검증 스크립트 v2: Direct API 배치 덤프 → DB 비교

BGF 단품별 발주(STBJ030_M0) Direct API로 전체 상품 재고를 일괄 조회하여
DB realtime_inventory.stock_qty와 비교합니다.

기존 verify_inventory.py 대비:
  - Direct API 배치 (5 동시 요청, ~0.1초/건) vs Selenium 1건씩 (~3.5초/건)
  - 매출수집 선행 옵션 (--collect-sales)
  - 전체 상품 기본 처리 + 청크 분할
  - 멀티 매장 + 엑셀 내보내기 (매장별 시트)
  - 주간 자동 스케줄 (run_scheduler 연동)

Usage:
    python scripts/verify_inventory_direct_api.py                          # 기본매장 검증
    python scripts/verify_inventory_direct_api.py --all-stores             # 전 매장 검증 + 엑셀
    python scripts/verify_inventory_direct_api.py --store-id 46704         # 특정 매장
    python scripts/verify_inventory_direct_api.py --collect-sales          # 매출수집 후 검증
    python scripts/verify_inventory_direct_api.py --max-items 100          # 100개만
    python scripts/verify_inventory_direct_api.py --threshold 2            # 갭 2 이상만 불일치
    python scripts/verify_inventory_direct_api.py --sync                   # 불일치 → DB 동기화
"""

import sys
import io

# Windows CP949 콘솔 -> UTF-8 래핑 (한글 깨짐 방지)
if sys.platform == "win32":
    for _sn in ("stdout", "stderr"):
        _s = getattr(sys, _sn, None)
        if _s and hasattr(_s, "buffer") and (getattr(_s, "encoding", "") or "").lower().replace("-", "") != "utf8":
            setattr(sys, _sn, io.TextIOWrapper(_s.buffer, encoding="utf-8", errors="replace", line_buffering=True))

import argparse
import csv
import time
import traceback
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

# 프로젝트 루트를 path에 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from src.infrastructure.database.repos import RealtimeInventoryRepository
from src.settings.constants import DEFAULT_STORE_ID

# =============================================================================
# 상수
# =============================================================================
OUTPUT_DIR = project_root / "data" / "reports"
CHUNK_SIZE = 200           # Direct API 배치 단위
STALE_HOURS = 36
DEFAULT_MAX_ITEMS = None   # None = 전체

# 상태
STATUS_MATCH = "MATCH"
STATUS_MISMATCH = "MISMATCH"
STATUS_BGF_ONLY = "BGF_ONLY"      # BGF에만 존재 (DB 미등록)
STATUS_DB_ONLY = "DB_ONLY"        # DB에만 존재 (BGF 조회 실패)
STATUS_QUERY_FAIL = "FAIL"


def _is_stale(queried_at: Optional[str]) -> bool:
    """queried_at이 STALE_HOURS보다 오래됐는지"""
    if not queried_at:
        return True
    try:
        queried_dt = datetime.fromisoformat(queried_at)
        return queried_dt < datetime.now() - timedelta(hours=STALE_HOURS)
    except (ValueError, TypeError):
        return True


# =============================================================================
# Step 1: 매출수집
# =============================================================================
def run_sales_collection(analyzer: Any, store_id: str) -> bool:
    """BGF 사이트에서 오늘 매출 데이터 수집 + DB 저장"""
    from src.infrastructure.database.repos import SalesRepository

    print("\n[매출수집] 매출분석 메뉴 이동...")
    if not analyzer.navigate_to_sales_menu():
        print("[ERROR] 매출분석 메뉴 이동 실패")
        return False

    target_date = datetime.now().strftime("%Y%m%d")
    print(f"[매출수집] 날짜: {target_date}")

    print("[매출수집] 데이터 수집 시작...")
    start = time.time()
    data = analyzer.collect_all_mid_category_data(target_date)
    elapsed = time.time() - start

    if not data:
        print(f"[매출수집] 데이터 없음 (소요: {elapsed:.1f}초)")
        return False

    print(f"[매출수집] {len(data)}건 수집 완료 (소요: {elapsed:.1f}초)")

    # DB 저장
    sales_date = datetime.now().strftime("%Y-%m-%d")
    repo = SalesRepository(store_id=store_id)
    stats = repo.save_daily_sales(data, sales_date, store_id=store_id)
    print(f"[매출수집] DB 저장: 총 {stats.get('total', 0)}건 (신규 {stats.get('new', 0)}, 갱신 {stats.get('updated', 0)})")

    # 매출 메뉴 탭 닫기
    try:
        from src.utils.nexacro_helpers import close_tab_by_frame_id
        close_tab_by_frame_id(analyzer.driver, "STMB011_M0")
        time.sleep(1)
    except Exception:
        pass

    return True


# =============================================================================
# Step 2: Direct API 재고 덤프
# =============================================================================
def dump_bgf_inventory(
    driver: Any,
    item_codes: List[str],
    store_id: str,
    save_to_db: bool = True,
) -> Dict[str, Dict[str, Any]]:
    """Direct API로 전체 상품 재고 일괄 조회

    Args:
        save_to_db: True면 조회 결과를 DB에도 저장 (재고 갱신 후 비교 가능)
    """
    from src.collectors.order_prep_collector import OrderPrepCollector

    collector = OrderPrepCollector(driver=driver, save_to_db=save_to_db, store_id=store_id)

    # 메뉴 이동 + 날짜 선택 (Direct API 템플릿 캡처에 필요)
    print("  메뉴 이동: 발주 > 단품별 발주...")
    if not collector.navigate_to_menu():
        print("[ERROR] 메뉴 이동 실패")
        return {}

    print("  발주일 선택...")
    if not collector.select_order_date():
        print("[ERROR] 날짜 선택 실패")
        return {}

    # Direct API 배치 수집
    total = len(item_codes)
    all_results = {}

    # 청크 분할
    chunks = [item_codes[i:i + CHUNK_SIZE] for i in range(0, total, CHUNK_SIZE)]
    print(f"  Direct API 배치: {total}개 상품, {len(chunks)}개 청크 (각 {CHUNK_SIZE}개)")

    for ci, chunk in enumerate(chunks, 1):
        chunk_start = time.time()
        print(f"  청크 [{ci}/{len(chunks)}] {len(chunk)}개 처리 중...", end="", flush=True)

        try:
            results = collector.collect_for_items(chunk, use_direct_api=True)
            success = sum(1 for v in results.values() if v.get('success'))
            chunk_elapsed = time.time() - chunk_start
            print(f" 성공 {success}/{len(chunk)} ({chunk_elapsed:.1f}초)")
            all_results.update(results)
        except Exception as e:
            chunk_elapsed = time.time() - chunk_start
            print(f" 오류: {e} ({chunk_elapsed:.1f}초)")

    # 메뉴 닫기
    try:
        collector.close_menu()
    except Exception:
        pass

    return all_results


# =============================================================================
# Step 3: DB vs BGF 비교
# =============================================================================
def compare_inventory(
    db_items: List[Dict[str, Any]],
    bgf_data: Dict[str, Dict[str, Any]],
    threshold: int,
) -> List[Dict[str, Any]]:
    """DB 재고와 BGF 재고 비교"""
    results = []
    db_map = {item["item_cd"]: item for item in db_items}

    # DB에 있는 상품 비교
    for item_cd, db_item in db_map.items():
        db_qty = db_item.get("stock_qty") or 0
        item_nm = db_item.get("item_nm") or ""
        queried_at = db_item.get("queried_at") or ""
        stale = _is_stale(queried_at)

        bgf_item = bgf_data.get(item_cd)

        if not bgf_item or not bgf_item.get("success"):
            results.append({
                "item_cd": item_cd,
                "item_nm": item_nm,
                "db_stock_qty": db_qty,
                "bgf_now_qty": None,
                "gap": None,
                "status": STATUS_QUERY_FAIL,
                "queried_at": queried_at,
                "is_stale": stale,
            })
            continue

        bgf_qty = bgf_item.get("current_stock", 0)
        gap = bgf_qty - db_qty
        status = STATUS_MISMATCH if abs(gap) >= threshold else STATUS_MATCH

        results.append({
            "item_cd": item_cd,
            "item_nm": item_nm or bgf_item.get("item_nm", ""),
            "db_stock_qty": db_qty,
            "bgf_now_qty": bgf_qty,
            "gap": gap,
            "status": status,
            "queried_at": queried_at,
            "is_stale": stale,
        })

    # BGF에만 존재하는 상품 (DB 미등록)
    for item_cd, bgf_item in bgf_data.items():
        if item_cd not in db_map and bgf_item.get("success"):
            bgf_qty = bgf_item.get("current_stock", 0)
            results.append({
                "item_cd": item_cd,
                "item_nm": bgf_item.get("item_nm", ""),
                "db_stock_qty": 0,
                "bgf_now_qty": bgf_qty,
                "gap": bgf_qty,
                "status": STATUS_BGF_ONLY,
                "queried_at": "",
                "is_stale": False,
            })

    return results


# =============================================================================
# 리포트 출력
# =============================================================================
def _print_report(results: List[Dict[str, Any]], total_db_items: int) -> None:
    """콘솔 리포트"""
    sep = "-" * 110

    # 불일치 항목만 먼저 출력
    mismatches = [r for r in results if r["status"] in (STATUS_MISMATCH, STATUS_BGF_ONLY)]
    fails = [r for r in results if r["status"] == STATUS_QUERY_FAIL]
    matches = [r for r in results if r["status"] == STATUS_MATCH]

    if mismatches:
        print(f"\n{'=' * 110}")
        print(f"[불일치 상세] ({len(mismatches)}건)")
        print(sep)
        print(f"{'No.':>4}  {'상품코드':<16} {'상품명':<22} {'DB재고':>6} {'BGF재고':>7} {'차이':>6} {'상태':<12} {'DB조회시각':<14}")
        print(sep)

        # 갭 절대값 기준 내림차순
        mismatches.sort(key=lambda r: abs(r.get("gap") or 0), reverse=True)
        for i, r in enumerate(mismatches, 1):
            bgf_str = str(r["bgf_now_qty"]) if r["bgf_now_qty"] is not None else "-"
            gap_val = r.get("gap") or 0
            gap_str = f"+{gap_val}" if gap_val > 0 else str(gap_val)
            qa = r.get("queried_at") or ""
            qa_short = ""
            if qa:
                try:
                    dt = datetime.fromisoformat(qa)
                    qa_short = dt.strftime("%m-%d %H:%M")
                except (ValueError, TypeError):
                    qa_short = qa[:14]
            stale_mark = " [STALE]" if r["is_stale"] else ""
            print(
                f"{i:>4}  {r['item_cd']:<16} {(r['item_nm'] or '')[:20]:<22} "
                f"{r['db_stock_qty']:>6} {bgf_str:>7} {gap_str:>6} "
                f"{r['status']:<12} {qa_short}{stale_mark}"
            )
        print(sep)

    # 요약 통계
    total = len(results)
    n_match = len(matches)
    n_mismatch = len(mismatches)
    n_fail = len(fails)
    stales = sum(1 for r in results if r["is_stale"])

    match_pct = (n_match / total * 100) if total else 0
    mismatch_pct = (n_mismatch / total * 100) if total else 0
    fail_pct = (n_fail / total * 100) if total else 0

    mismatch_with_gap = [r for r in mismatches if r["gap"] is not None]
    bgf_higher = [r for r in mismatch_with_gap if r["gap"] > 0]
    bgf_lower = [r for r in mismatch_with_gap if r["gap"] < 0]

    print(f"\n{'=' * 110}")
    print("[검증 요약]")
    print(sep)
    print(f"  DB 전체 상품:          {total_db_items}개")
    print(f"  검증 대상:             {total}개")
    print(f"  일치 (MATCH):          {n_match}개 ({match_pct:.1f}%)")
    print(f"  불일치 (MISMATCH):     {n_mismatch}개 ({mismatch_pct:.1f}%)")
    print(f"  조회실패 (FAIL):       {n_fail}개 ({fail_pct:.1f}%)")

    if mismatch_with_gap:
        gaps = [abs(r["gap"]) for r in mismatch_with_gap]
        avg_gap = sum(gaps) / len(gaps)
        max_gap_item = max(mismatch_with_gap, key=lambda r: abs(r["gap"]))

        print()
        print(f"  불일치 상세:")
        if bgf_higher:
            total_pos = sum(r["gap"] for r in bgf_higher)
            print(f"    BGF > DB (DB 과소): {len(bgf_higher)}개, 총 갭 +{total_pos}")
        if bgf_lower:
            total_neg = sum(r["gap"] for r in bgf_lower)
            print(f"    BGF < DB (DB 과다): {len(bgf_lower)}개, 총 갭 {total_neg}")
        gap_val = max_gap_item["gap"]
        gap_display = f"+{gap_val}" if gap_val > 0 else str(gap_val)
        print(f"    최대 갭: {gap_display} ({max_gap_item['item_cd']} {(max_gap_item['item_nm'] or '')[:16]})")
        print(f"    평균 갭: {avg_gap:.1f}")

    if stales:
        print(f"\n  경고: DB 데이터 스테일 (>{STALE_HOURS}h): {stales}개")

    print(f"{'=' * 110}")


def _write_csv(results: List[Dict[str, Any]], output_path: Path) -> None:
    """CSV 파일 저장"""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "item_cd", "item_nm", "db_stock_qty", "bgf_now_qty",
            "gap", "status", "queried_at", "is_stale",
        ])
        writer.writeheader()
        writer.writerows(results)
    print(f"\nCSV 저장: {output_path}")


def _write_excel(
    all_store_results: Dict[str, Dict[str, Any]],
    output_path: Path,
) -> None:
    """전 매장 결과를 엑셀 파일로 저장 (매장별 시트 + 요약 시트)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── 색상/스타일 정의 ──
    header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    match_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    stale_font = Font(name="맑은 고딕", color="999999", italic=True)
    num_font = Font(name="맑은 고딕", size=10)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    def apply_header(ws, headers, col_widths):
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ── 요약 시트 ──
    ws_summary = wb.active
    ws_summary.title = "요약"
    summary_headers = ["매장코드", "매장명", "검증상품", "일치", "불일치",
                       "일치율(%)", "BGF>DB", "BGF<DB", "총갭", "STALE"]
    summary_widths = [10, 18, 10, 10, 10, 10, 10, 10, 10, 10]
    apply_header(ws_summary, summary_headers, summary_widths)

    row_idx = 2
    for store_id, store_data in all_store_results.items():
        results = store_data.get("results", [])
        store_name = store_data.get("store_name", store_id)

        matches = [r for r in results if r["status"] == STATUS_MATCH]
        mismatches = [r for r in results if r["status"] in (STATUS_MISMATCH, STATUS_BGF_ONLY)]
        stales = sum(1 for r in results if r["is_stale"])
        bgf_higher = [r for r in mismatches if (r.get("gap") or 0) > 0]
        bgf_lower = [r for r in mismatches if (r.get("gap") or 0) < 0]
        total_gap = sum(abs(r.get("gap") or 0) for r in mismatches)
        total = len(results)
        match_pct = (len(matches) / total * 100) if total else 0

        vals = [store_id, store_name, total, len(matches), len(mismatches),
                round(match_pct, 1), len(bgf_higher), len(bgf_lower), total_gap, stales]
        for ci, v in enumerate(vals, 1):
            cell = ws_summary.cell(row=row_idx, column=ci, value=v)
            cell.font = num_font
            cell.border = thin_border
            if ci == 6:  # 일치율
                cell.number_format = "0.0"
                if match_pct >= 95:
                    cell.fill = match_fill
                elif match_pct < 80:
                    cell.fill = mismatch_fill
        row_idx += 1

    # ── 매장별 상세 시트 ──
    detail_headers = ["No", "상품코드", "상품명", "DB재고", "BGF재고",
                      "차이", "상태", "DB조회시각", "STALE"]
    detail_widths = [6, 16, 24, 8, 8, 8, 10, 20, 6]

    for store_id, store_data in all_store_results.items():
        results = store_data.get("results", [])
        store_name = store_data.get("store_name", store_id)

        # 불일치 먼저, 갭 절대값 큰 순 → 일치
        mismatches = sorted(
            [r for r in results if r["status"] != STATUS_MATCH],
            key=lambda r: abs(r.get("gap") or 0), reverse=True
        )
        matches = [r for r in results if r["status"] == STATUS_MATCH]
        sorted_results = mismatches + matches

        sheet_name = f"{store_id}"[:31]  # 시트명 31자 제한
        ws = wb.create_sheet(title=sheet_name)
        apply_header(ws, detail_headers, detail_widths)

        for ri, r in enumerate(sorted_results, 1):
            row_num = ri + 1
            bgf_qty = r["bgf_now_qty"] if r["bgf_now_qty"] is not None else ""
            gap_val = r.get("gap")
            gap_display = gap_val if gap_val is not None else ""

            qa = r.get("queried_at") or ""
            qa_short = ""
            if qa:
                try:
                    dt = datetime.fromisoformat(qa)
                    qa_short = dt.strftime("%Y-%m-%d %H:%M")
                except (ValueError, TypeError):
                    qa_short = qa[:16]

            vals = [ri, r["item_cd"], r.get("item_nm") or "", r["db_stock_qty"],
                    bgf_qty, gap_display, r["status"], qa_short,
                    "Y" if r["is_stale"] else ""]

            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=ci, value=v)
                cell.font = num_font
                cell.border = thin_border

                # 행 색상
                if r["status"] in (STATUS_MISMATCH, STATUS_BGF_ONLY):
                    cell.fill = mismatch_fill
                elif r["status"] == STATUS_MATCH:
                    cell.fill = match_fill
                if r["is_stale"]:
                    cell.font = stale_font

    # 저장
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"\n엑셀 저장: {output_path}")


def sync_db_from_results(
    results: List[Dict[str, Any]],
    store_id: str,
) -> int:
    """불일치 상품의 DB 재고를 BGF 값으로 동기화"""
    import sqlite3

    mismatches = [r for r in results
                  if r["status"] in (STATUS_MISMATCH, STATUS_BGF_ONLY)
                  and r["bgf_now_qty"] is not None]

    if not mismatches:
        return 0

    db_path = project_root / "data" / "stores" / f"{store_id}.db"
    conn = sqlite3.connect(str(db_path))
    now = datetime.now().isoformat()
    updated = 0

    try:
        cursor = conn.cursor()
        for r in mismatches:
            cursor.execute(
                "UPDATE realtime_inventory SET stock_qty = ?, queried_at = ? "
                "WHERE store_id = ? AND item_cd = ?",
                (r["bgf_now_qty"], now, store_id, r["item_cd"])
            )
            if cursor.rowcount > 0:
                updated += 1
        conn.commit()
    finally:
        conn.close()

    return updated


# =============================================================================
# 메인 실행
# =============================================================================
def run_verification_single(
    store_id: str,
    max_items: Optional[int] = DEFAULT_MAX_ITEMS,
    item_codes: Optional[List[str]] = None,
    threshold: int = 1,
    collect_sales: bool = False,
    sync_db: bool = False,
) -> Dict[str, Any]:
    """단일 매장 Direct API 재고 검증

    Returns:
        {"store_id", "store_name", "results", "synced"}
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # 매장명 조회
    store_name = store_id
    try:
        from src.settings.store_context import StoreContext
        ctx = StoreContext.from_store_id(store_id)
        store_name = ctx.store_name or store_id
    except Exception:
        pass

    print(f"\n{'=' * 80}")
    print(f"  재고 검증: {store_id} ({store_name})")
    print(f"  시작: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'=' * 80}")

    # 1. DB 재고 로드
    print("\n  [1/4] DB 재고 로드...")
    repo = RealtimeInventoryRepository(store_id=store_id)
    db_items = repo.get_all(available_only=True, exclude_cut=True, store_id=store_id)
    total_db = len(db_items)
    print(f"    취급 상품: {total_db}개")

    if total_db == 0 and not item_codes:
        print("  [ERROR] DB에 취급 상품이 없습니다.")
        return {"store_id": store_id, "store_name": store_name, "results": [], "synced": 0}

    # 검증 대상 상품 코드
    if item_codes:
        target_codes = item_codes
    elif max_items:
        with_stock = [it for it in db_items if (it.get("stock_qty") or 0) > 0]
        without_stock = [it for it in db_items if (it.get("stock_qty") or 0) == 0]
        import random
        random.shuffle(with_stock)
        random.shuffle(without_stock)
        combined = with_stock + without_stock
        target_codes = [it["item_cd"] for it in combined[:max_items]]
    else:
        target_codes = [it["item_cd"] for it in db_items]

    n = len(target_codes)
    est_sec = n * 0.15
    print(f"  [2/4] 검증 대상: {n}개 (예상 {est_sec:.0f}초)")

    # 2. BGF 로그인
    analyzer = None
    results = []
    synced = 0

    try:
        print(f"  [3/4] BGF 로그인 (store_id={store_id})...")
        from src.sales_analyzer import SalesAnalyzer

        analyzer = SalesAnalyzer(store_id=store_id)
        analyzer.setup_driver()
        analyzer.connect()

        if not analyzer.do_login():
            print("  [ERROR] 로그인 실패")
            return {"store_id": store_id, "store_name": store_name, "results": [], "synced": 0}
        print("    로그인 성공")
        time.sleep(2)

        try:
            analyzer.close_popup()
        except Exception:
            pass

        # 매출수집 (선택적)
        if collect_sales:
            print("\n  --- 매출수집 ---")
            run_sales_collection(analyzer, store_id)
            time.sleep(2)

        # Direct API 재고 덤프
        print(f"  [3/4] Direct API 재고 덤프 ({n}개)...")
        dump_start = time.time()
        bgf_data = dump_bgf_inventory(analyzer.driver, target_codes, store_id, save_to_db=False)
        dump_elapsed = time.time() - dump_start

        success_count = sum(1 for v in bgf_data.values() if v.get("success"))
        print(f"    완료: 성공 {success_count}/{n} ({dump_elapsed:.1f}초)")

        # 비교
        print(f"  [4/4] DB vs BGF 비교...")
        results = compare_inventory(db_items, bgf_data, threshold)
        _print_report(results, total_db)

        # DB 동기화 (옵션)
        if sync_db:
            synced = sync_db_from_results(results, store_id)
            print(f"\n  [SYNC] DB 동기화: {synced}건 업데이트")

    except Exception as e:
        print(f"\n  [ERROR] {store_id}: {e}")
        traceback.print_exc()

    finally:
        if analyzer:
            try:
                analyzer.close()
            except Exception:
                pass

    return {"store_id": store_id, "store_name": store_name, "results": results, "synced": synced}


def run_verification_all_stores(
    threshold: int = 1,
    sync_db: bool = False,
) -> str:
    """전 매장 재고 검증 + 엑셀 내보내기

    Returns:
        엑셀 파일 경로
    """
    from src.settings.store_context import StoreContext

    stores = StoreContext.get_all_active()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print(f"\n{'#' * 80}")
    print(f"  전 매장 재고 검증")
    print(f"  대상: {len(stores)}개 매장")
    print(f"  시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'#' * 80}")

    all_results = {}

    for idx, ctx in enumerate(stores, 1):
        print(f"\n  ▶ [{idx}/{len(stores)}] {ctx.store_id} ({ctx.store_name})")
        result = run_verification_single(
            store_id=ctx.store_id,
            threshold=threshold,
            sync_db=sync_db,
        )
        all_results[ctx.store_id] = result

    # 엑셀 저장
    excel_path = OUTPUT_DIR / f"inventory_verify_{timestamp}.xlsx"
    _write_excel(all_results, excel_path)

    # 최종 요약
    print(f"\n{'=' * 80}")
    print(f"  전체 검증 완료")
    print(f"  {'매장':10s} | {'검증':>6s} | {'일치':>6s} | {'불일치':>6s} | {'일치율':>7s} | {'동기화':>6s}")
    print(f"  {'-' * 55}")
    for sid, data in all_results.items():
        results = data.get("results", [])
        total = len(results)
        matches = sum(1 for r in results if r["status"] == STATUS_MATCH)
        mismatches = total - matches
        pct = (matches / total * 100) if total else 0
        synced = data.get("synced", 0)
        print(f"  {sid:10s} | {total:>6d} | {matches:>6d} | {mismatches:>6d} | {pct:>6.1f}% | {synced:>6d}")
    print(f"\n  엑셀: {excel_path}")
    print(f"{'=' * 80}")

    return str(excel_path)


# legacy 호환
def run_verification(
    max_items: Optional[int] = DEFAULT_MAX_ITEMS,
    item_codes: Optional[List[str]] = None,
    store_id: str = DEFAULT_STORE_ID,
    save_csv: bool = False,
    threshold: int = 1,
    collect_sales: bool = False,
    sync_db: bool = False,
) -> None:
    """단일 매장 검증 (기존 인터페이스 호환)"""
    result = run_verification_single(
        store_id=store_id,
        max_items=max_items,
        item_codes=item_codes,
        threshold=threshold,
        collect_sales=collect_sales,
        sync_db=sync_db,
    )

    if save_csv:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        csv_path = OUTPUT_DIR / f"verify_inventory_v2_{timestamp}.csv"
        _write_csv(result.get("results", []), csv_path)


# =============================================================================
# CLI
# =============================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="재고 검증 v2: Direct API 배치 덤프 → DB realtime_inventory 비교"
    )
    parser.add_argument(
        "--all-stores", action="store_true",
        help="전 매장 검증 + 엑셀 내보내기"
    )
    parser.add_argument(
        "--max-items", type=int, default=DEFAULT_MAX_ITEMS,
        help="검증할 최대 상품 수 (기본: 전체)"
    )
    parser.add_argument(
        "--item-cd", nargs="+", type=str,
        help="특정 상품코드만 검증 (공백 구분)"
    )
    parser.add_argument(
        "--store-id", type=str, default=DEFAULT_STORE_ID,
        help=f"점포 코드 (기본: {DEFAULT_STORE_ID})"
    )
    parser.add_argument(
        "--csv", action="store_true",
        help="CSV 파일로 결과 저장"
    )
    parser.add_argument(
        "--threshold", type=int, default=1,
        help="불일치 판정 임계값 (기본: 1)"
    )
    parser.add_argument(
        "--collect-sales", action="store_true",
        help="검증 전 매출수집 실행"
    )
    parser.add_argument(
        "--sync", action="store_true",
        help="불일치 상품 DB 동기화 (BGF 재고로 덮어쓰기)"
    )

    args = parser.parse_args()

    if args.all_stores:
        run_verification_all_stores(
            threshold=args.threshold,
            sync_db=args.sync,
        )
    else:
        run_verification(
            max_items=args.max_items,
            item_codes=args.item_cd,
            store_id=args.store_id,
            save_csv=args.csv,
            threshold=args.threshold,
            collect_sales=args.collect_sales,
            sync_db=args.sync,
        )
