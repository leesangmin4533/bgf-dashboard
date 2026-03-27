"""
푸드 발주 평가 엑셀 리포트 생성 스크립트
기간: 2026-03-01 ~ 2026-03-07

v2: 1차/2차 판매 윈도우 분리 + 만성 품절위험 시트 + 매장47863 표시 수정
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 설정 ──
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
COMMON_DB = os.path.join(DATA_DIR, "common.db")
STORES_DIR = os.path.join(DATA_DIR, "stores")
OUTPUT_PATH = os.path.join(
    DATA_DIR, "exports", "food_order_evaluation_0301_0307.xlsx"
)

# ── 설정 ──
ORDER_START = "2026-03-01"
ORDER_END = "2026-03-07"
SALES_END = "2026-03-09"  # 2차 상품 마지막 폐기까지
FOOD_MID_PREFIXES = ("001", "002", "003", "004", "005", "012")

# ── 스타일 ──
HEADER_FILL = PatternFill(
    start_color="4472C4", end_color="4472C4", fill_type="solid"
)
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
RED_FILL = PatternFill(
    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
)
RED_FONT = Font(name="맑은 고딕", color="9C0006")
YELLOW_FILL = PatternFill(
    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
)
YELLOW_FONT = Font(name="맑은 고딕", color="9C6500")
GREEN_FILL = PatternFill(
    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
)
GREEN_FONT = Font(name="맑은 고딕", color="006100")
ORANGE_FILL = PatternFill(
    start_color="F4B084", end_color="F4B084", fill_type="solid"
)
ORANGE_FONT = Font(name="맑은 고딕", bold=True, color="843C0C")
NORMAL_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_store_names():
    """common.db에서 매장명 조회, 미등록 매장은 '{store_id}점' 표시"""
    conn = sqlite3.connect(COMMON_DB)
    rows = conn.execute("SELECT store_id, store_name FROM stores").fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}


def get_store_dbs():
    """활성 매장 DB 파일 목록 반환"""
    store_dbs = []
    for f in os.listdir(STORES_DIR):
        if (
            f.endswith(".db")
            and not f.endswith(".db.db")
            and "backup" not in f
            and "bak" not in f
            and f not in ("store.db", "TEST.db", "99999.db")
        ):
            store_id = f.replace(".db", "")
            db_path = os.path.join(STORES_DIR, f)
            store_dbs.append((store_id, db_path))
    return store_dbs


def determine_delivery_type(item_nm):
    """상품명 끝자리로 차수 판별"""
    if not item_nm or not item_nm.strip():
        return None
    last_char = item_nm.strip()[-1]
    if last_char == "1":
        return "1차"
    elif last_char == "2":
        return "2차"
    return None


def get_sales_dates_for_order(order_date_str, delivery_type):
    """
    발주일 기준 판매/폐기 조회 날짜 결정 (차수별 분리)

    1차 상품 (당일 20:00 도착 ~ 발주일+1일 22:00):
      → sales_date = order_date+1 (주 판매일: 도착 다음날 하루)

    2차 상품 (익일 07:00 도착 ~ 발주일+2일 02:00):
      → sales_date = order_date+1 (도착일 하루 판매)
      → sales_date = order_date+2 (익일 02:00까지 폐기 반영)
    """
    order_date = datetime.strptime(order_date_str, "%Y-%m-%d")

    if delivery_type == "1차":
        # 당일 20:00 도착 → 다음날이 주 판매일
        return [(order_date + timedelta(days=1)).strftime("%Y-%m-%d")]
    else:
        # 익일 07:00 도착 → 도착일 + 그 다음날 02:00 폐기분
        d1 = (order_date + timedelta(days=1)).strftime("%Y-%m-%d")
        d2 = (order_date + timedelta(days=2)).strftime("%Y-%m-%d")
        return [d1, d2]


def evaluate(order_qty, sale_qty, disuse_qty):
    """평가 판정"""
    if order_qty <= 0:
        return "데이터없음"
    waste_rate = (disuse_qty / order_qty) * 100
    sell_rate = (sale_qty / order_qty) * 100
    if waste_rate > 20:
        return "과발주"
    elif sell_rate >= 95:
        return "품절위험"
    else:
        return "적정"


def collect_data():
    """전 매장 푸드 발주 평가 데이터 수집 (1차/2차 판매 윈도우 분리)"""
    store_names = get_store_names()
    store_dbs = get_store_dbs()
    all_records = []

    mid_filter = " OR ".join(
        [f"ot.mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES]
    )

    for store_id, db_path in store_dbs:
        # 매장명: DB 등록 → 미등록이면 "{store_id}점"
        store_name = store_names.get(store_id, f"{store_id}점")
        conn = sqlite3.connect(db_path)

        # 테이블 존재 확인
        tables = [
            r[0]
            for r in conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()
        ]
        if "order_tracking" not in tables or "daily_sales" not in tables:
            conn.close()
            continue

        # 1) 푸드 발주 데이터 조회
        orders = conn.execute(
            f"""
            SELECT ot.order_date, ot.item_cd, ot.item_nm, ot.mid_cd,
                   ot.delivery_type, ot.order_qty
            FROM order_tracking ot
            WHERE ot.order_date BETWEEN ? AND ?
              AND ({mid_filter})
            ORDER BY ot.order_date, ot.item_cd
            """,
            (ORDER_START, ORDER_END),
        ).fetchall()

        # 2) 판매/폐기 데이터 로드 (item_cd, sales_date 기준)
        sales_data = {}
        sales_rows = conn.execute(
            f"""
            SELECT ds.sales_date, ds.item_cd, ds.sale_qty, ds.disuse_qty
            FROM daily_sales ds
            WHERE ds.sales_date BETWEEN ? AND ?
              AND ({mid_filter.replace('ot.', 'ds.')})
            """,
            (ORDER_START, SALES_END),
        ).fetchall()

        for row in sales_rows:
            key = (row[0], row[1])  # (sales_date, item_cd)
            if key not in sales_data:
                sales_data[key] = {
                    "sale_qty": row[2] or 0,
                    "disuse_qty": row[3] or 0,
                }
            else:
                sales_data[key]["sale_qty"] = max(
                    sales_data[key]["sale_qty"], row[2] or 0
                )
                sales_data[key]["disuse_qty"] = max(
                    sales_data[key]["disuse_qty"], row[3] or 0
                )

        # 3) 발주별 평가 (차수별 판매 윈도우 분리)
        for order_date, item_cd, item_nm, mid_cd, db_delivery, order_qty in orders:
            delivery = determine_delivery_type(item_nm)
            if delivery is None:
                if db_delivery in ("1차", "2차"):
                    delivery = db_delivery
                else:
                    continue

            order_qty = order_qty or 0
            if order_qty <= 0:
                continue

            # 차수별 판매 윈도우: 해당 날짜들의 판매/폐기 합산
            sales_dates = get_sales_dates_for_order(order_date, delivery)
            sale_qty = 0
            disuse_qty = 0
            for sd in sales_dates:
                info = sales_data.get((sd, item_cd), {})
                sale_qty += info.get("sale_qty", 0)
                disuse_qty += info.get("disuse_qty", 0)

            waste_rate = (disuse_qty / order_qty * 100) if order_qty > 0 else 0
            sell_rate = (sale_qty / order_qty * 100) if order_qty > 0 else 0
            eval_result = evaluate(order_qty, sale_qty, disuse_qty)

            all_records.append(
                {
                    "store_id": store_id,
                    "store_name": store_name,
                    "order_date": order_date,
                    "item_cd": item_cd,
                    "item_nm": item_nm,
                    "delivery": delivery,
                    "order_qty": order_qty,
                    "sale_qty": sale_qty,
                    "disuse_qty": disuse_qty,
                    "sell_rate": round(sell_rate, 1),
                    "waste_rate": round(waste_rate, 1),
                    "eval": eval_result,
                }
            )

        conn.close()

    return all_records


def apply_eval_style(ws, cell, eval_result):
    """평가 결과에 따른 셀 스타일 적용"""
    if eval_result == "과발주":
        cell.fill = RED_FILL
        cell.font = RED_FONT
    elif eval_result == "품절위험":
        cell.fill = YELLOW_FILL
        cell.font = YELLOW_FONT
    elif eval_result == "적정":
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT


def write_sheet1_detail(wb, records):
    """시트 1 - 상세 데이터"""
    ws = wb.active
    ws.title = "상세 데이터"

    headers = [
        "매장명",
        "발주일",
        "상품명",
        "차수",
        "발주량",
        "실판매량",
        "폐기량",
        "판매율(%)",
        "폐기율(%)",
        "평가",
    ]
    col_widths = [20, 12, 35, 8, 10, 10, 10, 12, 12, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    for row_idx, rec in enumerate(records, 2):
        values = [
            rec["store_name"],
            rec["order_date"],
            rec["item_nm"],
            rec["delivery"],
            rec["order_qty"],
            rec["sale_qty"],
            rec["disuse_qty"],
            rec["sell_rate"],
            rec["waste_rate"],
            rec["eval"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col >= 5:
                cell.alignment = Alignment(horizontal="center")
            if col == 10:
                apply_eval_style(ws, cell, val)

    ws.auto_filter.ref = f"A1:J{len(records) + 1}"
    ws.freeze_panes = "A2"


def write_sheet2_summary(wb, records):
    """시트 2 - 매장별 요약"""
    ws = wb.create_sheet("매장별 요약")

    store_stats = defaultdict(
        lambda: {
            "store_name": "",
            "total_order": 0,
            "total_sale": 0,
            "total_disuse": 0,
            "over_count": 0,
            "stockout_count": 0,
            "total_count": 0,
            "waste_rates": [],
        }
    )

    for rec in records:
        sid = rec["store_id"]
        st = store_stats[sid]
        st["store_name"] = rec["store_name"]
        st["total_order"] += rec["order_qty"]
        st["total_sale"] += rec["sale_qty"]
        st["total_disuse"] += rec["disuse_qty"]
        st["total_count"] += 1
        st["waste_rates"].append(rec["waste_rate"])
        if rec["eval"] == "과발주":
            st["over_count"] += 1
        elif rec["eval"] == "품절위험":
            st["stockout_count"] += 1

    headers = [
        "매장명",
        "총발주량",
        "총판매량",
        "총폐기량",
        "평균폐기율(%)",
        "과발주건수",
        "품절위험건수",
        "종합평가점수",
    ]
    col_widths = [20, 12, 12, 12, 14, 14, 14, 14]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    row_idx = 2
    for sid, st in sorted(store_stats.items()):
        avg_waste = (
            round(sum(st["waste_rates"]) / len(st["waste_rates"]), 1)
            if st["waste_rates"]
            else 0
        )
        total = st["total_count"] or 1
        over_ratio = st["over_count"] / total
        stockout_ratio = st["stockout_count"] / total
        score = round(
            100
            - (over_ratio * 40 + stockout_ratio * 30 + avg_waste / 100 * 30),
            1,
        )
        score = max(0, min(100, score))

        values = [
            st["store_name"],
            st["total_order"],
            st["total_sale"],
            st["total_disuse"],
            avg_waste,
            st["over_count"],
            st["stockout_count"],
            score,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col >= 2:
                cell.alignment = Alignment(horizontal="center")
            if col == 8:
                if val >= 80:
                    cell.fill = GREEN_FILL
                    cell.font = Font(
                        name="맑은 고딕", bold=True, color="006100"
                    )
                elif val >= 60:
                    cell.fill = YELLOW_FILL
                    cell.font = Font(
                        name="맑은 고딕", bold=True, color="9C6500"
                    )
                else:
                    cell.fill = RED_FILL
                    cell.font = Font(
                        name="맑은 고딕", bold=True, color="9C0006"
                    )
        row_idx += 1

    ws.freeze_panes = "A2"


def write_sheet3_trend(wb, records):
    """시트 3 - 일별 트렌드"""
    ws = wb.create_sheet("일별 트렌드")

    daily_store = defaultdict(
        lambda: {"waste_rates": [], "order_qty": 0, "disuse_qty": 0}
    )
    stores_set = set()
    dates_set = set()

    for rec in records:
        key = (rec["order_date"], rec["store_name"])
        daily_store[key]["waste_rates"].append(rec["waste_rate"])
        daily_store[key]["order_qty"] += rec["order_qty"]
        daily_store[key]["disuse_qty"] += rec["disuse_qty"]
        stores_set.add(rec["store_name"])
        dates_set.add(rec["order_date"])

    stores = sorted(stores_set)
    dates = sorted(dates_set)

    headers = ["발주일"]
    for store in stores:
        headers.extend(
            [
                f"{store}\n폐기율(%)",
                f"{store}\n발주량",
                f"{store}\n폐기량",
            ]
        )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 12
    for i in range(len(stores)):
        for j in range(3):
            ws.column_dimensions[get_column_letter(2 + i * 3 + j)].width = 14

    ws.row_dimensions[1].height = 35

    for row_idx, date in enumerate(dates, 2):
        ws.cell(row=row_idx, column=1, value=date).font = NORMAL_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = Alignment(
            horizontal="center"
        )

        for s_idx, store in enumerate(stores):
            key = (date, store)
            info = daily_store.get(
                key, {"waste_rates": [], "order_qty": 0, "disuse_qty": 0}
            )
            avg_waste = (
                round(
                    sum(info["waste_rates"]) / len(info["waste_rates"]), 1
                )
                if info["waste_rates"]
                else 0
            )

            base_col = 2 + s_idx * 3
            cell_w = ws.cell(row=row_idx, column=base_col, value=avg_waste)
            cell_w.font = NORMAL_FONT
            cell_w.border = THIN_BORDER
            cell_w.alignment = Alignment(horizontal="center")
            if avg_waste > 20:
                cell_w.fill = RED_FILL
                cell_w.font = RED_FONT
            elif avg_waste > 10:
                cell_w.fill = YELLOW_FILL
                cell_w.font = YELLOW_FONT
            else:
                cell_w.fill = GREEN_FILL
                cell_w.font = GREEN_FONT

            cell_o = ws.cell(
                row=row_idx, column=base_col + 1, value=info["order_qty"]
            )
            cell_o.font = NORMAL_FONT
            cell_o.border = THIN_BORDER
            cell_o.alignment = Alignment(horizontal="center")

            cell_d = ws.cell(
                row=row_idx, column=base_col + 2, value=info["disuse_qty"]
            )
            cell_d.font = NORMAL_FONT
            cell_d.border = THIN_BORDER
            cell_d.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B2"


def write_sheet4_chronic_stockout(wb, records):
    """시트 4 - 만성 품절위험 (3회 이상 반복 상품)"""
    ws = wb.create_sheet("만성 품절위험")

    # 상품명 × 매장별 품절위험 횟수 집계
    stockout_by_product = defaultdict(
        lambda: {
            "count": 0,
            "stores": defaultdict(int),
            "dates": [],
            "total_order": 0,
            "total_sale": 0,
            "avg_sell_rate": [],
        }
    )

    for rec in records:
        if rec["eval"] == "품절위험":
            item_key = rec["item_nm"]
            info = stockout_by_product[item_key]
            info["count"] += 1
            info["stores"][rec["store_name"]] += 1
            info["dates"].append(rec["order_date"])
            info["total_order"] += rec["order_qty"]
            info["total_sale"] += rec["sale_qty"]
            info["avg_sell_rate"].append(rec["sell_rate"])

    # 3회 이상만 필터
    chronic = {
        k: v for k, v in stockout_by_product.items() if v["count"] >= 3
    }

    # 헤더
    headers = [
        "상품명",
        "품절위험 횟수",
        "해당 매장",
        "발생일",
        "총발주량",
        "총판매량",
        "평균판매율(%)",
        "심각도",
    ]
    col_widths = [35, 14, 25, 30, 12, 12, 14, 12]

    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = PatternFill(
            start_color="ED7D31", end_color="ED7D31", fill_type="solid"
        )
        cell.font = Font(
            name="맑은 고딕", bold=True, color="FFFFFF", size=11
        )
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    # 횟수 내림차순 정렬
    sorted_chronic = sorted(
        chronic.items(), key=lambda x: x[1]["count"], reverse=True
    )

    row_idx = 2
    for item_nm, info in sorted_chronic:
        avg_sell = (
            round(sum(info["avg_sell_rate"]) / len(info["avg_sell_rate"]), 1)
            if info["avg_sell_rate"]
            else 0
        )
        stores_str = ", ".join(
            f"{s}({c}회)"
            for s, c in sorted(
                info["stores"].items(), key=lambda x: x[1], reverse=True
            )
        )
        unique_dates = sorted(set(info["dates"]))
        dates_str = ", ".join(unique_dates)

        # 심각도 판정
        if info["count"] >= 6:
            severity = "매우심각"
        elif info["count"] >= 5:
            severity = "심각"
        elif info["count"] >= 4:
            severity = "주의"
        else:
            severity = "경고"

        values = [
            item_nm,
            info["count"],
            stores_str,
            dates_str,
            info["total_order"],
            info["total_sale"],
            avg_sell,
            severity,
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if col in (2, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")
            if col == 8:
                if severity == "매우심각":
                    cell.fill = RED_FILL
                    cell.font = Font(
                        name="맑은 고딕", bold=True, color="9C0006"
                    )
                elif severity == "심각":
                    cell.fill = ORANGE_FILL
                    cell.font = ORANGE_FONT
                elif severity == "주의":
                    cell.fill = YELLOW_FILL
                    cell.font = YELLOW_FONT
                else:
                    cell.fill = PatternFill(
                        start_color="DDEBF7",
                        end_color="DDEBF7",
                        fill_type="solid",
                    )
                    cell.font = Font(
                        name="맑은 고딕", bold=True, color="2F5496"
                    )
        row_idx += 1

    # 요약 행
    row_idx += 1
    summary_cell = ws.cell(
        row=row_idx,
        column=1,
        value=f"만성 품절위험 상품 총 {len(chronic)}개 (3회 이상 반복)",
    )
    summary_cell.font = Font(name="맑은 고딕", bold=True, size=11)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{row_idx - 1}"


def main():
    print("푸드 발주 평가 리포트 v2 생성 시작...")
    print(f"기간: {ORDER_START} ~ {ORDER_END}")
    print("개선사항: 1차/2차 판매윈도우 분리, 만성 품절위험 시트 추가")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    records = collect_data()
    print(f"총 {len(records)}건 수집 완료")

    if not records:
        print("데이터가 없습니다.")
        return

    stores = set(r["store_name"] for r in records)
    evals = defaultdict(int)
    for r in records:
        evals[r["eval"]] += 1
    print(f"매장 수: {len(stores)}")
    print(f"평가 분포: {dict(evals)}")

    # 품절위험 상품 통계
    stockout_items = defaultdict(int)
    for r in records:
        if r["eval"] == "품절위험":
            stockout_items[r["item_nm"]] += 1
    chronic_count = sum(1 for c in stockout_items.values() if c >= 3)
    print(f"품절위험 상품: {len(stockout_items)}종, 만성(3회+): {chronic_count}종")

    wb = Workbook()
    write_sheet1_detail(wb, records)
    write_sheet2_summary(wb, records)
    write_sheet3_trend(wb, records)
    write_sheet4_chronic_stockout(wb, records)

    wb.save(OUTPUT_PATH)
    print(f"\n리포트 저장 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
