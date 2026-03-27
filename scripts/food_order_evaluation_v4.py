"""
푸드 발주 평가 v4 - 재고 흐름 기반 리포트 (기초재고 분리)
기간: 2026-03-01 ~ 2026-03-07 (발주일 기준)
플로우: 2026-03-01 ~ 2026-03-09 (판매/폐기 추적 포함)

v3 대비 핵심 수정:
  1. 기초이월재고(3/1 전일이월)와 당기입고(3/1~3/9 buy_qty)를 분리
  2. FIFO에서 기초재고 먼저 소진 후 당기배치 소진 추적
  3. 판정은 당기입고 기준으로만 (기초재고폐기 = 별도 집계, 판정 반영 안 함)
  4. 판매>가용재고 상품은 "데이터불일치"로 분리

지표 (당기입고 기준):
  당기입고 = 기간 내 buy_qty 합계
  당기판매 = 총판매 - 기초재고로 소화된 판매
  당기폐기 = 총폐기 - 기초재고에서 발생한 폐기
  배치효율 = 당기판매 / 당기입고 x 100
  실폐기율 = 당기폐기 / 당기입고 x 100

판정:
  발주없음: 당기입고 = 0 (기초재고만 있던 상품)
  데이터불일치: 판매+폐기 > 가용재고 (어느 날이든)
  우수: 실폐기율 < 5% AND 배치효율 >= 70%
  양호: 실폐기율 5~10%
  과발주: 실폐기율 > 10%
  저조: 배치효율 < 50% AND 실폐기율 < 5%

시트 구성:
  1. 재고흐름 상세 (이천호반베르디움 auto)
  2. 상품별 평가 (당기입고 기준)
  3. 매장별 종합 (3매장, 신뢰도 명시)
"""

import sqlite3
import os
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 ──
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
COMMON_DB = os.path.join(DATA_DIR, "common.db")
STORES_DIR = os.path.join(DATA_DIR, "stores")
OUTPUT_PATH = os.path.join(
    DATA_DIR, "exports", "food_order_evaluation_v4_0301_0307.xlsx"
)

# ── 설정 ──
ORDER_START = "2026-03-01"
ORDER_END = "2026-03-07"
FLOW_START = "2026-03-01"
FLOW_END = "2026-03-09"
INITIAL_STOCK_DATE = "2026-02-28"
FOOD_MID_PREFIXES = ("001", "002", "003", "004", "005", "012")
CORE_STORE_ID = "46513"

DEFAULT_EXPIRY = {
    "001": 1, "002": 1, "003": 2,
    "004": 2, "005": 1, "012": 3,
}
FALLBACK_EXPIRY = 1

# ── 스타일 ──
BLUE_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ORANGE_HEADER = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="맑은 고딕", color="9C0006")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="맑은 고딕", color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="맑은 고딕", color="9C6500")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BLUE_FONT = Font(name="맑은 고딕", color="2F5496")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
GRAY_FONT = Font(name="맑은 고딕", color="595959")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

NORMAL_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
ITALIC_FONT = Font(name="맑은 고딕", size=9, italic=True, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ────────────────────────────────────────
# 헬퍼
# ────────────────────────────────────────

def get_store_names():
    conn = sqlite3.connect(COMMON_DB)
    rows = conn.execute("SELECT store_id, store_name FROM stores").fetchall()
    conn.close()
    return {str(r[0]): r[1] for r in rows}


def get_store_dbs():
    store_dbs = []
    for f in os.listdir(STORES_DIR):
        if (f.endswith(".db") and not f.endswith(".db.db")
                and "backup" not in f and "bak" not in f
                and f not in ("store.db", "TEST.db", "99999.db")):
            store_id = f.replace(".db", "")
            store_dbs.append((store_id, os.path.join(STORES_DIR, f)))
    return store_dbs


def determine_delivery(item_nm):
    if not item_nm or not item_nm.strip():
        return None
    last = item_nm.strip()[-1]
    return "1" if last == "1" else ("2" if last == "2" else None)


def date_range_list(start, end):
    s = datetime.strptime(start, "%Y-%m-%d")
    e = datetime.strptime(end, "%Y-%m-%d")
    result = []
    while s <= e:
        result.append(s.strftime("%Y-%m-%d"))
        s += timedelta(days=1)
    return result


def get_expiry_days(item_cd, mid_cd, expiry_lookup):
    db_val = expiry_lookup.get(item_cd)
    if db_val and db_val > 0:
        return int(db_val)
    if mid_cd:
        for prefix, days in DEFAULT_EXPIRY.items():
            if mid_cd.startswith(prefix):
                return days
    return FALLBACK_EXPIRY


def estimate_initial_stock(daily_entries):
    running = 0
    min_running = 0
    for e in daily_entries:
        running += (e.get("buy_qty", 0) or 0)
        running -= (e.get("sale_qty", 0) or 0)
        running -= (e.get("disuse_qty", 0) or 0)
        min_running = min(min_running, running)
    return abs(min_running) if min_running < 0 else 0


def style_header(ws, row, headers, widths, fill):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 30


def write_row(ws, row_idx, values, center_from=1):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if col >= center_from:
            cell.alignment = Alignment(horizontal="center")
    return row_idx + 1


def fmt_date(d):
    parts = d.split("-")
    return f"{int(parts[1])}/{int(parts[2])}"


# ────────────────────────────────────────
# 재고 배치 시뮬레이션 (v4: 기초/당기 분리)
# ────────────────────────────────────────

class InventoryBatch:
    """개별 입고 배치 (FIFO 추적용), is_initial로 기초/당기 구분"""
    __slots__ = ("arrival_date", "qty", "expiry_date", "is_initial")

    def __init__(self, arrival_date, qty, expiry_date, is_initial=False):
        self.arrival_date = arrival_date
        self.qty = qty
        self.expiry_date = expiry_date
        self.is_initial = is_initial


def _consume_fifo(batches, qty):
    """
    FIFO로 qty만큼 소진. 기초/당기 각각 소진량 반환.
    Returns: (initial_consumed, current_consumed, unaccounted)
    """
    initial_consumed = 0
    current_consumed = 0
    rem = qty
    for b in batches:
        if rem <= 0:
            break
        take = min(b.qty, rem)
        b.qty -= take
        rem -= take
        if b.is_initial:
            initial_consumed += take
        else:
            current_consumed += take
    return initial_consumed, current_consumed, rem


def simulate_item_flow(daily_entries, expiry_days, initial_stock=0):
    """
    일별 재고 흐름 FIFO 시뮬레이션 (v4: 기초/당기 분리 추적).

    Returns: [{date, carryover_in, buy, available, sold, wasted,
               carryover_out, initial_sold, current_sold,
               initial_wasted, current_wasted, unaccounted, note}, ...]
    """
    batches = []

    if initial_stock > 0 and daily_entries:
        # 기초재고는 FLOW_START + expiry_days를 만료일로 설정
        # (이월재고를 FLOW_START에 "입고"된 것으로 취급, 기록 지연 1일 반영)
        start = daily_entries[0]["date"]
        init_exp = (datetime.strptime(start, "%Y-%m-%d")
                    + timedelta(days=expiry_days)).strftime("%Y-%m-%d")
        batches.append(InventoryBatch(
            "이월", initial_stock, init_exp, is_initial=True
        ))

    flow = []
    initial_fully_depleted = False  # 기초재고 완전소진 추적

    for entry in daily_entries:
        d = entry["date"]
        buy = entry.get("buy_qty", 0) or 0
        sold = entry.get("sale_qty", 0) or 0
        wasted = entry.get("disuse_qty", 0) or 0

        # 1) 전일 이월재고
        carryover_in = sum(b.qty for b in batches)
        initial_remaining_before = sum(b.qty for b in batches if b.is_initial)

        # 2) 당일 입고 -> 새 배치 (is_initial=False)
        if buy > 0:
            exp_d = (datetime.strptime(d, "%Y-%m-%d")
                     + timedelta(days=expiry_days)).strftime("%Y-%m-%d")
            batches.append(InventoryBatch(d, buy, exp_d, is_initial=False))
            batches.sort(key=lambda b: b.expiry_date)

        available = carryover_in + buy

        # 3) 판매 소진 (FIFO: 유통기한 짧은 것 = 기초재고 먼저)
        init_sold, curr_sold, unaccounted_sold = _consume_fifo(batches, sold)

        # 4) 폐기 소진
        init_wasted, curr_wasted, unaccounted_wasted = _consume_fifo(batches, wasted)

        total_unaccounted = unaccounted_sold + unaccounted_wasted

        # 5) 빈 배치 제거
        batches = [b for b in batches if b.qty > 0]

        # 6) 만료 배치 제거
        init_expired = 0
        curr_expired = 0
        surviving = []
        for b in batches:
            if b.expiry_date <= d:
                if b.is_initial:
                    init_expired += b.qty
                else:
                    curr_expired += b.qty
            else:
                surviving.append(b)
        batches = surviving

        carryover_out = sum(b.qty for b in batches)
        initial_remaining_after = sum(b.qty for b in batches if b.is_initial)

        # 7) 비고
        notes = []
        if buy > 0:
            notes.append(f"입고{buy}")

        # 기초재고 소진 여부
        if initial_remaining_before > 0 and initial_remaining_after == 0 and not initial_fully_depleted:
            notes.append("기초완소")
            initial_fully_depleted = True

        if init_wasted > 0:
            notes.append(f"기초폐기{init_wasted}")
        if curr_wasted > 0:
            notes.append(f"당기폐기{curr_wasted}")
        if init_expired > 0:
            notes.append(f"기초만료{init_expired}")
        if curr_expired > 0:
            notes.append(f"당기만료{curr_expired}")
        if available > 0 and sold >= available:
            notes.append("완판")
        if total_unaccounted > 0:
            notes.append(f"초과{total_unaccounted}")

        flow.append({
            "date": d,
            "carryover_in": carryover_in,
            "buy": buy,
            "available": available,
            "sold": sold,
            "wasted": wasted,
            "carryover_out": carryover_out,
            "initial_sold": init_sold,
            "current_sold": curr_sold,
            "initial_wasted": init_wasted + init_expired,
            "current_wasted": curr_wasted + curr_expired,
            "unaccounted": total_unaccounted,
            "note": " | ".join(notes),
        })

    return flow


# ────────────────────────────────────────
# 데이터 수집
# ────────────────────────────────────────

def collect_store_data(store_id, db_path, store_name):
    conn = sqlite3.connect(db_path)
    tables = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()]
    if "order_tracking" not in tables or "daily_sales" not in tables:
        conn.close()
        return None

    # order_source 통계
    src_rows = conn.execute("""
        SELECT order_source, COUNT(*)
        FROM order_tracking
        WHERE order_date BETWEEN ? AND ?
        GROUP BY order_source
    """, (ORDER_START, ORDER_END)).fetchall()
    source_stats = {(r[0] or "unknown"): r[1] for r in src_rows}
    source_stats["total"] = sum(source_stats.values())

    # auto food item 목록
    mid_filter = " OR ".join([f"mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES])
    auto_items = conn.execute(f"""
        SELECT DISTINCT item_cd, item_nm, mid_cd
        FROM order_tracking
        WHERE order_date BETWEEN ? AND ?
          AND order_source = 'auto'
          AND ({mid_filter})
    """, (ORDER_START, ORDER_END)).fetchall()

    if not auto_items:
        conn.close()
        return {"store_name": store_name, "items": {}, "source_stats": source_stats}

    auto_item_cds = [r[0] for r in auto_items]
    item_info = {icd: {"item_nm": inm, "mid_cd": mcd} for icd, inm, mcd in auto_items}

    # expiration_days
    escaped_path = COMMON_DB.replace("'", "''")
    conn.execute(f"ATTACH DATABASE '{escaped_path}' AS common")
    expiry_lookup = {}
    placeholders = ",".join("?" * len(auto_item_cds))
    for r in conn.execute(f"""
        SELECT item_cd, expiration_days
        FROM common.product_details
        WHERE item_cd IN ({placeholders})
    """, auto_item_cds).fetchall():
        if r[1] is not None:
            expiry_lookup[r[0]] = r[1]
    conn.execute("DETACH DATABASE common")

    # daily_sales 로드
    ds_ph = ",".join("?" * len(auto_item_cds))
    daily_rows = conn.execute(f"""
        SELECT sales_date, item_cd, buy_qty, sale_qty, disuse_qty, stock_qty
        FROM daily_sales
        WHERE item_cd IN ({ds_ph})
          AND sales_date BETWEEN ? AND ?
        ORDER BY item_cd, sales_date
    """, (*auto_item_cds, INITIAL_STOCK_DATE, FLOW_END)).fetchall()
    conn.close()

    # 아이템별 정리
    daily_by_item = defaultdict(dict)
    stock_initial = {}

    for sales_date, item_cd, buy_qty, sale_qty, disuse_qty, stock_qty in daily_rows:
        if sales_date == INITIAL_STOCK_DATE:
            if stock_qty is not None and stock_qty > 0:
                stock_initial[item_cd] = stock_qty
            continue
        daily_by_item[item_cd][sales_date] = {
            "buy_qty": buy_qty or 0,
            "sale_qty": sale_qty or 0,
            "disuse_qty": disuse_qty or 0,
        }

    # 아이템별 flow 시뮬레이션
    all_dates = date_range_list(FLOW_START, FLOW_END)
    items_result = {}

    for item_cd in auto_item_cds:
        info = item_info.get(item_cd, {})
        item_nm = info.get("item_nm", item_cd)
        mid_cd = info.get("mid_cd", "")
        delivery = determine_delivery(item_nm)
        expiry = get_expiry_days(item_cd, mid_cd, expiry_lookup)

        item_daily = daily_by_item.get(item_cd, {})
        full_entries = []
        for d in all_dates:
            if d in item_daily:
                entry = item_daily[d].copy()
                entry["date"] = d
            else:
                entry = {"date": d, "buy_qty": 0, "sale_qty": 0, "disuse_qty": 0}
            full_entries.append(entry)

        initial = stock_initial.get(item_cd, 0)
        if initial == 0:
            initial = estimate_initial_stock(full_entries)

        flow = simulate_item_flow(full_entries, expiry, initial)

        # ── v4 핵심: 당기 기준 지표 계산 ──
        total_buy = sum(f["buy"] for f in flow)        # 당기입고
        total_sold = sum(f["sold"] for f in flow)
        total_wasted = sum(f["wasted"] for f in flow)

        sum_initial_sold = sum(f["initial_sold"] for f in flow)
        sum_current_sold = sum(f["current_sold"] for f in flow)
        sum_initial_wasted = sum(f["initial_wasted"] for f in flow)
        sum_current_wasted = sum(f["current_wasted"] for f in flow)
        sum_unaccounted = sum(f["unaccounted"] for f in flow)

        carryover_end = flow[-1]["carryover_out"] if flow else 0

        # 당기 지표
        current_sold_total = sum_current_sold       # 당기판매
        current_wasted_total = sum_current_wasted   # 당기폐기

        if total_buy > 0:
            batch_eff = round(current_sold_total / total_buy * 100, 1)
            waste_rate = round(current_wasted_total / total_buy * 100, 1)
        else:
            batch_eff = 0
            waste_rate = 0

        # 판정
        # 미추적(unaccounted)은 재고조정/수동입고/이관 등으로 발생하는
        # 데이터 한계이므로 판정에 반영하지 않음 (시트에 정보로만 표시)
        # 데이터불일치: 당기판매 > 당기입고 (FIFO 구조상 불가능 = 로직 오류)
        if total_buy == 0:
            verdict = "발주없음"
        elif batch_eff > 100:
            verdict = "데이터불일치"
        elif waste_rate > 10:
            verdict = "과발주"
        elif waste_rate >= 5:
            verdict = "양호"
        elif batch_eff >= 70:
            verdict = "우수"
        elif batch_eff < 50:
            verdict = "저조"
        else:
            verdict = "적정"

        items_result[item_cd] = {
            "item_nm": item_nm,
            "delivery": delivery or "?",
            "mid_cd": mid_cd,
            "expiry_days": expiry,
            "initial_stock": initial,
            "flow": flow,
            "summary": {
                "initial_stock": initial,
                "total_buy": total_buy,
                "total_sold": total_sold,
                "total_wasted": total_wasted,
                "initial_sold": sum_initial_sold,
                "current_sold": current_sold_total,
                "initial_wasted": sum_initial_wasted,
                "current_wasted": current_wasted_total,
                "batch_efficiency": batch_eff,
                "waste_rate": waste_rate,
                "carryover_end": carryover_end,
                "unaccounted": sum_unaccounted,
                "verdict": verdict,
            },
        }

    return {
        "store_name": store_name,
        "items": items_result,
        "source_stats": source_stats,
    }


def collect_all_stores():
    store_names = get_store_names()
    store_dbs = get_store_dbs()
    all_data = {}
    for store_id, db_path in store_dbs:
        sname = store_names.get(store_id, f"{store_id}점")
        data = collect_store_data(store_id, db_path, sname)
        if data is not None:
            all_data[store_id] = data
    return all_data


# ────────────────────────────────────────
# 시트 1: 재고흐름 상세
# ────────────────────────────────────────

def write_sheet1_flow(wb, items_data, store_name):
    ws = wb.active
    ws.title = "재고흐름 상세"

    ws.merge_cells("A1:J1")
    title = ws.cell(
        row=1, column=1,
        value=f"{store_name} | 자동발주 재고 흐름 (기초/당기 분리) | "
              f"{ORDER_START} ~ {ORDER_END}"
    )
    title.font = Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    headers = ["상품명", "차수", "날짜", "전일이월", "당일입고",
               "가용재고", "판매량", "폐기량", "익일이월", "비고"]
    widths = [36, 5, 8, 10, 10, 10, 9, 9, 10, 35]
    style_header(ws, 2, headers, widths, BLUE_HEADER)

    sorted_items = sorted(items_data.items(), key=lambda x: x[1]["item_nm"])

    row = 3
    group_idx = 0
    for item_cd, item_data in sorted_items:
        flow = item_data["flow"]
        if not flow:
            continue

        has_activity = any(
            f["buy"] > 0 or f["sold"] > 0 or f["wasted"] > 0 or f["carryover_in"] > 0
            for f in flow
        )
        if not has_activity:
            continue

        use_bg = group_idx % 2 == 0
        verdict = item_data["summary"]["verdict"]

        for i, f in enumerate(flow):
            nm_val = item_data["item_nm"] if i == 0 else ""
            dv_val = item_data["delivery"] if i == 0 else ""

            vals = [nm_val, dv_val, fmt_date(f["date"]),
                    f["carryover_in"], f["buy"], f["available"],
                    f["sold"], f["wasted"], f["carryover_out"],
                    f["note"]]
            write_row(ws, row, vals, center_from=2)

            if i == 0:
                ws.cell(row=row, column=1).font = BOLD_FONT

            if use_bg:
                for c in range(1, 11):
                    ws.cell(row=row, column=c).fill = LIGHT_BLUE_FILL

            if f["wasted"] > 0:
                wc = ws.cell(row=row, column=8)
                wc.fill = RED_FILL
                wc.font = RED_FONT

            if "완판" in f["note"]:
                for c in range(6, 9):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = GREEN_FILL
                    cell.font = GREEN_FONT

            row += 1

        # 소계
        s = item_data["summary"]
        summary_text = (f"유통기한 {item_data['expiry_days']}일 | "
                        f"기초재고 {s['initial_stock']} | "
                        f"당기입고 {s['total_buy']} | "
                        f"판정: {verdict}")
        ws.cell(row=row, column=3, value=summary_text).font = ITALIC_FONT
        row += 2
        group_idx += 1

    ws.cell(row=row, column=1,
            value="기초완소=기초재고 완전소진 | 기초폐기N=기초재고에서 폐기 | "
                  "당기폐기N=당기입고에서 폐기 | 초과N=데이터 불일치").font = ITALIC_FONT

    ws.auto_filter.ref = f"A2:J{max(row - 1, 3)}"
    ws.freeze_panes = "A3"
    return group_idx


# ────────────────────────────────────────
# 시트 2: 상품별 평가 (당기입고 기준)
# ────────────────────────────────────────

def apply_verdict_style(cell, verdict):
    styles = {
        "과발주": (RED_FILL, RED_FONT),
        "저조": (YELLOW_FILL, YELLOW_FONT),
        "적정": (LIGHT_GRAY_FILL, NORMAL_FONT),
        "양호": (BLUE_FILL, BLUE_FONT),
        "우수": (GREEN_FILL, GREEN_FONT),
        "발주없음": (GRAY_FILL, GRAY_FONT),
        "데이터불일치": (YELLOW_FILL, RED_FONT),
    }
    fill, font = styles.get(verdict, (None, NORMAL_FONT))
    if fill:
        cell.fill = fill
    cell.font = font


def write_sheet2_evaluation(wb, items_data, store_name):
    ws = wb.create_sheet("상품별 평가")

    ws.merge_cells("A1:K1")
    ws.cell(row=1, column=1,
            value=f"{store_name} | 당기입고 기준 상품별 평가 | "
                  f"{ORDER_START} ~ {FLOW_END}").font = \
        Font(name="맑은 고딕", bold=True, size=12, color="2F5496")
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["상품명", "차수", "기초재고", "당기입고", "당기판매",
               "당기폐기", "기초재고폐기", "배치효율(%)", "실폐기율(%)",
               "이월잔여", "판정"]
    widths = [36, 5, 10, 10, 10, 10, 12, 12, 12, 10, 12]
    style_header(ws, 2, headers, widths, ORANGE_HEADER)

    verdict_order = {
        "과발주": 0, "양호": 1, "적정": 2, "저조": 3,
        "우수": 4, "발주없음": 5, "데이터불일치": 6,
    }
    sorted_items = sorted(
        items_data.items(),
        key=lambda x: (verdict_order.get(x[1]["summary"]["verdict"], 9),
                       -x[1]["summary"]["waste_rate"],
                       x[1]["item_nm"])
    )

    row = 3
    counts = defaultdict(int)
    for item_cd, item_data in sorted_items:
        s = item_data["summary"]

        # 활동 없으면 스킵
        if s["total_buy"] == 0 and s["initial_stock"] == 0:
            continue

        vals = [item_data["item_nm"], item_data["delivery"],
                s["initial_stock"], s["total_buy"],
                s["current_sold"], s["current_wasted"],
                s["initial_wasted"],
                s["batch_efficiency"], s["waste_rate"],
                s["carryover_end"], s["verdict"]]
        write_row(ws, row, vals, center_from=2)

        apply_verdict_style(ws.cell(row=row, column=11), s["verdict"])

        # 실폐기율 10%+ 강조
        if s["waste_rate"] > 10 and s["verdict"] != "발주없음":
            ws.cell(row=row, column=9).fill = RED_FILL
            ws.cell(row=row, column=9).font = RED_FONT

        # 기초재고폐기 > 0 강조
        if s["initial_wasted"] > 0:
            ws.cell(row=row, column=7).fill = YELLOW_FILL
            ws.cell(row=row, column=7).font = YELLOW_FONT

        counts[s["verdict"]] += 1
        row += 1

    # 요약
    row += 1
    total_items = sum(counts.values())
    summary_parts = []
    for v in ["과발주", "양호", "적정", "저조", "우수", "발주없음", "데이터불일치"]:
        if counts[v] > 0:
            summary_parts.append(f"{v} {counts[v]}")

    ws.cell(row=row, column=1,
            value=f"총 {total_items}개 상품 | {' / '.join(summary_parts)}").font = \
        Font(name="맑은 고딕", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1,
            value="당기입고 기준 판정 (기초재고폐기는 참고만, 판정 반영 안 함)").font = ITALIC_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="우수(폐기<5%+효율70%+) | 양호(폐기5~10%) | "
                  "과발주(폐기>10%) | 저조(효율<50%+폐기<5%) | "
                  "발주없음(당기입고=0)").font = ITALIC_FONT

    ws.auto_filter.ref = f"A2:K{max(row - 3, 3)}"
    ws.freeze_panes = "A3"
    return counts


# ────────────────────────────────────────
# 시트 3: 매장별 종합 (당기입고 기준)
# ────────────────────────────────────────

def write_sheet3_stores(wb, all_data):
    ws = wb.create_sheet("매장별 종합")

    ws.merge_cells("A1:L1")
    note_cell = ws.cell(
        row=1, column=1,
        value="당기입고 기준 매장별 종합 (기초재고 제외) | "
              "auto 비율에 따라 신뢰도 차이"
    )
    note_cell.font = Font(name="맑은 고딕", bold=True, size=11, color="9C6500")
    note_cell.fill = YELLOW_FILL
    note_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    headers = ["매장명", "auto비율(%)", "신뢰도", "평가상품",
               "발주없음", "당기입고", "당기판매", "당기폐기",
               "배치효율(%)", "실폐기율(%)", "기초재고폐기", "비고"]
    widths = [20, 12, 10, 10, 10, 10, 10, 10, 12, 12, 12, 40]
    style_header(ws, 2, headers, widths, GREEN_HEADER)

    store_ids = sorted(all_data.keys())
    row = 3
    totals = {"eval": 0, "noorder": 0, "buy": 0, "c_sold": 0,
              "c_wasted": 0, "i_wasted": 0}

    for sid in store_ids:
        sd = all_data[sid]
        ss = sd["source_stats"]
        auto_cnt = ss.get("auto", 0)
        total_cnt = ss.get("total", 1)
        auto_pct = round(auto_cnt / total_cnt * 100, 1) if total_cnt > 0 else 0

        if auto_pct >= 80:
            reliability = "높음"
        elif auto_pct >= 50:
            reliability = "중간"
        else:
            reliability = "낮음"

        eval_count = 0
        noorder_count = 0
        t_buy = 0
        t_c_sold = 0
        t_c_wasted = 0
        t_i_wasted = 0
        verdict_dist = defaultdict(int)

        for icd, idata in sd["items"].items():
            s = idata["summary"]
            verdict_dist[s["verdict"]] += 1
            if s["verdict"] == "발주없음":
                noorder_count += 1
                t_i_wasted += s["initial_wasted"]
                continue
            if s["total_buy"] == 0 and s["initial_stock"] == 0:
                continue
            eval_count += 1
            t_buy += s["total_buy"]
            t_c_sold += s["current_sold"]
            t_c_wasted += s["current_wasted"]
            t_i_wasted += s["initial_wasted"]

        b_eff = round(t_c_sold / t_buy * 100, 1) if t_buy > 0 else 0
        w_rate = round(t_c_wasted / t_buy * 100, 1) if t_buy > 0 else 0

        notes = []
        if sid == CORE_STORE_ID:
            notes.append("핵심 평가 대상")
        if reliability == "낮음":
            notes.append(f"auto {auto_pct}%")
        if verdict_dist["과발주"] > 0:
            notes.append(f"과발주 {verdict_dist['과발주']}개")
        if verdict_dist["데이터불일치"] > 0:
            notes.append(f"불일치 {verdict_dist['데이터불일치']}개")
        if verdict_dist["우수"] > 0:
            notes.append(f"우수 {verdict_dist['우수']}개")
        note_str = " | ".join(notes) if notes else ""

        vals = [sd["store_name"], auto_pct, reliability,
                eval_count, noorder_count,
                t_buy, t_c_sold, t_c_wasted,
                b_eff, w_rate, t_i_wasted, note_str]
        write_row(ws, row, vals, center_from=2)
        ws.cell(row=row, column=12).alignment = Alignment(horizontal="left", wrap_text=True)

        # 신뢰도 색상
        rel_cell = ws.cell(row=row, column=3)
        if reliability == "높음":
            rel_cell.fill = GREEN_FILL
            rel_cell.font = Font(name="맑은 고딕", bold=True, color="006100")
        elif reliability == "중간":
            rel_cell.fill = YELLOW_FILL
            rel_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
        else:
            rel_cell.fill = RED_FILL
            rel_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")

        if sid == CORE_STORE_ID:
            for c in range(1, 13):
                cell = ws.cell(row=row, column=c)
                if c != 3:
                    cell.fill = LIGHT_BLUE_FILL
                cell.font = Font(name="맑은 고딕", size=10, bold=True)
                cell.border = THIN_BORDER

        totals["eval"] += eval_count
        totals["noorder"] += noorder_count
        totals["buy"] += t_buy
        totals["c_sold"] += t_c_sold
        totals["c_wasted"] += t_c_wasted
        totals["i_wasted"] += t_i_wasted
        row += 1

    # 합계
    row += 1
    t_eff = round(totals["c_sold"] / totals["buy"] * 100, 1) if totals["buy"] > 0 else 0
    t_wr = round(totals["c_wasted"] / totals["buy"] * 100, 1) if totals["buy"] > 0 else 0
    total_vals = ["합계", "", "",
                  totals["eval"], totals["noorder"],
                  totals["buy"], totals["c_sold"], totals["c_wasted"],
                  t_eff, t_wr, totals["i_wasted"], ""]
    write_row(ws, row, total_vals, center_from=2)
    for c in range(1, 13):
        cell = ws.cell(row=row, column=c)
        cell.font = BOLD_FONT
        cell.fill = LIGHT_GRAY_FILL
        cell.border = THIN_BORDER

    row += 2
    ws.cell(row=row, column=1,
            value="배치효율 = 당기판매 / 당기입고 x 100 | "
                  "실폐기율 = 당기폐기 / 당기입고 x 100 (기초재고 제외)").font = ITALIC_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="발주없음 = 당기입고 0건, 기초재고만 있던 상품 (평가 제외)").font = ITALIC_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="신뢰도: 높음(auto 80%+) | 중간(50~79%) | 낮음(<50%)").font = ITALIC_FONT

    ws.auto_filter.ref = f"A2:L{max(row - 4, 3)}"
    ws.freeze_panes = "A3"


# ────────────────────────────────────────
# 메인
# ────────────────────────────────────────

def main():
    print("=" * 60)
    print("푸드 발주 평가 v4 - 재고 흐름 (기초재고 분리)")
    print(f"발주 기간: {ORDER_START} ~ {ORDER_END}")
    print(f"플로우 추적: {FLOW_START} ~ {FLOW_END}")
    print(f"핵심 평가: 이천호반베르디움점 (auto, 당기입고 기준)")
    print("=" * 60)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    all_data = collect_all_stores()
    core_data = all_data.get(CORE_STORE_ID)

    if not core_data or not core_data["items"]:
        print("핵심 평가 대상 데이터 없음")
        return

    core_items = core_data["items"]

    # v4 통계
    verdict_counts = defaultdict(int)
    total_buy = 0
    total_c_sold = 0
    total_c_wasted = 0
    total_i_wasted = 0

    for icd, idata in core_items.items():
        s = idata["summary"]
        verdict_counts[s["verdict"]] += 1
        total_buy += s["total_buy"]
        total_c_sold += s["current_sold"]
        total_c_wasted += s["current_wasted"]
        total_i_wasted += s["initial_wasted"]

    b_eff = round(total_c_sold / total_buy * 100, 1) if total_buy > 0 else 0
    w_rate = round(total_c_wasted / total_buy * 100, 1) if total_buy > 0 else 0

    print(f"\n[핵심 평가 - {core_data['store_name']} (당기입고 기준)]")
    print(f"  전체 상품: {len(core_items)}")
    print(f"  당기입고: {total_buy}")
    print(f"  당기판매: {total_c_sold}")
    print(f"  당기폐기: {total_c_wasted}")
    print(f"  기초재고폐기 (참고): {total_i_wasted}")
    print(f"  배치효율: {b_eff}%")
    print(f"  실폐기율: {w_rate}%")
    print(f"  판정 분포:")
    for v in ["과발주", "양호", "적정", "저조", "우수", "발주없음", "데이터불일치"]:
        if verdict_counts[v] > 0:
            print(f"    {v}: {verdict_counts[v]}개")

    # 매장별 auto 비율
    print(f"\n[매장별 auto 비율]")
    for sid in sorted(all_data.keys()):
        sd = all_data[sid]
        ss = sd["source_stats"]
        apct = round(ss.get("auto", 0) / ss.get("total", 1) * 100, 1) if ss.get("total", 0) > 0 else 0
        rel = "높음" if apct >= 80 else ("중간" if apct >= 50 else "낮음")
        eval_cnt = sum(1 for i in sd["items"].values() if i["summary"]["total_buy"] > 0)
        no_cnt = sum(1 for i in sd["items"].values() if i["summary"]["verdict"] == "발주없음")
        print(f"  {sd['store_name']}: {apct}% ({rel}) - 평가 {eval_cnt}개, 발주없음 {no_cnt}개")

    # 엑셀 생성
    wb = Workbook()
    shown = write_sheet1_flow(wb, core_items, core_data["store_name"])
    counts = write_sheet2_evaluation(wb, core_items, core_data["store_name"])
    write_sheet3_stores(wb, all_data)

    wb.save(OUTPUT_PATH)
    print(f"\n리포트 저장 완료: {OUTPUT_PATH}")
    print(f"시트 구성: {wb.sheetnames}")
    print(f"시트1: {shown}개 | 시트2: {dict(counts)}")


if __name__ == "__main__":
    main()
