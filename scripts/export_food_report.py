"""푸드 발주 vs 판매 비교 엑셀 리포트 (2026-03-01 ~ 03-23, 3매장)

컬럼 정의:
- 자동/수동/사이트 발주: order_tracking 기준
- 입고확정: receiving_history.receiving_qty (센터매입조회)
- 매입: daily_sales.buy_qty (BGF 판매현황)
- 실판매: daily_sales.sale_qty
- 폐기: daily_sales.disuse_qty
"""
import sqlite3, os, sys
from collections import defaultdict

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# === 설정 ===
stores = {"46704": "CU호반점", "46513": "CU동양대점", "47863": "CU세번째점"}
FOOD_MID_CDS = ("001", "002", "003", "004", "005", "012")
MID_NAMES = {"001": "도시락", "002": "주먹밥", "003": "김밥",
             "004": "샌드위치", "005": "햄버거", "012": "빵"}
DATE_FROM, DATE_TO = "2026-03-01", "2026-03-23"
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
common_db = os.path.join(BASE_DIR, "data", "common.db")
output_path = os.path.join(BASE_DIR, "data", "reports", "food_order_vs_sales_2026-03.xlsx")
os.makedirs(os.path.dirname(output_path), exist_ok=True)

# === 스타일 ===
header_font_white = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
auto_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
manual_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
mixed_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_row(ws, row, max_col, fill=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fill:
            cell.fill = fill


def get_order_fill(auto_q, manual_q, site_q):
    """발주 구분에 따른 배경색 반환"""
    if auto_q > 0 and (manual_q > 0 or site_q > 0):
        return mixed_fill
    elif auto_q > 0:
        return auto_fill
    elif manual_q > 0 or site_q > 0:
        return manual_fill
    return None


wb = Workbook()
wb.remove(wb.active)
mid_tuple_str = str(FOOD_MID_CDS)


def query_store_data(db_path):
    """매장 DB에서 발주/입고/판매 데이터를 모두 조회"""
    conn = sqlite3.connect(db_path)
    conn.execute(f"ATTACH DATABASE '{common_db}' AS common")
    c = conn.cursor()

    # 1) 발주 (order_tracking)
    c.execute(f"""
        SELECT order_date, mid_cd, order_source, SUM(order_qty), COUNT(*)
        FROM order_tracking
        WHERE order_date BETWEEN ? AND ? AND mid_cd IN {mid_tuple_str} AND order_qty > 0
        GROUP BY order_date, mid_cd, order_source
        ORDER BY order_date, mid_cd
    """, (DATE_FROM, DATE_TO))
    order_rows = c.fetchall()

    # 2) 판매 (daily_sales) — buy_qty 추가
    c.execute(f"""
        SELECT sales_date, mid_cd,
               SUM(sale_qty), SUM(buy_qty), SUM(disuse_qty), SUM(stock_qty), COUNT(*)
        FROM daily_sales
        WHERE sales_date BETWEEN ? AND ? AND mid_cd IN {mid_tuple_str}
        GROUP BY sales_date, mid_cd
        ORDER BY sales_date, mid_cd
    """, (DATE_FROM, DATE_TO))
    sales_rows = c.fetchall()

    # 3) 입고확정 (receiving_history)
    c.execute(f"""
        SELECT receiving_date, mid_cd, SUM(receiving_qty)
        FROM receiving_history
        WHERE receiving_date BETWEEN ? AND ? AND mid_cd IN {mid_tuple_str}
        GROUP BY receiving_date, mid_cd
        ORDER BY receiving_date, mid_cd
    """, (DATE_FROM, DATE_TO))
    recv_rows = c.fetchall()

    # 4) 상품별 발주
    c.execute(f"""
        SELECT ot.order_date, ot.item_cd, COALESCE(p.item_nm, ot.item_nm) as item_nm,
               ot.mid_cd, ot.order_source, ot.order_qty
        FROM order_tracking ot
        LEFT JOIN common.products p ON ot.item_cd = p.item_cd
        WHERE ot.order_date BETWEEN ? AND ? AND ot.mid_cd IN {mid_tuple_str} AND ot.order_qty > 0
        ORDER BY ot.order_date, ot.item_cd
    """, (DATE_FROM, DATE_TO))
    item_orders = c.fetchall()

    # 5) 상품별 판매 — buy_qty 추가
    c.execute(f"""
        SELECT ds.sales_date, ds.item_cd, COALESCE(p.item_nm, '') as item_nm,
               ds.mid_cd, ds.sale_qty, ds.buy_qty, ds.disuse_qty, ds.stock_qty
        FROM daily_sales ds
        LEFT JOIN common.products p ON ds.item_cd = p.item_cd
        WHERE ds.sales_date BETWEEN ? AND ? AND ds.mid_cd IN {mid_tuple_str}
        ORDER BY ds.sales_date, ds.item_cd
    """, (DATE_FROM, DATE_TO))
    item_sales = c.fetchall()

    # 6) 상품별 입고확정
    c.execute(f"""
        SELECT receiving_date, item_cd, SUM(receiving_qty)
        FROM receiving_history
        WHERE receiving_date BETWEEN ? AND ? AND mid_cd IN {mid_tuple_str}
        GROUP BY receiving_date, item_cd
    """, (DATE_FROM, DATE_TO))
    item_recv = c.fetchall()

    conn.close()
    return order_rows, sales_rows, recv_rows, item_orders, item_sales, item_recv


# ═══════════════════════════════════════════════════════════════
# 매장별 시트 (일자×중분류 요약)
# ═══════════════════════════════════════════════════════════════
for store_id, store_name in stores.items():
    db_path = os.path.join(BASE_DIR, "data", "stores", f"{store_id}.db")
    if not os.path.exists(db_path):
        print(f"{store_id}: DB not found, skip")
        continue

    ws = wb.create_sheet(title=f"{store_name}_{store_id}")
    order_rows, sales_rows, recv_rows, _, _, _ = query_store_data(db_path)

    # 발주 집계
    orders = defaultdict(lambda: {"auto": 0, "manual": 0, "site": 0,
                                   "auto_cnt": 0, "manual_cnt": 0, "site_cnt": 0})
    for date, mid, src, qty, cnt in order_rows:
        k = (date, mid)
        if src == "auto":
            orders[k]["auto"] += qty; orders[k]["auto_cnt"] += cnt
        elif src == "manual":
            orders[k]["manual"] += qty; orders[k]["manual_cnt"] += cnt
        elif src == "site":
            orders[k]["site"] += qty; orders[k]["site_cnt"] += cnt

    # 판매 집계 (buy_qty 추가)
    sales = {}
    for date, mid, sale, buy, disuse, stock, cnt in sales_rows:
        sales[(date, mid)] = {"sale": sale or 0, "buy": buy or 0,
                               "disuse": disuse or 0, "stock": stock or 0}

    # 입고확정 집계
    recv = {}
    for date, mid, qty in recv_rows:
        recv[(date, mid)] = qty or 0

    all_dates = sorted(set(
        d for d, m in list(orders.keys()) + list(sales.keys()) + list(recv.keys())
    ))

    headers = ["날짜", "중분류", "카테고리명",
               "자동발주(수량)", "자동발주(건수)",
               "수동발주(수량)", "수동발주(건수)",
               "사이트발주(수량)", "사이트발주(건수)",
               "발주합계",
               "입고확정(RH)", "매입(DS)",
               "실판매량", "폐기량", "재고량",
               "과잉률(%)"]
    for ci, h in enumerate(headers, 1):
        ws.cell(row=1, column=ci, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for date in all_dates:
        for mid in FOOD_MID_CDS:
            k = (date, mid)
            o = orders.get(k, {"auto": 0, "manual": 0, "site": 0,
                                "auto_cnt": 0, "manual_cnt": 0, "site_cnt": 0})
            s = sales.get(k, {"sale": 0, "buy": 0, "disuse": 0, "stock": 0})
            rh = recv.get(k, 0)

            auto_q, manual_q, site_q = o["auto"], o["manual"], o["site"]
            total_order = auto_q + manual_q + site_q
            sale_q = s["sale"]

            if total_order == 0 and sale_q == 0 and rh == 0:
                continue

            if sale_q > 0:
                excess = round((total_order - sale_q) / sale_q * 100, 1)
            elif total_order > 0:
                excess = "N/A"
            else:
                excess = "-"

            vals = [date, mid, MID_NAMES.get(mid, mid),
                    auto_q, o["auto_cnt"], manual_q, o["manual_cnt"],
                    site_q, o["site_cnt"], total_order,
                    rh, s["buy"],
                    sale_q, s["disuse"], s["stock"], excess]
            for ci, v in enumerate(vals, 1):
                ws.cell(row=row, column=ci, value=v)
            style_row(ws, row, len(headers), get_order_fill(auto_q, manual_q, site_q))
            row += 1

    widths = [12, 8, 10, 14, 14, 14, 14, 14, 14, 10, 12, 10, 10, 8, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    ws.freeze_panes = "A2"
    print(f"  {store_name}: {row - 2}행")

# ═══════════════════════════════════════════════════════════════
# 상품별 상세 시트 (3매장 통합)
# ═══════════════════════════════════════════════════════════════
ws2 = wb.create_sheet(title="상품별상세_3매장통합")
headers2 = ["매장", "날짜", "상품코드", "상품명", "중분류", "카테고리",
            "자동발주", "수동발주", "사이트발주", "발주합계",
            "입고확정(RH)", "매입(DS)",
            "실판매", "폐기", "재고", "발주구분"]
for ci, h in enumerate(headers2, 1):
    ws2.cell(row=1, column=ci, value=h)
style_header(ws2, 1, len(headers2))

row2 = 2
for store_id, store_name in stores.items():
    db_path = os.path.join(BASE_DIR, "data", "stores", f"{store_id}.db")
    if not os.path.exists(db_path):
        continue

    _, _, _, item_orders, item_sales, item_recv = query_store_data(db_path)

    # 상품별 발주 집계
    order_map = defaultdict(lambda: {"auto": 0, "manual": 0, "site": 0, "nm": "", "mid": ""})
    for date, item, nm, mid, src, qty in item_orders:
        k = (date, item)
        order_map[k]["nm"] = nm
        order_map[k]["mid"] = mid
        if src == "auto":
            order_map[k]["auto"] += qty
        elif src == "manual":
            order_map[k]["manual"] += qty
        elif src == "site":
            order_map[k]["site"] += qty

    # 상품별 판매 집계 (buy_qty 추가)
    sale_map = {}
    for date, item, nm, mid, sale, buy, disuse, stock in item_sales:
        sale_map[(date, item)] = {"sale": sale or 0, "buy": buy or 0,
                                   "disuse": disuse or 0, "stock": stock or 0,
                                   "nm": nm, "mid": mid}

    # 상품별 입고확정 집계
    recv_map = {}
    for date, item, qty in item_recv:
        recv_map[(date, item)] = qty or 0

    all_keys = sorted(set(
        list(order_map.keys()) + list(sale_map.keys()) + list(recv_map.keys())
    ))

    for date, item in all_keys:
        o = order_map.get((date, item), {"auto": 0, "manual": 0, "site": 0, "nm": "", "mid": ""})
        s = sale_map.get((date, item), {"sale": 0, "buy": 0, "disuse": 0, "stock": 0, "nm": "", "mid": ""})
        rh = recv_map.get((date, item), 0)

        nm = o["nm"] or s["nm"]
        mid = o["mid"] or s["mid"]
        auto_q, manual_q, site_q = o["auto"], o["manual"], o["site"]
        total = auto_q + manual_q + site_q

        sources = []
        if auto_q > 0: sources.append("자동")
        if manual_q > 0: sources.append("수동")
        if site_q > 0: sources.append("사이트")
        label = "+".join(sources) if sources else "판매만"

        vals = [store_name, date, item, nm, mid, MID_NAMES.get(mid, mid),
                auto_q, manual_q, site_q, total,
                rh, s["buy"],
                s["sale"], s["disuse"], s["stock"], label]
        for ci, v in enumerate(vals, 1):
            ws2.cell(row=row2, column=ci, value=v)
        style_row(ws2, row2, len(headers2), get_order_fill(auto_q, manual_q, site_q))
        row2 += 1

widths2 = [14, 12, 16, 25, 8, 10, 10, 10, 12, 10, 12, 10, 10, 8, 8, 12]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers2))}{row2 - 1}"
ws2.freeze_panes = "A2"

wb.save(output_path)
print(f"\n 엑셀 생성 완료: {output_path}")
print(f" 시트: {[ws.title for ws in wb.worksheets]}")
print(f" 상품별상세: {row2 - 1}행")
