"""
푸드 발주 평가 v2 엑셀 리포트 (개선판)
기간: 2026-03-01 ~ 2026-03-07

평가 철학: "폐기를 최소화하면서 팔 수 있는 만큼 파는 것"
- 입고량(buy_qty) 기준 평가
- 폐기 발생 = 과발주 (가장 중요)
- 완판 반복 = 수요초과 가능성 (관찰 대상)
- order_source별 분리 집계 + 이중발주 감지 + 입고 미기록 추적
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 경로 ──
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
COMMON_DB = os.path.join(DATA_DIR, "common.db")
STORES_DIR = os.path.join(DATA_DIR, "stores")
OUTPUT_PATH = os.path.join(
    DATA_DIR, "exports", "food_order_evaluation_v2_0301_0307.xlsx"
)

# ── 설정 ──
ORDER_START = "2026-03-01"
ORDER_END = "2026-03-07"
SALES_END = "2026-03-09"
FOOD_MID_PREFIXES = ("001", "002", "003", "004", "005", "012")
SELLOUT_THRESHOLD_DAYS = 5  # 7일 중 N일 이상 완판 → 수요초과 가능성

# ── 스타일 ──
BLUE_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ORANGE_HEADER = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
GREEN_HEADER = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
PURPLE_HEADER = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
GRAY_HEADER = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
DARK_BLUE_HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
TEAL_HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
RED_FONT = Font(name="맑은 고딕", color="9C0006")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREEN_FONT = Font(name="맑은 고딕", color="006100")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
YELLOW_FONT = Font(name="맑은 고딕", color="9C6500")
BLUE_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
BLUE_FONT = Font(name="맑은 고딕", color="2F5496")
GRAY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
GRAY_FONT = Font(name="맑은 고딕", color="375623")
LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

NORMAL_FONT = Font(name="맑은 고딕", size=10)
BOLD_FONT = Font(name="맑은 고딕", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

SOURCE_LABELS = {"auto": "자동", "site": "사이트", "manual": "수동", "receiving": "입고감지"}


# ────────────────────────────────────────
# 데이터 수집
# ────────────────────────────────────────

def get_store_names():
    conn = sqlite3.connect(COMMON_DB)
    rows = conn.execute("SELECT store_id, store_name FROM stores").fetchall()
    conn.close()
    return {r[0]: r[1] for r in rows}


def get_store_dbs():
    store_dbs = []
    for f in os.listdir(STORES_DIR):
        if (f.endswith(".db") and not f.endswith(".db.db")
                and "backup" not in f and "bak" not in f
                and f not in ("store.db", "TEST.db", "99999.db")):
            store_id = f.replace(".db", "")
            store_dbs.append((store_id, os.path.join(STORES_DIR, f)))
    return store_dbs


def determine_delivery(item_nm):
    if not item_nm or not item_nm.strip():
        return None
    last = item_nm.strip()[-1]
    return "1차" if last == "1" else ("2차" if last == "2" else None)


def get_sales_dates(order_date_str, delivery_type):
    od = datetime.strptime(order_date_str, "%Y-%m-%d")
    if delivery_type == "1차":
        return [(od + timedelta(days=1)).strftime("%Y-%m-%d")]
    else:
        return [
            (od + timedelta(days=1)).strftime("%Y-%m-%d"),
            (od + timedelta(days=2)).strftime("%Y-%m-%d"),
        ]


def collect_data():
    """Returns (records, buy0_records, duplicates, source_stats)"""
    store_names = get_store_names()
    store_dbs = get_store_dbs()
    all_records = []
    all_buy0_records = []
    all_duplicates = []
    all_source_stats = {}

    mid_filter_ot = " OR ".join([f"ot.mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES])
    mid_filter_ds = " OR ".join([f"ds.mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES])
    mid_filter_a = " OR ".join([f"a.mid_cd LIKE '{p}%'" for p in FOOD_MID_PREFIXES])

    for store_id, db_path in store_dbs:
        store_name = store_names.get(store_id, f"{store_id}점")
        conn = sqlite3.connect(db_path)
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table'"
        ).fetchall()]
        if "order_tracking" not in tables or "daily_sales" not in tables:
            conn.close()
            continue

        # ── order_source별 건수 집계 (전체 카테고리) ──
        src_rows = conn.execute("""
            SELECT order_source, COUNT(*)
            FROM order_tracking
            WHERE order_date BETWEEN ? AND ?
            GROUP BY order_source
        """, (ORDER_START, ORDER_END)).fetchall()
        store_src = {"store_name": store_name}
        total_all = 0
        for src, cnt in src_rows:
            store_src[src or "unknown"] = cnt
            total_all += cnt
        store_src["total"] = total_all
        all_source_stats[store_id] = store_src

        # ── 이중발주 감지 (auto + manual/site 동일 날짜/상품, 전체 카테고리) ──
        dupes = conn.execute("""
            SELECT a.order_date, a.item_cd, a.item_nm, a.mid_cd,
                   a.delivery_type,
                   a.order_source, a.order_qty,
                   b.order_source, b.order_qty
            FROM order_tracking a
            JOIN order_tracking b
              ON a.order_date = b.order_date
             AND a.item_cd = b.item_cd
             AND a.id < b.id
            WHERE a.order_date BETWEEN ? AND ?
              AND a.order_source = 'auto'
              AND b.order_source IN ('manual', 'site')
            ORDER BY a.order_date, a.item_cd
        """, (ORDER_START, ORDER_END)).fetchall()

        # ── 발주 데이터 (order_source 포함) ──
        orders = conn.execute(f"""
            SELECT ot.order_date, ot.item_cd, ot.item_nm, ot.mid_cd,
                   ot.delivery_type, ot.order_qty, ot.order_source
            FROM order_tracking ot
            WHERE ot.order_date BETWEEN ? AND ?
              AND ({mid_filter_ot})
            ORDER BY ot.order_date, ot.item_cd
        """, (ORDER_START, ORDER_END)).fetchall()

        # ── 판매/폐기/입고 데이터 로드 ──
        sales_data = {}
        for row in conn.execute(f"""
            SELECT ds.sales_date, ds.item_cd, ds.sale_qty, ds.disuse_qty, ds.buy_qty
            FROM daily_sales ds
            WHERE ds.sales_date BETWEEN ? AND ?
              AND ({mid_filter_ds})
        """, (ORDER_START, SALES_END)).fetchall():
            key = (row[0], row[1])
            if key not in sales_data:
                sales_data[key] = {"sale_qty": row[2] or 0, "disuse_qty": row[3] or 0, "buy_qty": row[4] or 0}
            else:
                sales_data[key]["sale_qty"] = max(sales_data[key]["sale_qty"], row[2] or 0)
                sales_data[key]["disuse_qty"] = max(sales_data[key]["disuse_qty"], row[3] or 0)
                sales_data[key]["buy_qty"] = max(sales_data[key]["buy_qty"], row[4] or 0)

        # ── hourly_sales_detail 로드 ──
        hourly_data = defaultdict(lambda: defaultdict(int))
        if "hourly_sales_detail" in tables:
            for row in conn.execute("""
                SELECT sales_date, item_cd, hour, sale_qty
                FROM hourly_sales_detail
                WHERE sales_date BETWEEN ? AND ?
            """, (ORDER_START, SALES_END)).fetchall():
                hourly_data[(row[0], row[1])][row[2]] += row[3] or 0

        # ── 레코드 생성 ──
        for order_date, item_cd, item_nm, mid_cd, db_delivery, order_qty, order_source in orders:
            delivery = determine_delivery(item_nm)
            if delivery is None:
                delivery = db_delivery if db_delivery in ("1차", "2차") else None
            if delivery is None:
                continue
            order_qty = order_qty or 0
            if order_qty <= 0:
                continue

            sales_dates = get_sales_dates(order_date, delivery)
            sale_qty = 0
            disuse_qty = 0
            buy_qty = 0
            for sd in sales_dates:
                info = sales_data.get((sd, item_cd), {})
                sale_qty += info.get("sale_qty", 0)
                disuse_qty += info.get("disuse_qty", 0)
                buy_qty += info.get("buy_qty", 0)

            # 오전(7~13시) 판매 비중
            morning_qty = 0
            total_hourly = 0
            for sd in sales_dates:
                hd = hourly_data.get((sd, item_cd), {})
                for h, q in hd.items():
                    total_hourly += q
                    if 7 <= h <= 13:
                        morning_qty += q
            morning_ratio = (morning_qty / total_hourly * 100) if total_hourly > 0 else None

            rec = {
                "store_id": store_id,
                "store_name": store_name,
                "order_date": order_date,
                "item_cd": item_cd,
                "item_nm": item_nm,
                "delivery": delivery,
                "order_source": order_source or "unknown",
                "order_qty": order_qty,
                "buy_qty": buy_qty,
                "sale_qty": sale_qty,
                "disuse_qty": disuse_qty,
                "morning_ratio": morning_ratio,
            }

            # buy_qty=0 → 별도 수집
            if buy_qty <= 0:
                rec["sell_rate"] = 0
                rec["sellout"] = 0
                all_buy0_records.append(rec)
                continue

            rec["sellout"] = 1 if (sale_qty >= buy_qty and disuse_qty == 0) else 0
            rec["sell_rate"] = round(sale_qty / buy_qty * 100, 1) if buy_qty > 0 else 0
            all_records.append(rec)

        # ── 이중발주 건에 판매 데이터 매핑 (전체 카테고리) ──
        # 이중발주 건용 sales 데이터 (카테고리 필터 없이 조회)
        dup_item_cds = set(d[1] for d in dupes)
        dup_sales = {}
        if dup_item_cds:
            placeholders = ",".join("?" * len(dup_item_cds))
            for row in conn.execute(f"""
                SELECT sales_date, item_cd, sale_qty, disuse_qty, buy_qty
                FROM daily_sales
                WHERE sales_date BETWEEN ? AND ?
                  AND item_cd IN ({placeholders})
            """, (ORDER_START, SALES_END, *dup_item_cds)).fetchall():
                dk = (row[0], row[1])
                dup_sales[dk] = {"sale_qty": row[2] or 0, "disuse_qty": row[3] or 0, "buy_qty": row[4] or 0}

        for (od, icd, inm, mcd, dtyp, src_a, qty_a, src_b, qty_b) in dupes:
            delivery = determine_delivery(inm)
            if delivery is None:
                delivery = dtyp if dtyp in ("1차", "2차") else None
            if delivery is None:
                # 비푸드는 배송차수 불명 → order_date+1 기본
                s_dates = [(datetime.strptime(od, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")]
            else:
                s_dates = get_sales_dates(od, delivery)
            # 푸드는 기존 sales_data, 비푸드는 dup_sales
            sq = sum((sales_data.get((sd, icd), {}) or dup_sales.get((sd, icd), {})).get("sale_qty", 0) for sd in s_dates)
            dq = sum((sales_data.get((sd, icd), {}) or dup_sales.get((sd, icd), {})).get("disuse_qty", 0) for sd in s_dates)
            bq = sum((sales_data.get((sd, icd), {}) or dup_sales.get((sd, icd), {})).get("buy_qty", 0) for sd in s_dates)
            all_duplicates.append({
                "store_id": store_id,
                "store_name": store_name,
                "order_date": od,
                "item_cd": icd,
                "item_nm": inm,
                "delivery": delivery,
                "auto_qty": qty_a,
                "other_source": src_b,
                "other_qty": qty_b,
                "combined_order": qty_a + qty_b,
                "buy_qty": bq,
                "sale_qty": sq,
                "disuse_qty": dq,
            })

        conn.close()

    # ── 7일 완판 횟수 계산 ──
    sellout_counts = defaultdict(int)
    for rec in all_records:
        if rec["sellout"]:
            sellout_counts[(rec["store_id"], rec["item_cd"])] += 1
    for rec in all_records:
        rec["sellout_7d"] = sellout_counts.get((rec["store_id"], rec["item_cd"]), 0)

    # ── 판정 ──
    for rec in all_records:
        if rec["disuse_qty"] > 0:
            rec["verdict"] = "과발주"
        elif rec["sellout"] and rec["sellout_7d"] >= SELLOUT_THRESHOLD_DAYS:
            rec["verdict"] = "수요초과가능성"
        elif rec["sale_qty"] < rec["buy_qty"] * 0.7:
            rec["verdict"] = "저조"
        else:
            rec["verdict"] = "적정"

    return all_records, all_buy0_records, all_duplicates, all_source_stats


# ────────────────────────────────────────
# 스타일 헬퍼
# ────────────────────────────────────────

def style_header(ws, row, headers, widths, fill):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 30


def apply_verdict_style(cell, verdict):
    if verdict == "과발주":
        cell.fill = RED_FILL
        cell.font = RED_FONT
    elif verdict == "적정":
        cell.fill = GREEN_FILL
        cell.font = GREEN_FONT
    elif verdict == "수요초과가능성":
        cell.fill = BLUE_FILL
        cell.font = BLUE_FONT
    elif verdict == "저조":
        cell.fill = YELLOW_FILL
        cell.font = YELLOW_FONT


def write_row(ws, row_idx, values, center_from=1):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER
        if col >= center_from:
            cell.alignment = Alignment(horizontal="center")
    return row_idx + 1


# ────────────────────────────────────────
# 시트 1: 상세 평가 (order_source 추가)
# ────────────────────────────────────────

def write_sheet1(wb, records):
    ws = wb.active
    ws.title = "상세 평가"
    headers = ["매장명", "발주일", "상품명", "발주출처", "차수", "발주량", "입고량",
               "판매량", "폐기량", "완판율(%)", "판정", "7일완판횟수"]
    widths = [18, 12, 35, 10, 7, 9, 9, 9, 9, 11, 14, 12]
    style_header(ws, 1, headers, widths, BLUE_HEADER)

    for i, rec in enumerate(records, 2):
        src_label = SOURCE_LABELS.get(rec["order_source"], rec["order_source"])
        vals = [rec["store_name"], rec["order_date"], rec["item_nm"],
                src_label, rec["delivery"], rec["order_qty"], rec["buy_qty"],
                rec["sale_qty"], rec["disuse_qty"], rec["sell_rate"],
                rec["verdict"], rec["sellout_7d"]]
        write_row(ws, i, vals, center_from=4)
        apply_verdict_style(ws.cell(row=i, column=11), rec["verdict"])

    ws.auto_filter.ref = f"A1:L{len(records)+1}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 2: 과발주 분석 (변경 없음)
# ────────────────────────────────────────

def write_sheet2(wb, records):
    ws = wb.create_sheet("과발주 분석")
    headers = ["매장명", "상품명", "폐기발생횟수", "총폐기량", "평균입고량",
               "평균판매량", "폐기율(%)", "원인추정"]
    widths = [18, 35, 13, 11, 11, 11, 11, 16]
    style_header(ws, 1, headers, widths, ORANGE_HEADER)

    waste_stats = defaultdict(lambda: {
        "store_name": "", "count": 0, "total_disuse": 0,
        "total_buy": 0, "total_sale": 0, "days": 0,
    })
    for rec in records:
        if rec["disuse_qty"] > 0:
            key = (rec["store_id"], rec["item_cd"])
            ws_item = waste_stats[key]
            ws_item["store_name"] = rec["store_name"]
            ws_item["item_nm"] = rec["item_nm"]
            ws_item["count"] += 1
            ws_item["total_disuse"] += rec["disuse_qty"]
            ws_item["total_buy"] += rec["buy_qty"]
            ws_item["total_sale"] += rec["sale_qty"]
            ws_item["days"] += 1

    sorted_waste = sorted(waste_stats.items(),
                          key=lambda x: x[1]["total_disuse"], reverse=True)

    row = 2
    for (sid, icd), st in sorted_waste:
        avg_buy = round(st["total_buy"] / st["days"], 1)
        avg_sale = round(st["total_sale"] / st["days"], 1)
        denom = st["total_sale"] + st["total_disuse"]
        waste_rate = round(st["total_disuse"] / denom * 100, 1) if denom > 0 else 0

        if waste_rate > 30:
            cause = "과발주"
        elif waste_rate >= 10:
            cause = "발주 조정 검토"
        else:
            cause = "허용 범위"

        vals = [st["store_name"], st["item_nm"], st["count"],
                st["total_disuse"], avg_buy, avg_sale, waste_rate, cause]
        write_row(ws, row, vals, center_from=3)

        cause_cell = ws.cell(row=row, column=8)
        if cause == "과발주":
            cause_cell.fill = RED_FILL
            cause_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")
        elif cause == "발주 조정 검토":
            cause_cell.fill = YELLOW_FILL
            cause_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
        else:
            cause_cell.fill = GREEN_FILL
            cause_cell.font = Font(name="맑은 고딕", bold=True, color="006100")
        row += 1

    row += 1
    summary = ws.cell(row=row, column=1,
                      value=f"과발주 상품 총 {len(sorted_waste)}건 (폐기율 30%+: "
                            f"{sum(1 for _,s in sorted_waste if s['total_disuse']/max(s['total_sale']+s['total_disuse'],1)*100>30)}건)")
    summary.font = Font(name="맑은 고딕", bold=True, size=11)

    ws.auto_filter.ref = f"A1:H{row-1}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 3: 매장별 종합 (auto/site+manual/전체 분리 + 신뢰도)
# ────────────────────────────────────────

def write_sheet3(wb, records, source_stats):
    ws = wb.create_sheet("매장별 종합")
    headers = ["매장명", "구분", "건수", "총입고량", "총판매량", "총폐기량",
               "폐기율(%)", "과발주", "적정", "저조", "수요초과",
               "auto비율(%)", "평가 신뢰도"]
    widths = [18, 14, 8, 11, 11, 11, 11, 9, 9, 9, 9, 12, 12]
    style_header(ws, 1, headers, widths, GREEN_HEADER)

    # 매장×구분별 집계
    # key: (store_id, category) where category = "auto" | "site_manual" | "전체"
    stats = defaultdict(lambda: {
        "name": "", "count": 0, "buy": 0, "sale": 0, "disuse": 0,
        "over": 0, "ok": 0, "low": 0, "excess": 0,
    })
    for rec in records:
        sid = rec["store_id"]
        src = rec["order_source"]

        # 전체
        k_all = (sid, "전체")
        stats[k_all]["name"] = rec["store_name"]
        stats[k_all]["count"] += 1
        stats[k_all]["buy"] += rec["buy_qty"]
        stats[k_all]["sale"] += rec["sale_qty"]
        stats[k_all]["disuse"] += rec["disuse_qty"]
        v = rec["verdict"]
        if v == "과발주": stats[k_all]["over"] += 1
        elif v == "적정": stats[k_all]["ok"] += 1
        elif v == "저조": stats[k_all]["low"] += 1
        elif v == "수요초과가능성": stats[k_all]["excess"] += 1

        # auto vs site+manual
        if src == "auto":
            cat = "자동(auto)"
        else:
            cat = "사이트+수동"
        k_cat = (sid, cat)
        stats[k_cat]["name"] = rec["store_name"]
        stats[k_cat]["count"] += 1
        stats[k_cat]["buy"] += rec["buy_qty"]
        stats[k_cat]["sale"] += rec["sale_qty"]
        stats[k_cat]["disuse"] += rec["disuse_qty"]
        if v == "과발주": stats[k_cat]["over"] += 1
        elif v == "적정": stats[k_cat]["ok"] += 1
        elif v == "저조": stats[k_cat]["low"] += 1
        elif v == "수요초과가능성": stats[k_cat]["excess"] += 1

    # 매장 목록 (정렬)
    store_ids = sorted(set(rec["store_id"] for rec in records))

    row = 2
    for sid in store_ids:
        ss = source_stats.get(sid, {})
        auto_cnt = ss.get("auto", 0)
        total_cnt = ss.get("total", 1)
        auto_pct = round(auto_cnt / total_cnt * 100, 1) if total_cnt > 0 else 0

        if auto_pct >= 80:
            reliability = "높음"
        elif auto_pct >= 50:
            reliability = "중간"
        else:
            reliability = "낮음"

        for cat in ["자동(auto)", "사이트+수동", "전체"]:
            k = (sid, cat)
            st = stats.get(k)
            if st is None or st["count"] == 0:
                continue
            wr = round(st["disuse"] / st["buy"] * 100, 1) if st["buy"] > 0 else 0

            # auto비율/신뢰도는 전체 행에만 표시
            show_auto = auto_pct if cat == "전체" else ""
            show_rel = reliability if cat == "전체" else ""

            vals = [st["name"], cat, st["count"],
                    st["buy"], st["sale"], st["disuse"], wr,
                    st["over"], st["ok"], st["low"], st["excess"],
                    show_auto, show_rel]
            write_row(ws, row, vals, center_from=2)

            # 전체 행 굵게 + 배경
            if cat == "전체":
                for c in range(1, 14):
                    cell = ws.cell(row=row, column=c)
                    cell.font = BOLD_FONT
                    cell.fill = LIGHT_GRAY_FILL
                    cell.border = THIN_BORDER

                # 신뢰도 색상
                rel_cell = ws.cell(row=row, column=13)
                if reliability == "높음":
                    rel_cell.fill = GREEN_FILL
                    rel_cell.font = Font(name="맑은 고딕", bold=True, color="006100")
                elif reliability == "중간":
                    rel_cell.fill = YELLOW_FILL
                    rel_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
                else:
                    rel_cell.fill = RED_FILL
                    rel_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")

            row += 1

        # 매장 간 빈 행 (구분선)
        row += 1

    # 범례
    ws.cell(row=row, column=1,
            value="auto비율: 전체 order_tracking 기준 | 높음(80%+) 중간(50~79%) 낮음(<50%)").font = \
        Font(name="맑은 고딕", italic=True, size=9, color="666666")

    ws.auto_filter.ref = f"A1:M{row-1}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 4: 수요초과 가능성 (변경 없음)
# ────────────────────────────────────────

def write_sheet4(wb, records):
    ws = wb.create_sheet("수요초과 가능성")
    headers = ["매장명", "발주일", "상품명", "차수", "입고량",
               "판매량", "완판율(%)", "7일완판횟수", "오전비중(7~13시, %)"]
    widths = [18, 12, 35, 7, 9, 9, 11, 12, 18]
    style_header(ws, 1, headers, widths, PURPLE_HEADER)

    excess_records = [r for r in records if r["verdict"] == "수요초과가능성"]
    excess_records.sort(key=lambda r: (r["store_name"], r["item_nm"], r["order_date"]))

    row = 2
    for rec in excess_records:
        morning_str = (f"{rec['morning_ratio']:.1f}%"
                       if rec["morning_ratio"] is not None else "데이터없음")
        vals = [rec["store_name"], rec["order_date"], rec["item_nm"],
                rec["delivery"], rec["buy_qty"], rec["sale_qty"],
                rec["sell_rate"], rec["sellout_7d"], morning_str]
        write_row(ws, row, vals, center_from=4)

        if rec["morning_ratio"] is not None and rec["morning_ratio"] >= 60:
            mc = ws.cell(row=row, column=9)
            mc.fill = YELLOW_FILL
            mc.font = YELLOW_FONT
        row += 1

    unique_items = len(set((r["store_id"], r["item_cd"]) for r in excess_records))
    row += 1
    ws.cell(row=row, column=1,
            value=f"수요초과 가능성 총 {len(excess_records)}건 "
                  f"({unique_items}개 상품)").font = \
        Font(name="맑은 고딕", bold=True, size=11)

    ws.auto_filter.ref = f"A1:I{max(row-1, 2)}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 5: 저조 상품 (변경 없음)
# ────────────────────────────────────────

def write_sheet5(wb, records):
    ws = wb.create_sheet("저조 상품")
    headers = ["매장명", "상품명", "저조일수", "총발주일수",
               "평균입고량", "평균판매량", "평균완판율(%)", "발주중단 검토"]
    widths = [18, 35, 10, 11, 11, 11, 13, 14]
    style_header(ws, 1, headers, widths, GRAY_HEADER)

    low_days = defaultdict(lambda: {
        "store_name": "", "item_nm": "", "low_count": 0,
        "total_days": 0, "total_buy": 0, "total_sale": 0,
        "low_sell_rates": [],
    })
    for rec in records:
        key = (rec["store_id"], rec["item_cd"])
        li = low_days[key]
        li["store_name"] = rec["store_name"]
        li["item_nm"] = rec["item_nm"]
        li["total_days"] += 1
        if rec["disuse_qty"] == 0 and rec["sell_rate"] < 70:
            li["low_count"] += 1
            li["total_buy"] += rec["buy_qty"]
            li["total_sale"] += rec["sale_qty"]
            li["low_sell_rates"].append(rec["sell_rate"])

    filtered = {k: v for k, v in low_days.items() if v["low_count"] >= 2}
    sorted_low = sorted(filtered.items(),
                        key=lambda x: x[1]["low_count"], reverse=True)

    row = 2
    for (sid, icd), li in sorted_low:
        avg_buy = round(li["total_buy"] / li["low_count"], 1) if li["low_count"] > 0 else 0
        avg_sale = round(li["total_sale"] / li["low_count"], 1) if li["low_count"] > 0 else 0
        avg_rate = (round(sum(li["low_sell_rates"]) / len(li["low_sell_rates"]), 1)
                    if li["low_sell_rates"] else 0)

        recommend = "중단 검토" if li["low_count"] >= 4 else ("주의 관찰" if li["low_count"] >= 2 else "유지")

        vals = [li["store_name"], li["item_nm"], li["low_count"],
                li["total_days"], avg_buy, avg_sale, avg_rate, recommend]
        write_row(ws, row, vals, center_from=3)

        rec_cell = ws.cell(row=row, column=8)
        if recommend == "중단 검토":
            rec_cell.fill = RED_FILL
            rec_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")
        elif recommend == "주의 관찰":
            rec_cell.fill = YELLOW_FILL
            rec_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
        row += 1

    row += 1
    ws.cell(row=row, column=1,
            value=f"저조 상품 총 {len(sorted_low)}건 "
                  f"(중단검토: {sum(1 for _,l in sorted_low if l['low_count']>=4)}건)").font = \
        Font(name="맑은 고딕", bold=True, size=11)

    ws.auto_filter.ref = f"A1:H{max(row-1, 2)}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 6: 이중발주 의심 (신규)
# ────────────────────────────────────────

def write_sheet6(wb, duplicates):
    ws = wb.create_sheet("이중발주 의심")
    headers = ["매장명", "발주일", "상품명", "차수",
               "auto발주량", "추가출처", "추가발주량", "합산발주량",
               "실입고량", "판매량", "폐기량", "비고"]
    widths = [18, 12, 35, 7, 11, 10, 11, 11, 10, 10, 10, 16]
    style_header(ws, 1, headers, widths, DARK_BLUE_HEADER)

    duplicates_sorted = sorted(duplicates,
                               key=lambda d: (d["store_name"], d["order_date"], d["item_nm"]))

    total_buy = 0
    total_sale = 0
    total_disuse = 0

    row = 2
    for dup in duplicates_sorted:
        other_label = SOURCE_LABELS.get(dup["other_source"], dup["other_source"])

        # 비고: 이중발주 패턴 판별
        if dup["auto_qty"] == dup["other_qty"]:
            note = "동일수량 중복"
        elif dup["other_qty"] > dup["auto_qty"] * 3:
            note = "대량 추가발주"
        else:
            note = "추가 보충"

        vals = [dup["store_name"], dup["order_date"], dup["item_nm"],
                dup["delivery"], dup["auto_qty"], other_label,
                dup["other_qty"], dup["combined_order"],
                dup["buy_qty"], dup["sale_qty"], dup["disuse_qty"], note]
        write_row(ws, row, vals, center_from=4)

        total_buy += dup["buy_qty"]
        total_sale += dup["sale_qty"]
        total_disuse += dup["disuse_qty"]

        # 비고 색상
        note_cell = ws.cell(row=row, column=12)
        if note == "동일수량 중복":
            note_cell.fill = RED_FILL
            note_cell.font = Font(name="맑은 고딕", bold=True, color="9C0006")
        elif note == "대량 추가발주":
            note_cell.fill = YELLOW_FILL
            note_cell.font = Font(name="맑은 고딕", bold=True, color="9C6500")
        else:
            note_cell.fill = BLUE_FILL
            note_cell.font = Font(name="맑은 고딕", bold=True, color="2F5496")
        row += 1

    # 요약
    row += 1
    ws.cell(row=row, column=1,
            value=f"이중발주 의심 총 {len(duplicates)}건").font = \
        Font(name="맑은 고딕", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1,
            value=f"합계: 입고={total_buy} / 판매={total_sale} / 폐기={total_disuse}").font = \
        Font(name="맑은 고딕", size=10)
    dup_waste = round(total_disuse / (total_sale + total_disuse) * 100, 1) if (total_sale + total_disuse) > 0 else 0
    row += 1
    ws.cell(row=row, column=1,
            value=f"이중발주 건의 폐기율: {dup_waste}%").font = \
        Font(name="맑은 고딕", size=10)

    ws.auto_filter.ref = f"A1:L{max(row-3, 2)}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 시트 7: 입고 미기록 (신규)
# ────────────────────────────────────────

def write_sheet7(wb, buy0_records, records):
    ws = wb.create_sheet("입고 미기록")
    headers = ["매장명", "발주일", "상품명", "발주출처", "차수",
               "발주량", "입고량(0)", "판매량", "비고"]
    widths = [18, 12, 35, 10, 7, 9, 10, 9, 18]
    style_header(ws, 1, headers, widths, TEAL_HEADER)

    # sale_qty > 0 인 케이스만
    sale_records = [r for r in buy0_records if r.get("sale_qty", 0) > 0]
    sale_records.sort(key=lambda r: (r["store_name"], r["order_date"], r["item_nm"]))

    row = 2
    sale_sum = 0
    for rec in sale_records:
        src_label = SOURCE_LABELS.get(rec["order_source"], rec["order_source"])
        note = "입고미기록+판매발생"

        vals = [rec["store_name"], rec["order_date"], rec["item_nm"],
                src_label, rec["delivery"], rec["order_qty"],
                0, rec["sale_qty"], note]
        write_row(ws, row, vals, center_from=4)

        sale_sum += rec["sale_qty"]

        # 판매량 셀 강조
        sale_cell = ws.cell(row=row, column=8)
        sale_cell.fill = YELLOW_FILL
        sale_cell.font = YELLOW_FONT
        row += 1

    # 전체 판매량 대비 비중 계산
    total_main_sale = sum(r["sale_qty"] for r in records)
    pct = round(sale_sum / (total_main_sale + sale_sum) * 100, 1) if (total_main_sale + sale_sum) > 0 else 0

    # 요약
    row += 1
    ws.cell(row=row, column=1,
            value=f"입고 미기록이지만 판매 발생: {len(sale_records)}건").font = \
        Font(name="맑은 고딕", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1,
            value=f"미기록 건 판매량 합계: {sale_sum}개").font = \
        Font(name="맑은 고딕", size=10)
    row += 1
    ws.cell(row=row, column=1,
            value=f"전체 판매량 대비 비중: {sale_sum} / {total_main_sale + sale_sum} = {pct}%").font = \
        Font(name="맑은 고딕", size=10)
    row += 1
    ws.cell(row=row, column=1,
            value="원인 추정: BGF 수집 시점에 입고 미확정 또는 데이터 타이밍 이슈").font = \
        Font(name="맑은 고딕", italic=True, size=9, color="666666")

    ws.auto_filter.ref = f"A1:I{max(row-4, 2)}"
    ws.freeze_panes = "A2"


# ────────────────────────────────────────
# 메인
# ────────────────────────────────────────

def main():
    print("푸드 발주 평가 v2 리포트 생성 시작...")
    print(f"기간: {ORDER_START} ~ {ORDER_END}")
    print(f"평가 기준: buy_qty(실입고) 기반, 폐기 중심 판정")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

    records, buy0_records, duplicates, source_stats = collect_data()
    print(f"총 {len(records)}건 수집 (buy_qty>0 필터 적용)")
    print(f"buy_qty=0 제외 건: {len(buy0_records)}건 (판매발생: {sum(1 for r in buy0_records if r.get('sale_qty',0)>0)}건)")
    print(f"이중발주 의심: {len(duplicates)}건")

    if not records:
        print("데이터 없음")
        return

    # 통계
    verdict_counts = defaultdict(int)
    for r in records:
        verdict_counts[r["verdict"]] += 1
    stores = set(r["store_name"] for r in records)
    print(f"매장 수: {len(stores)}")
    print(f"판정 분포:")
    for v in ["과발주", "적정", "저조", "수요초과가능성"]:
        print(f"  {v}: {verdict_counts[v]}건")

    # order_source 분포
    src_counts = defaultdict(int)
    for r in records:
        src_counts[r["order_source"]] += 1
    print(f"발주출처 분포:")
    for src in sorted(src_counts.keys()):
        label = SOURCE_LABELS.get(src, src)
        print(f"  {label}({src}): {src_counts[src]}건")

    total_buy = sum(r["buy_qty"] for r in records)
    total_disuse = sum(r["disuse_qty"] for r in records)
    print(f"전체 폐기율: {total_disuse}/{total_buy} = {total_disuse/total_buy*100:.1f}%")

    # 매장별 auto 비율
    print(f"매장별 auto 비율:")
    for sid, ss in sorted(source_stats.items()):
        if ss.get("total", 0) == 0:
            continue
        apct = round(ss.get("auto", 0) / ss["total"] * 100, 1)
        print(f"  {ss['store_name']}: {apct}%")

    wb = Workbook()
    write_sheet1(wb, records)
    write_sheet2(wb, records)
    write_sheet3(wb, records, source_stats)
    write_sheet4(wb, records)
    write_sheet5(wb, records)
    write_sheet6(wb, duplicates)
    write_sheet7(wb, buy0_records, records)

    wb.save(OUTPUT_PATH)
    print(f"\n리포트 저장 완료: {OUTPUT_PATH}")
    print(f"시트 구성: {wb.sheetnames}")


if __name__ == "__main__":
    main()
